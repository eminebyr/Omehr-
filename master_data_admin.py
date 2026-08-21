"""Web panelinden ana Excel input verilerini güvenli biçimde yönetir.

Ana veri kaynağı değişmez: input/OMEHR_AI_NORM_TRANSFER_INPUT.xlsx.
Bu servis panelde yapılan değişiklikleri yedek alarak aynı çalışma kitabına
atomik şekilde yazar. Böylece mevcut motorlar, raporlar ve Outlook akışları
kod değişmeden güncel veriyi okumaya devam eder.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
import csv
import os
import shutil
import tempfile
import time
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

SHEETS = {
    "Mağazalar": "Dim_Magaza",
    "Unvanlar": "Dim_Unvan",
    "Norm Kadro": "Fact_Norm",
    "Personel": "Fact_Mevcut",
}

EDITABLE_COLUMNS = {
    "Dim_Magaza": ["MağazaID", "Mağaza", "Bölge Sorumlusu"],
    "Dim_Unvan": ["UnvanID", "Unvan"],
    "Fact_Norm": ["MağazaID", "UnvanID", "Norm Kadro"],
    "Fact_Mevcut": [
        "MağazaID", "UnvanID", "Departman", "Açıklama", "İsim Soyisim", "İşe Giriş",
        "İşten Çıkış", "Çıkış Kodu", "CikisNedeniID", "Çıkış Nedeni",
    ],
}

YELLOW = "FFF2CC"  # formül/otomatik
RED = "F4CCCC"     # manuel veri girişi


def read_tables(input_path: Path) -> dict[str, pd.DataFrame]:
    result: dict[str, pd.DataFrame] = {}
    for sheet, cols in EDITABLE_COLUMNS.items():
        df = pd.read_excel(input_path, sheet_name=sheet, dtype=object)
        for col in cols:
            if col not in df.columns:
                df[col] = None
        result[sheet] = df[cols].copy()
    return result


def _clean_text(value: object) -> object:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    text = str(value).strip()
    return text if text else None


def _clean_frame(sheet: str, df: pd.DataFrame) -> pd.DataFrame:
    cols = EDITABLE_COLUMNS[sheet]
    out = df.copy()
    for col in cols:
        if col not in out.columns:
            out[col] = None
    out = out[cols]
    # Tamamen boş satırları kaldır.
    out = out.dropna(how="all")
    for col in cols:
        if col not in {"Norm Kadro", "İşe Giriş", "İşten Çıkış"}:
            out[col] = out[col].map(_clean_text)
    if sheet == "Fact_Norm":
        out["Norm Kadro"] = pd.to_numeric(out["Norm Kadro"], errors="coerce")
    if sheet == "Fact_Mevcut":
        for col in ("İşe Giriş", "İşten Çıkış"):
            out[col] = pd.to_datetime(out[col], errors="coerce")
    return out.reset_index(drop=True)


def validate_tables(tables: dict[str, pd.DataFrame]) -> list[str]:
    errors: list[str] = []
    dm = _clean_frame("Dim_Magaza", tables["Dim_Magaza"])
    du = _clean_frame("Dim_Unvan", tables["Dim_Unvan"])
    fn = _clean_frame("Fact_Norm", tables["Fact_Norm"])
    fm = _clean_frame("Fact_Mevcut", tables["Fact_Mevcut"])

    for label, df, key in (
        ("Mağaza", dm, "MağazaID"), ("Unvan", du, "UnvanID")
    ):
        if df[key].isna().any():
            errors.append(f"{label} tablosunda boş {key} var.")
        dup = df[df[key].duplicated(keep=False) & df[key].notna()][key].unique().tolist()
        if dup:
            errors.append(f"{label} tablosunda mükerrer {key}: {', '.join(map(str, dup[:10]))}")

    valid_store = set(dm["MağazaID"].dropna().astype(str))
    valid_title = set(du["UnvanID"].dropna().astype(str))
    for label, df in (("Norm Kadro", fn), ("Personel", fm)):
        bad_store = sorted(set(df["MağazaID"].dropna().astype(str)) - valid_store)
        bad_title = sorted(set(df["UnvanID"].dropna().astype(str)) - valid_title)
        if bad_store:
            errors.append(f"{label}: Dim_Magaza'da olmayan MağazaID: {', '.join(bad_store[:10])}")
        if bad_title:
            errors.append(f"{label}: Dim_Unvan'da olmayan UnvanID: {', '.join(bad_title[:10])}")

    if fn[["MağazaID", "UnvanID"]].duplicated().any():
        errors.append("Fact_Norm içinde aynı MağazaID + UnvanID birden fazla kez bulunuyor.")
    if (fn["Norm Kadro"].fillna(-1) < 0).any():
        errors.append("Norm Kadro boş veya negatif olamaz.")
    if fm["İsim Soyisim"].isna().any():
        errors.append("Fact_Mevcut içinde boş İsim Soyisim var.")
    dup_names = fm[fm["İsim Soyisim"].duplicated(keep=False) & fm["İsim Soyisim"].notna()]["İsim Soyisim"].unique().tolist()
    if dup_names:
        errors.append("İsim Soyisim benzersiz olmalı. Mükerrer: " + ", ".join(map(str, dup_names[:10])))
    return errors


def _clear_data(ws, start_row: int = 2) -> None:
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def _write_dim_magaza(ws, df: pd.DataFrame) -> None:
    _clear_data(ws)
    for r, row in enumerate(df.itertuples(index=False), 2):
        for c, value in enumerate(row, 1):
            ws.cell(r, c, value)
            ws.cell(r, c).fill = PatternFill("solid", fgColor=RED)


def _write_dim_unvan(ws, df: pd.DataFrame) -> None:
    _write_dim_magaza(ws, df)


def _write_fact_norm(ws, df: pd.DataFrame, dim_magaza: pd.DataFrame, dim_unvan: pd.DataFrame) -> None:
    # Fact_Norm yalniz magaza-unvan normunu tutar; personel aciklamasi burada tutulmaz.
    # DUZELTME (KRITIK -- daha once Fact_Mevcut'ta bulunup duzeltilen AYNI hata
    # sinifi, burada fark edilmemisti): Magaza/Bolge Sorumlusu/Unvan artik
    # Excel VLOOKUP formulu DEGIL, Python'da ANINDA cozulmus gercek degerler
    # olarak yaziliyor -- hicbir LibreOffice yeniden hesaplama adimi bu yazma
    # yolunda calismadigi icin, formul olarak yazilan bu sutunlar pandas/
    # openpyxl ile okundugunda HER ZAMAN bos/NaN geliyordu (bizzat kanitlandi).
    _clear_data(ws)
    if ws.max_column >= 8:
        ws.delete_cols(8, ws.max_column - 7)
    magaza_map = dict(zip(dim_magaza["MağazaID"].astype(str), dim_magaza["Mağaza"].astype(str)))
    bolge_map = dict(zip(dim_magaza["MağazaID"].astype(str), dim_magaza.get("Bölge Sorumlusu", pd.Series(dtype=str)).astype(str)))
    unvan_map = dict(zip(dim_unvan["UnvanID"].astype(str), dim_unvan["Unvan"].astype(str)))
    for r, row in enumerate(df.itertuples(index=False), 2):
        store_id, title_id, norm = row
        sid = "" if pd.isna(store_id) else str(store_id)
        tid = "" if pd.isna(title_id) else str(title_id)
        ws.cell(r, 1, store_id)
        ws.cell(r, 2, magaza_map.get(sid, ""))
        ws.cell(r, 3, bolge_map.get(sid, ""))
        ws.cell(r, 4, title_id)
        ws.cell(r, 5, unvan_map.get(tid, ""))
        ws.cell(r, 6, None if pd.isna(norm) else int(norm))
        for c in (1, 4, 6): ws.cell(r, c).fill = PatternFill("solid", fgColor=RED)
        for c in (2, 3, 5): ws.cell(r, c).fill = PatternFill("solid", fgColor=YELLOW)

def _write_fact_mevcut(ws, df: pd.DataFrame, dim_magaza: pd.DataFrame, dim_unvan: pd.DataFrame, dim_cikis: pd.DataFrame) -> None:
    _clear_data(ws)
    magaza_map = dict(zip(dim_magaza["MağazaID"].astype(str), dim_magaza["Mağaza"].astype(str)))
    bolge_map = dict(zip(dim_magaza["MağazaID"].astype(str), dim_magaza.get("Bölge Sorumlusu", pd.Series(dtype=str)).astype(str)))
    unvan_map = dict(zip(dim_unvan["UnvanID"].astype(str), dim_unvan["Unvan"].astype(str)))
    cikis_map = {}
    if dim_cikis is not None and not dim_cikis.empty and {"CikisNedeniID", "CikisNedeni"}.issubset(dim_cikis.columns):
        cikis_map = dict(zip(dim_cikis["CikisNedeniID"], dim_cikis["CikisNedeni"].astype(str)))

    # H sutunu personel aciklamasidir; isim ve devamindaki alanlar bir sutun saga kayar.
    # DUZELTME: Magaza/Unvan/Cikis Nedeni/Durum artik Excel VLOOKUP formulu
    # DEGIL, Python'da ANINDA cozulmus gercek degerler olarak yaziliyor --
    # pandas/openpyxl ile okunan HER rapor (PDF/Excel/panel) LibreOffice'in
    # formulu hesaplamasini beklemeden dogru veriyi gorur. Kok neden: bu
    # yazma yolunda hicbir yerde formul yeniden hesaplama adimi calismiyordu,
    # bu da bir kisi isten cikartildiginda satirinin load() ciktisindan
    # TAMAMEN kaybolmasina yol aciyordu (Magaza/Unvan bos kaldigi icin).
    headers=["MağazaID","Mağaza","Bölge Sorumlusu","UnvanID","Unvan","Departman","Norm fazlası Norm eksiği","Açıklama","İsim Soyisim","İşe Giriş","İşten Çıkış","Çıkış Kodu","CikisNedeniID","Çıkış Nedeni","Durum","Kıdem (Gün)","Kıdem (Yıl)"]
    for c,h in enumerate(headers,1): ws.cell(1,c,h)
    for r, row in enumerate(df.itertuples(index=False), 2):
        store_id, title_id, dept, explanation, name, hire, exit_date, exit_code, exit_reason_id, exit_reason_text = row
        sid = "" if pd.isna(store_id) else str(store_id)
        tid = "" if pd.isna(title_id) else str(title_id)
        rid = None if pd.isna(exit_reason_id) else exit_reason_id
        ws.cell(r, 1, store_id)
        ws.cell(r, 2, magaza_map.get(sid, ""))
        ws.cell(r, 3, bolge_map.get(sid, ""))
        ws.cell(r, 4, title_id)
        ws.cell(r, 5, unvan_map.get(tid, ""))
        ws.cell(r, 6, dept)
        ws.cell(r, 7, "=IF(OR(A{0}=\"\",F{0}=\"\",I{0}=\"\",K{0}<>\"\"),\"\",IF(COUNTIFS($A$2:A{0},A{0},$F$2:F{0},F{0},$K$2:K{0},\"\")<=SUMIFS(Fact_Norm!$F:$F,Fact_Norm!$A:$A,A{0},Fact_Norm!$E:$E,F{0}),\"Norma Uygun\",\"Norm Fazlası\"))".format(r))
        ws.cell(r, 8, explanation)
        ws.cell(r, 9, name)
        ws.cell(r, 10, None if pd.isna(hire) else hire.to_pydatetime())
        ws.cell(r, 11, None if pd.isna(exit_date) else exit_date.to_pydatetime())
        ws.cell(r, 12, exit_code)
        ws.cell(r, 13, exit_reason_id)
        # DUZELTME: ozel/serbest metin cikis nedeni ONCELIKLIDIR -- yalniz
        # bos ise ID tabanli kanonik metne dusulur (FAST V9'un istedigi
        # "AYSE Istifa - Ucret, MEHMET Isveren feshi - Performans" gibi
        # kisiye ozel aciklamalarin KORUNMASI icin).
        _ozel_metin = "" if pd.isna(exit_reason_text) else str(exit_reason_text).strip()
        ws.cell(r, 14, _ozel_metin or (cikis_map.get(rid, "") if rid is not None else ""))
        satir_pasif = not pd.isna(exit_date)
        ws.cell(r, 15, ("" if pd.isna(name) or str(name).strip()=="" else ("Pasif" if satir_pasif else "Aktif")))
        ws.cell(r, 16, "=IF(J{0}=\"\",\"\",IF(K{0}<>\"\",K{0}-J{0},TODAY()-J{0}))".format(r))
        ws.cell(r, 17, "=IF(P{0}=\"\",\"\",ROUND(P{0}/365,1))".format(r))
        for c in (1,4,6,8,9,10,11,12,13): ws.cell(r,c).fill = PatternFill("solid", fgColor=RED)
        for c in (2,3,5,7,14,15,16,17): ws.cell(r,c).fill = PatternFill("solid", fgColor=YELLOW)
        ws.cell(r, 10).number_format = "DD.MM.YYYY"
        ws.cell(r, 11).number_format = "DD.MM.YYYY"

def _audit(root: Path, username: str, counts: dict[str, int]) -> None:
    path = root / "logs" / "ana_veri_degisiklik_gecmisi.csv"
    path.parent.mkdir(parents=True, exist_ok=True)
    new = not path.exists()
    with path.open("a", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, delimiter=";")
        if new: writer.writerow(["TarihSaat", "Kullanıcı", "Dim_Magaza", "Dim_Unvan", "Fact_Norm", "Fact_Mevcut"])
        writer.writerow([datetime.now().isoformat(timespec="seconds"), username, counts["Dim_Magaza"], counts["Dim_Unvan"], counts["Fact_Norm"], counts["Fact_Mevcut"]])


def _invalidate_generated_reports(root: Path) -> None:
    """Ana veri değişince eski güncel raporları kaldır; arşive dokunma."""
    output = root / 'output'
    for name in (
        'OMEHR_Yonetici_Raporu.pdf', 'OMEHR_Yonetici_Raporu.xlsx',
        'OMEHR_Kutucuklu_Yonetici_Raporu.xlsx', 'OMEHR_Executive_Data.xlsx',
    ):
        try:
            (output / name).unlink(missing_ok=True)
        except Exception:
            pass
    region_dir = output / 'Bolge_Raporlari'
    if region_dir.is_dir():
        for fp in region_dir.glob('*'):
            if fp.is_file() and fp.suffix.lower() in {'.pdf', '.xlsx'}:
                try:
                    fp.unlink()
                except Exception:
                    pass


def _replace_excel_with_retry(temp: Path, input_path: Path, *, attempts: int = 5, delay_seconds: float = 0.8) -> None:
    """Geçici Excel dosyasını ana input üzerine güvenli biçimde geçir.

    DÜZELTME (FAST V13): Windows/Excel dosya kilidinde (kullanıcı ana
    dosyayı Excel'de açık bırakmışsa) kısa süreli kilitleri otomatik
    tekrar dener. Kilit kalıcıysa ana dosyaya DOKUNMADAN, ham
    "[WinError 5]" yerine anlaşılır bir hata mesajı verir.
    """
    last_error: OSError | None = None
    for attempt in range(1, max(1, attempts) + 1):
        try:
            os.replace(temp, input_path)
            return
        except OSError as exc:
            last_error = exc
            winerror = getattr(exc, "winerror", None)
            retryable = isinstance(exc, PermissionError) or winerror in {5, 32}
            if not retryable or attempt >= attempts:
                break
            time.sleep(delay_seconds)

    if last_error is not None:
        winerror = getattr(last_error, "winerror", None)
        if isinstance(last_error, PermissionError) or winerror in {5, 32}:
            raise PermissionError(
                "Ana Excel dosyası şu anda başka bir program tarafından kullanılıyor veya Windows yazma izni vermiyor. "
                "OMEHR_AI_NORM_TRANSFER_INPUT.xlsx dosyasını Excel/LibreOffice'te kapatın ve işlemi tekrar deneyin. "
                "Kayıt uygulanmadı; mevcut ana Excel dosyası korunmuştur."
            ) from last_error
        raise last_error


def _save_tables_unlocked(root: Path, input_path: Path, tables: dict[str, pd.DataFrame], username: str) -> Path:
    cleaned = {name: _clean_frame(name, tables[name]) for name in EDITABLE_COLUMNS}
    errors = validate_tables(cleaned)
    if errors:
        raise ValueError("\n".join(errors))

    # DÜZELTME (SaaS kota uygulaması): sube_kotasi tabloda tanımlıydı ama
    # hiçbir yerde kontrol edilmiyordu. Mağaza sayısı kotayı aşıyorsa
    # kayıt burada, herhangi bir dosya değişmeden ÖNCE reddedilir.
    try:
        from services.tenant_context import current_tenant_id
        from services.tenant_registry import check_quota
        tenant_id = current_tenant_id()
        yeni_magaza_sayisi = len(cleaned["Dim_Magaza"])
        uygun, mesaj = check_quota(tenant_id, "sube", yeni_magaza_sayisi)
        if not uygun:
            raise ValueError(mesaj)
    except ImportError:
        pass

    backups = root / "backups"
    backups.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = backups / f"BASDAS_AI_NORM_TRANSFER_INPUT_PANEL_{stamp}.xlsx"
    shutil.copy2(input_path, backup)

    wb = load_workbook(input_path)
    _write_dim_magaza(wb["Dim_Magaza"], cleaned["Dim_Magaza"])
    _write_dim_unvan(wb["Dim_Unvan"], cleaned["Dim_Unvan"])
    _write_fact_norm(wb["Fact_Norm"], cleaned["Fact_Norm"], cleaned["Dim_Magaza"], cleaned["Dim_Unvan"])
    try:
        _dim_cikis_ham = pd.read_excel(input_path, sheet_name="Dim_CikisNedeni", dtype=object)
    except Exception:
        _dim_cikis_ham = pd.DataFrame()
    _write_fact_mevcut(wb["Fact_Mevcut"], cleaned["Fact_Mevcut"], cleaned["Dim_Magaza"], cleaned["Dim_Unvan"], _dim_cikis_ham)
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    fd, temp_name = tempfile.mkstemp(prefix="basdas_panel_", suffix=".xlsx", dir=str(input_path.parent))
    os.close(fd)
    temp = Path(temp_name)
    try:
        wb.save(temp)
        wb.close()
        # Dosya tekrar açılabiliyor mu kontrol et.
        check = load_workbook(temp, read_only=True, data_only=False)
        check.close()
        _replace_excel_with_retry(temp, input_path)
    finally:
        try:
            wb.close()
        except Exception:
            pass
        temp.unlink(missing_ok=True)

    _audit(root, username, {k: len(v) for k, v in cleaned.items()})
    _invalidate_generated_reports(root)
    return backup


def save_tables(root: Path, input_path: Path, tables: dict[str, pd.DataFrame], username: str, *, acquire_lock: bool = True) -> Path:
    """DÜZELTME (FAST V15 — 3 PC ortak Excel): birden fazla bilgisayar aynı
    ağ Excel dosyasını kullanıyorsa, bu kilit eşzamanlı iki yazmanın
    birbirinin üzerine yazmasını (stale overwrite) önler."""
    if not acquire_lock:
        return _save_tables_unlocked(root, input_path, tables, username)
    from services.multi_pc_excel import excel_transaction_lock
    with excel_transaction_lock(input_path, user=username):
        return _save_tables_unlocked(root, input_path, tables, username)
