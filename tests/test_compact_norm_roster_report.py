from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.excel_report import build_compact_norm_roster_excel


def _frames():
    st = pd.DataFrame([
        {'Bölge Sorumlusu': 'DENEME BÖLGESİ', 'Mağaza': name}
        for name in ('MAĞAZA A', 'MAĞAZA B', 'MAĞAZA C')
    ])
    norm = pd.DataFrame([
        {'Mağaza': 'MAĞAZA A', 'Unvan': 'KASİYER', 'Norm Kadro': 2},
        {'Mağaza': 'MAĞAZA B', 'Unvan': 'REYON GÖREVLİSİ', 'Norm Kadro': 1},
        {'Mağaza': 'MAĞAZA C', 'Unvan': 'MANAV', 'Norm Kadro': 0},
    ])
    staff = pd.DataFrame([
        {'Mağaza': 'MAĞAZA A', 'Departman': 'KASİYER', 'Unvan': 'KASİYER', 'İsim Soyisim': 'AYŞE TEST', 'İşe Giriş': '2026-01-01', 'Açıklama': ''},
        {'Mağaza': 'MAĞAZA B', 'Departman': 'REYON GÖREVLİSİ', 'Unvan': 'REYON GÖREVLİSİ', 'İsim Soyisim': 'ALİ TEST', 'İşe Giriş': '2026-01-01', 'Açıklama': 'Raporlu'},
        {'Mağaza': 'MAĞAZA C', 'Departman': 'MANAV', 'Unvan': 'MANAV', 'İsim Soyisim': 'ECE TEST', 'İşe Giriş': '2026-08-01', 'Açıklama': ''},
    ])
    return st, norm, staff


def _all_values(ws):
    return [cell.value for row in ws.iter_rows() for cell in row if cell.value is not None]


def test_compact_report_uses_three_store_layout_and_live_values(tmp_path: Path):
    st, norm, staff = _frames()
    first = build_compact_norm_roster_excel(st, norm, staff, output_path=tmp_path/'first.xlsx')
    ws = load_workbook(first)['DENEME BÖLGESİ']
    assert ws['A5'].value == 'MAĞAZA A'
    assert ws['D5'].value == 'MAĞAZA B'
    assert ws['G5'].value == 'MAĞAZA C'
    values = _all_values(ws)
    assert 'BOŞ POZİSYON' in values and 'EKSİK PERSONEL: 1' in values
    assert 'FAZLA PERSONEL: 1' in values

    staff = pd.concat([staff, pd.DataFrame([{
        'Mağaza': 'MAĞAZA A', 'Departman': 'KASİYER', 'Unvan': 'KASİYER',
        'İsim Soyisim': 'YENİ PERSONEL', 'İşe Giriş': '2026-08-29', 'Açıklama': '',
    }])], ignore_index=True)
    second = build_compact_norm_roster_excel(st, norm, staff, output_path=tmp_path/'second.xlsx')
    values = _all_values(load_workbook(second)['DENEME BÖLGESİ'])
    assert 'YENİ PERSONEL' in values
    assert 'EKSİK PERSONEL: 0' in values


def test_compact_report_includes_norm_eksik_fazla_pivot_sheets(tmp_path: Path):
    st, norm, staff = _frames()
    out = build_compact_norm_roster_excel(st, norm, staff, output_path=tmp_path / 'pivot.xlsx')
    wb = load_workbook(out)
    assert {'NORM KADRO', 'EKSİK', 'FAZLA'}.issubset(set(wb.sheetnames))

    norm_ws = wb['NORM KADRO']
    header = [c.value for c in norm_ws[1]]
    assert 'KASİYER' in header and 'REYON GÖREVLİSİ' in header and 'MANAV' in header and header[-1] == 'TOPLAM'

    def _row(ws, store_name):
        for row in ws.iter_rows(min_row=2):
            if row[1].value == store_name:
                return {ws.cell(1, c.column).value: c.value for c in row}
        raise AssertionError(f"{store_name} bulunamadı")

    # MAĞAZA A: KASİYER norm=2, mevcut=1 (test fixture'ında) -> eksik=1
    eksik_a = _row(wb['EKSİK'], 'MAĞAZA A')
    assert eksik_a['KASİYER'] == 1
    assert eksik_a['TOPLAM'] == 1

    # MAĞAZA C: MANAV norm=0, mevcut=1 -> fazla=1
    fazla_c = _row(wb['FAZLA'], 'MAĞAZA C')
    assert fazla_c['MANAV'] == 1
    assert fazla_c['TOPLAM'] == 1

    # Alt satırda TOPLAM satırı, her sütunun toplamını içermeli
    total_row = [c.value for c in list(wb['EKSİK'].iter_rows())[-1]]
    assert total_row[1] == 'TOPLAM'
    assert total_row[-1] == sum(v for v in total_row[2:-1] if isinstance(v, int))


def test_compact_report_pivot_sheets_use_state_output_when_tt_provided(tmp_path: Path):
    """DÜZELTME (tutarlılık, 29 Ağustos 2026) regresyon testi: tt (state()
    çıktısı) verildiğinde EKSİK/FAZLA sayfaları KENDİ ham hesaplaması
    yerine tt'den okumalı — bu, KASİYER gibi config_norm_rules.json'da
    hiç tanımlı olmayan bir aile için bile otomatik dengeleme (Uzman
    Kasiyer dahil, Kasiyer Yardımcısı ile mahsup) yapılmasını sağlar.
    tt VERİLMEDEN aynı senaryo çalıştırılsaydı bu dengeleme HİÇ
    uygulanmazdı (ham normx/staffx hesaplaması aile kuralı bilmez)."""
    from src.state_engine import state

    norm = pd.DataFrame([
        {'MağazaID': '1', 'Mağaza': 'TEST MAĞAZA', 'Bölge Sorumlusu': 'TEST BÖLGE', 'Unvan': 'KASİYER', 'Norm Kadro': 2},
        {'MağazaID': '1', 'Mağaza': 'TEST MAĞAZA', 'Bölge Sorumlusu': 'TEST BÖLGE', 'Unvan': 'KASİYER YARDIMCISI', 'Norm Kadro': 1},
    ])
    staff = pd.DataFrame([
        {'MağazaID': '1', 'Mağaza': 'TEST MAĞAZA', 'Bölge Sorumlusu': 'TEST BÖLGE',
         'İsim Soyisim': 'KİŞİ 1', 'Unvan': 'UZMAN KASİYER', 'Departman': 'KASİYER'},
        {'MağazaID': '1', 'Mağaza': 'TEST MAĞAZA', 'Bölge Sorumlusu': 'TEST BÖLGE',
         'İsim Soyisim': 'KİŞİ 2', 'Unvan': 'KASİYER YARDIMCISI', 'Departman': 'KASİYER YARDIMCISI'},
        {'MağazaID': '1', 'Mağaza': 'TEST MAĞAZA', 'Bölge Sorumlusu': 'TEST BÖLGE',
         'İsim Soyisim': 'KİŞİ 3', 'Unvan': 'KASİYER YARDIMCISI', 'Departman': 'KASİYER YARDIMCISI'},
    ])
    st, tt = state(norm, staff, {})

    out_ile_tt = build_compact_norm_roster_excel(st, norm, staff, output_path=tmp_path / 'ile_tt.xlsx', tt=tt)
    wb = load_workbook(out_ile_tt)

    def _deger(ws, unvan_kolon_adi):
        header = [c.value for c in ws[1]]
        col_idx = header.index(unvan_kolon_adi) + 1
        for row in ws.iter_rows(min_row=2):
            if row[1].value and row[1].value != 'TOPLAM':
                v = row[col_idx - 1].value
                return v if v else 0
        return None

    eksik_ws = wb['EKSİK']
    fazla_ws = wb['FAZLA']
    kasiyer_eksik = _deger(eksik_ws, 'KASİYER')
    yardimci_fazla = _deger(fazla_ws, 'KASİYER YARDIMCISI')
    assert (kasiyer_eksik or 0) == 0, "tt verildiğinde Kasiyer ailesi (Uzman Kasiyer dahil) otomatik dengelenmeli"
    assert (yardimci_fazla or 0) == 0, "tt verildiğinde Yardımcı fazlası otomatik dengelenmeli"


def test_compact_report_applies_status_colors_and_comments(tmp_path: Path):
    st, norm, staff = _frames()
    out = build_compact_norm_roster_excel(st, norm, staff, output_path=tmp_path/'colors.xlsx')
    ws = load_workbook(out)['DENEME BÖLGESİ']
    cells = {cell.value: cell for row in ws.iter_rows() for cell in row if cell.value}
    assert cells['BOŞ POZİSYON'].fill.fgColor.rgb.endswith('44B3E1')
    assert cells['ECE TEST'].fill.fgColor.rgb.endswith('92D050')
    assert cells['ALİ TEST'].fill.fgColor.rgb.endswith('E4DFEC')
    assert cells['ALİ TEST'].comment and 'raporlu' in cells['ALİ TEST'].comment.text.lower()
