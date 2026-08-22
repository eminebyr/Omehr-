from __future__ import annotations

"""mail_router.py entegrasyonu — report_mail_engine.py üzerinden GERÇEK
gönderim akışı (regresyon testi).

Bu, önceki "hiçbir gerçek akışa bağlı değil" durumunu düzeltir. Gerçek
send_reports_via_outlook() akışı üzerinden, abonelik sütunu (Norm_Genel)
"Hayır" olan bir alıcının GERÇEKTEN mail almadığı, sütun hiç yoksa
davranışın DEĞİŞMEDİĞİ doğrulanır.
"""

import shutil

import openpyxl
import pandas as pd
import pytest


def _kutuphane_pdf_var_mi():
    try:
        import pypdf  # noqa: F401
        return True
    except ImportError:
        return False


@pytest.mark.skipif(not _kutuphane_pdf_var_mi(), reason="pypdf gerekli")
def test_subscription_opt_out_genuinely_blocks_send(tmp_path, monkeypatch):
    monkeypatch.setenv("OMEHR_RUNTIME_ROOT", str(tmp_path))
    monkeypatch.setenv("OMEHR_MAIL_DRY_RUN", "1")
    (tmp_path / "input").mkdir()
    (tmp_path / "output").mkdir()

    import importlib
    import report_mail_engine as rme
    importlib.reload(rme)

    from pypdf import PdfWriter
    w = PdfWriter()
    w.add_blank_page(width=200, height=200)
    with open(rme._output_dir() / "OMEHR_Yonetici_Raporu.pdf", "wb") as f:
        w.write(f)

    df_mail = pd.DataFrame([
        {"Aktif": "evet", "E-posta": "abone@test.com", "Bölge": "TUMU", "Norm_Genel": "Evet"},
        {"Aktif": "evet", "E-posta": "abone_degil@test.com", "Bölge": "TUMU", "Norm_Genel": "Hayır"},
    ])
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Mail_Listesi")
    ws.append(list(df_mail.columns))
    for _, row in df_mail.iterrows():
        ws.append(list(row))
    hedef = rme._input_file()
    wb.save(hedef)

    rme.send_reports_via_outlook(hedef, display_only=False)
    import json
    gonderilenler = [r["to"] for r in json.loads(open(rme._log_file()).read())]
    assert "abone@test.com" in gonderilenler
    assert "abone_degil@test.com" not in gonderilenler, (
        "REGRESYON: mail_router.py entegrasyonu bozulmuş — abonelikten "
        "çıkan kişi hâlâ mail alıyor."
    )


@pytest.mark.skipif(not _kutuphane_pdf_var_mi(), reason="pypdf gerekli")
def test_no_subscription_column_preserves_old_behavior(tmp_path, monkeypatch):
    monkeypatch.setenv("OMEHR_RUNTIME_ROOT", str(tmp_path))
    monkeypatch.setenv("OMEHR_MAIL_DRY_RUN", "1")
    (tmp_path / "input").mkdir()
    (tmp_path / "output").mkdir()

    import importlib
    import report_mail_engine as rme
    importlib.reload(rme)

    from pypdf import PdfWriter
    w = PdfWriter()
    w.add_blank_page(width=200, height=200)
    with open(rme._output_dir() / "OMEHR_Yonetici_Raporu.pdf", "wb") as f:
        w.write(f)

    df_mail = pd.DataFrame([
        {"Aktif": "evet", "E-posta": "kisi1@test.com", "Bölge": "TUMU"},
        {"Aktif": "evet", "E-posta": "kisi2@test.com", "Bölge": "TUMU"},
    ])
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Mail_Listesi")
    ws.append(list(df_mail.columns))
    for _, row in df_mail.iterrows():
        ws.append(list(row))
    hedef = rme._input_file()
    wb.save(hedef)

    rme.send_reports_via_outlook(hedef, display_only=False)
    import json
    gonderilenler = [r["to"] for r in json.loads(open(rme._log_file()).read())]
    assert "kisi1@test.com" in gonderilenler
    assert "kisi2@test.com" in gonderilenler
