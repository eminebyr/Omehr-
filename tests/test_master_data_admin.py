from pathlib import Path
import shutil

from openpyxl import load_workbook

from services.master_data_admin import read_tables, save_tables, validate_tables


def test_panel_master_data_roundtrip(tmp_path):
    root = tmp_path / "pkg"
    (root / "input").mkdir(parents=True)
    source = Path(__file__).resolve().parents[1] / "input" / "BASDAS_AI_NORM_TRANSFER_INPUT.xlsx"
    target = root / "input" / source.name
    shutil.copy2(source, target)

    tables = read_tables(target)
    assert not validate_tables(tables)
    backup = save_tables(root, target, tables, "PYTEST")
    assert backup.exists()

    wb = load_workbook(target, data_only=False)
    # DÜZELTME: Fact_Norm'da Mağaza/Bölge Sorumlusu/Unvan artık HESAPLANMAMIŞ
    # bir VLOOKUP formülü değil, Python'da anında çözülmüş GERÇEK bir değer
    # olarak yazılıyor — bu, Fact_Mevcut'ta bulunup düzeltilen AYNI hata
    # sınıfının Fact_Norm'daki, daha sonra fark edilen bir örneğiydi:
    # pandas/openpyxl bu formülleri hiç hesaplamıyordu, bu yüzden HER
    # Python tabanlı okuma bu hücreleri boş (NaN) görüyordu.
    assert str(wb["Fact_Norm"]["B2"].value).strip() not in ("", "None")
    assert not str(wb["Fact_Norm"]["B2"].value).startswith("=")
    # DÜZELTME: Fact_Mevcut'ta Mağaza/Unvan artık HESAPLANMAMIŞ bir VLOOKUP
    # formülü değil, Python'da anında çözülmüş GERÇEK bir değer olarak
    # yazılıyor — kök neden buydu: pandas/openpyxl bu formülleri hiç
    # hesaplamıyordu (LibreOffice recalculation adımı bu yolda hiç
    # çalışmıyordu), bu yüzden HER Python tabanlı okuma (PDF/Excel/panel)
    # bu hücreleri boş görüyordu ve işten çıkan bir personelin satırı
    # load()'un çıktısından TAMAMEN kayboluyordu. Artık gerçek değer
    # doğrudan yazıldığı için bu bağımlılık ortadan kalktı.
    assert str(wb["Fact_Mevcut"]["B2"].value).strip() not in ("", "None")
    assert not str(wb["Fact_Mevcut"]["B2"].value).startswith("=")
    assert "Norm Fazlası" in str(wb["Fact_Mevcut"]["G2"].value)
    assert wb["Fact_Mevcut"].max_row == len(tables["Fact_Mevcut"]) + 1
