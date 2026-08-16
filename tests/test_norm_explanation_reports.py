from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.excel_report import build_boxed_manager_excel


def test_fact_norm_explanation_becomes_excel_comment(tmp_path):
    """DÜZELTME: build_boxed_manager_excel() önceden Fact_Norm'daki
    unvan-düzeyi (kişiye bağlı olmayan) 'Açıklama' sütununu hiç
    okumuyordu (notes_by_key = {} sabitti, "Fact_Norm açıklaması
    kullanılmaz" yorumuyla bilerek atlanıyordu). Artık gerçekten
    okunuyor ve o unvana ait HER satırın 'Unvan' hücresine yorum
    olarak ekleniyor (personel-düzeyi notlar 'Ad Soyad' hücresinde
    kalmaya devam ediyor — ikisi karışmaz)."""
    st = pd.DataFrame([{
        "Bölge Sorumlusu":"ALİ ÇELİK", "Mağaza":"TEST MAĞAZA",
        "Aktif Mevcut":1, "Norm Kadro":1, "Norm Eksiği":0, "Norm Fazlası":0, "Net Fark":0
    }])
    norm = pd.DataFrame([{
        "Mağaza":"TEST MAĞAZA", "Unvan":"ŞARKÜTERİ", "Norm Kadro":1,
        "Açıklama":"Vardiya planında öğleden sonra desteklenmelidir."
    }])
    staff = pd.DataFrame([{
        "Mağaza":"TEST MAĞAZA", "Unvan":"UZMAN ŞARKÜTERİ", "Departman":"ŞARKÜTERİ",
        "İsim Soyisim":"TEST PERSONEL", "İşe Giriş":pd.Timestamp("2026-01-01")
    }])
    out=tmp_path/"boxed.xlsx"
    build_boxed_manager_excel(st,norm,staff,output_path=out)
    wb=load_workbook(out)
    ws=wb["ALİ ÇELİK"]
    comments=[c.comment.text for row in ws.iter_rows() for c in row if c.comment]
    assert any("öğleden sonra" in text for text in comments)
