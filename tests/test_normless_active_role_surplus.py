from __future__ import annotations

from pathlib import Path

import pandas as pd


def _sheets():
    norm = pd.DataFrame([
        {
            "MağazaID": "M1", "Mağaza": "AKEVLER", "Bölge Sorumlusu": "B1",
            "UnvanID": "U1", "Unvan": "KASİYER", "Norm Kadro": 1,
        }
    ])
    staff = pd.DataFrame([
        {
            "MağazaID": "M1", "Mağaza": "AKEVLER", "Bölge Sorumlusu": "B1",
            "PersonelID": "P1", "UnvanID": "U1", "Unvan": "KASİYER",
            "Departman": "KASİYER", "İsim Soyisim": "KİŞİ 1", "İşten Çıkış": None,
        },
        {
            "MağazaID": "M1", "Mağaza": "AKEVLER", "Bölge Sorumlusu": "B1",
            "PersonelID": "P2", "UnvanID": "U99", "Unvan": "ANAHTARCI",
            "Departman": "ANAHTARCI", "İsim Soyisim": "EMİNE AKSARI", "İşten Çıkış": None,
        },
        {
            "MağazaID": "M1", "Mağaza": "AKEVLER", "Bölge Sorumlusu": "B1",
            "PersonelID": "P3", "UnvanID": "U99", "Unvan": "ANAHTARCI",
            "Departman": "ANAHTARCI", "İsim Soyisim": "HAKAN AKSARI", "İşten Çıkış": None,
        },
    ])
    return norm, staff


def test_state_counts_active_role_without_fact_norm_as_surplus():
    from src.state_engine import state

    norm, staff = _sheets()
    stores, titles = state(norm, staff, {})
    anahtarci = titles[titles["Unvan"].eq("ANAHTARCI")].iloc[0]

    assert int(anahtarci["Norm Kadro"]) == 0
    assert int(anahtarci["Aktif Mevcut"]) == 2
    assert int(anahtarci["Norm Fazlası"]) == 2
    assert bool(anahtarci["Normda Tanımlı Değil"])
    assert anahtarci["Norm Tanımı Durumu"] == "NORMDA TANIMLI DEĞİL (+2)"
    assert stores.attrs["kpi_override"]["Norm Fazlası"] == 2


def test_dashboard_counts_active_role_without_fact_norm_as_surplus(tmp_path: Path):
    from services.dashboard_model import build_dashboard_model

    norm, staff = _sheets()
    sheets = {
        "Fact_Norm": norm, "Fact_Mevcut": staff,
        "Dim_Magaza": pd.DataFrame(), "Dim_Unvan": pd.DataFrame(),
    }
    _, detail, stores, kpis = build_dashboard_model(sheets, tmp_path / "olmayan.json")
    anahtarci = detail[detail["Unvan"].eq("ANAHTARCI")].iloc[0]

    assert int(anahtarci["Norm"]) == 0
    assert int(anahtarci["Mevcut"]) == 2
    assert int(anahtarci["Fazla"]) == 2
    assert anahtarci["Norm Tanımı Durumu"] == "NORMDA TANIMLI DEĞİL (+2)"
    assert int(stores.iloc[0]["Fazla"]) == 2
    assert kpis["Norm Fazlası"] == 2


def test_static_excel_counts_active_role_without_fact_norm_as_surplus(tmp_path: Path):
    import openpyxl
    from services.formula_bagimsiz_hesapla import statiklestir

    path = tmp_path / "input.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in {
        "Dim_Magaza": [["MağazaID", "Mağaza"], ["M1", "AKEVLER"]],
        "Dim_Unvan": [["UnvanID", "Unvan"], ["U1", "KASİYER"], ["U99", "ANAHTARCI"]],
        "Fact_Norm": [["MağazaID", "UnvanID", "Norm Kadro"], ["M1", "U1", 1]],
        "Fact_Mevcut": [
            ["MağazaID", "Departman", "Unvan", "İşten Çıkış"],
            ["M1", "KASİYER", "KASİYER", None],
            ["M1", "ANAHTARCI", "ANAHTARCI", None],
            ["M1", "ANAHTARCI", "ANAHTARCI", None],
        ],
        "Norm_Durumu": [[
            "MağazaID", "Mağaza", "UnvanID", "Unvan", "Norm", "Mevcut",
            "YardımcıUID", "YardımcıNorm", "YardımcıMevcut", "ToplamNorm",
            "ToplamMevcut", "Yardımcımı", "Eksik", "Fazla",
        ]],
    }.items():
        ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row)
    wb.save(path)

    assert statiklestir(path) is True
    ws = openpyxl.load_workbook(path, data_only=True)["Norm_Durumu"]
    anahtarci = next(row for row in ws.iter_rows(min_row=2, values_only=True) if row[3] == "ANAHTARCI")
    assert anahtarci[4] == 0
    assert anahtarci[5] == 2
    assert anahtarci[13] == 2


def test_dashboard_and_state_use_identical_kpis_for_shipped_input():
    from services.dashboard_model import build_dashboard_model
    from services.personnel_status import active_people
    from src.state_engine import state

    root = Path(__file__).resolve().parents[1]
    sheets = pd.read_excel(root / "input" / "OMEHR_AI_NORM_TRANSFER_INPUT.xlsx", sheet_name=None)
    stores, _ = state(sheets["Fact_Norm"], active_people(sheets["Fact_Mevcut"]), sheets)
    expected = stores.attrs["kpi_override"]
    _, _, _, actual = build_dashboard_model(
        sheets, root / "reference" / "KONTROL_NORM_KADRO_24_07_2026.xlsx"
    )

    assert actual == expected
    assert actual["Norm Eksiği"] == 48
    assert actual["Norm Fazlası"] == 37
