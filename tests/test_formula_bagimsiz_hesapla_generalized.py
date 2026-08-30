"""services/formula_bagimsiz_hesapla.py::statiklestir() — regresyon testleri.

Kapsam: bu modül, CEO Özet'in "Mağaza KPI Skor Kartı" bölümünün ve
Bölge&Mağaza/Unvan Analizi sekmelerinin (LibreOffice yoksa) okuduğu
Norm_Durumu Excel sayfasını dolduruyordu — ÖNCEDEN kendi SABİT kodlanmış
kademe/aile dengeleme mantığıyla (yalnız 4 aile, Yönetici hariç, 0 ana
personelken dengeleme yapmayan bir şart ile), src.state_engine::state()
ile SESSİZCE senkron olmayan ayrı bir hesaplama yapıyordu. Artık ikisi de
aynı ortak kaynağı (services.norm_rule_config.resolve_family_key /
resolve_assistant_pairs, src.state_engine._staff_norm_family) kullanıyor.

Bu testler, gerçek bir .xlsx dosyası üretip statiklestir()'i uçtan uca
çalıştırarak Norm_Durumu sayfasının doğru dolduğunu kanıtlar.
"""

import openpyxl
import pytest

from services.formula_bagimsiz_hesapla import statiklestir


def _test_workbook(tmp_path):
    path = tmp_path / "test_input.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Dim_Magaza")
    ws.append(["MağazaID", "Mağaza"])
    ws.append(["M1", "TEST MAĞAZA"])

    ws = wb.create_sheet("Dim_Unvan")
    ws.append(["UnvanID", "Unvan"])
    ws.append(["U1", "KASİYER"])
    ws.append(["U2", "KASİYER YARDIMCISI"])
    ws.append(["U3", "MANAV TERAZİ"])

    ws = wb.create_sheet("Fact_Norm")
    ws.append(["MağazaID", "UnvanID", "Norm Kadro"])
    ws.append(["M1", "U1", 2])
    ws.append(["M1", "U2", 1])
    ws.append(["M1", "U3", 1])

    ws = wb.create_sheet("Fact_Mevcut")
    ws.append(["MağazaID", "Departman", "Unvan", "İşten Çıkış"])
    ws.append(["M1", "KASİYER", "UZMAN KASİYER", None])
    ws.append(["M1", "KASİYER YARDIMCISI", "KASİYER YARDIMCISI", None])
    ws.append(["M1", "KASİYER YARDIMCISI", "KASİYER YARDIMCISI", None])
    # MANAV TERAZİ: normu var (1) ama hiç personeli yok -> eksik=1 kalmalı.

    ws = wb.create_sheet("Norm_Durumu")
    ws.append([
        "MağazaID", "Mağaza", "UnvanID", "Unvan", "Norm", "Mevcut",
        "YardımcıUID", "YardımcıNorm", "YardımcıMevcut", "ToplamNorm",
        "ToplamMevcut", "Yardımcımı", "Eksik", "Fazla",
    ])

    wb.save(path)
    return path


def _norm_durumu_rows(path):
    wb = openpyxl.load_workbook(path)
    ws = wb["Norm_Durumu"]
    rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            rows[row[3]] = row  # Unvan adına göre anahtarla
    return rows


def test_statiklestir_kasiyer_ailesini_config_disinda_otomatik_dengeler(tmp_path):
    """KASİYER config_norm_rules.json'da hiç tanımlı değil — ama 'Uzman
    Kasiyer' otomatik olarak KASİYER ailesine sayılmalı ve KASİYER +
    KASİYER YARDIMCISI aile toplamı dengede olduğu için ikisi de
    Eksik=0/Fazla=0 göstermeli (src.state_engine.state() ile aynı sonuç)."""
    path = _test_workbook(tmp_path)
    assert statiklestir(path) is True

    rows = _norm_durumu_rows(path)
    kasiyer = rows["KASİYER"]
    yardimci = rows["KASİYER YARDIMCISI"]
    # Kolon sırası: ... Norm(4) Mevcut(5) ... Eksik(12) Fazla(13)
    assert kasiyer[4] == 2 and kasiyer[5] == 1  # norm=2, mevcut=1 (Uzman Kasiyer dahil)
    assert kasiyer[12] == 0, "Kasiyer ailesi (Uzman Kasiyer dahil) otomatik dengelenmeli"
    assert yardimci[13] == 0, "Yardımcı fazlası otomatik dengelenmeli"


def test_statiklestir_manav_terazi_otomatik_kuraldan_etkilenmez(tmp_path):
    """MANAV TERAZİ, 'UZMAN '/'ELİT ' ile başlamadığı ve ' YARDIMCISI'
    ile bitmediği için otomatik kurala hiç girmemeli — normu var (1),
    hiç personeli yok, Eksik=1 olarak KALMALI."""
    path = _test_workbook(tmp_path)
    statiklestir(path)

    rows = _norm_durumu_rows(path)
    manav_teraz = rows["MANAV TERAZİ"]
    assert manav_teraz[4] == 1 and manav_teraz[5] == 0  # norm=1, mevcut=0
    assert manav_teraz[12] == 1, "MANAV TERAZİ dengelenmemeli, gerçek eksik görünmeli"
    assert manav_teraz[13] == 0
