from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

from src.excel_report import executive_excel
from src.pdf_report import _build_store_pdf


def _frames():
    staff = pd.DataFrame([{
        'MağazaID':'M015','Mağaza':'GAZİEMİR-1','Bölge Sorumlusu':'ALİ ÇELİK',
        'UnvanID':'U036','Unvan':'YÖNETİCİ','Departman':'YÖNETİCİ',
        'Açıklama':'raporlu','İsim Soyisim':'FAHRİYE DRAGONAVA',
        'İşe Giriş':'2020-01-01','İşten Çıkış':None,'Durum':'Aktif'
    }])
    norm = pd.DataFrame([{
        'MağazaID':'M015','Mağaza':'GAZİEMİR-1','Bölge Sorumlusu':'ALİ ÇELİK',
        'UnvanID':'U036','Unvan':'YÖNETİCİ','Norm Kadro':1,'Açıklama':''
    }])
    tt = pd.DataFrame([{
        'Bölge Sorumlusu':'ALİ ÇELİK','MağazaID':'M015','Mağaza':'GAZİEMİR-1',
        'Unvan':'YÖNETİCİ','Aktif Mevcut':1,'Norm Kadro':1,
        'Norm Eksiği':0,'Norm Fazlası':0,'Net Fark':0
    }])
    st = pd.DataFrame([{
        'Bölge Sorumlusu':'ALİ ÇELİK','MağazaID':'M015','Mağaza':'GAZİEMİR-1',
        'Aktif Mevcut':1,'Norm Kadro':1,'Norm Eksiği':0,'Norm Fazlası':0,'Net Fark':0
    }])
    kpi={'Aktif Mevcut':1,'Toplam Norm':1,'Norm Eksiği':0,'Norm Fazlası':0,'Net İhtiyaç':0}
    return staff,norm,tt,st,kpi


def test_personnel_explanation_in_executive_excel():
    staff,norm,tt,st,kpi=_frames()
    path=executive_excel(kpi,st,tt,{},pd.DataFrame(),'x',input_sheets={},staff=staff)
    wb=load_workbook(path)
    ws=wb['Mağaza-Unvan Bazlı']
    headers={c.value:c.column for c in ws[1]}
    assert ws.cell(2,headers['Açıklama']).value == 'raporlu'
    target=ws.cell(2,headers['Personel Adı Soyadı'])
    assert target.comment and 'raporlu' in target.comment.text.lower()
    assert target.fill.fgColor.rgb.endswith('FFF2CC')


def test_personnel_explanation_pdf_builds(tmp_path):
    staff,norm,tt,st,kpi=_frames()
    out=tmp_path/'person_note.pdf'
    _build_store_pdf(out,kpi,norm,staff,pd.DataFrame(),include_summary=False,sheets={})
    assert out.exists() and out.stat().st_size > 10000
