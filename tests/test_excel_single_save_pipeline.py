from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.excel_report import executive_excel
from src.kpi_engine import _yaz_kpi_mutabakat_sayfasi


def test_executive_excel_can_be_enriched_before_single_save(isolated_root):
    """Orkestratör aynı çalışma kitabını zenginleştirip yalnız bir kez kaydeder."""
    st = pd.DataFrame([{
        "Bölge Sorumlusu": "TEST", "MağazaID": "M1", "Mağaza": "TEST",
        "Aktif Mevcut": 1, "Norm Kadro": 1, "Norm Eksiği": 0,
        "Norm Fazlası": 0, "Net Fark": 0,
    }])
    tt = st.assign(Unvan="YÖNETİCİ")
    kpi = {"Aktif Mevcut": 1, "Toplam Norm": 1, "Norm Eksiği": 0,
           "Norm Fazlası": 0, "Net İhtiyaç": 0}
    mutabakat = {
        "tutarli": True, "toplam_aktif_mevcut": 1,
        "norm_kapsamindaki_aktif_mevcut": 1, "norm_disi_calisan_sayisi": 0,
        "toplam_norm": 1, "brut_norm_farki": 0, "toplam_norm_fazlasi": 0,
        "toplam_norm_eksigi": 0, "net_pozisyon_farki": 0,
        "net_ihtiyac_kpi": 0, "aciklanabilir_fark_tablosu": pd.DataFrame(),
    }

    output, workbook = executive_excel(
        kpi, st, tt, {}, pd.DataFrame(), "test", return_workbook=True,
    )
    assert not Path(output).exists()

    _yaz_kpi_mutabakat_sayfasi(output, mutabakat, workbook=workbook)
    assert "KPI_Mutabakat_Kontrolu" in workbook.sheetnames
    workbook.save(output)

    assert "KPI_Mutabakat_Kontrolu" in load_workbook(output, read_only=True).sheetnames
