"""services.powerbi_export — Power BI'a hazır star şema testleri.

Kapsam: Dim tablolarının tekilleştirilmesi, ID sütunlarının Power BI
ilişkileri için tutarlı METİN'e sabitlenmesi, yetim (Dim'de karşılığı
olmayan) fact kayıtlarının sessizce atılmayıp ayrı bir sayfada
raporlanması, ve Dim_Tarih takvim boyutunun üretilmesi.
"""
from __future__ import annotations

import pandas as pd
import pytest


def _ornek_sheets():
    dim_magaza = pd.DataFrame([
        {"MağazaID": 1, "Mağaza": "A Mağazası", "Bölge Sorumlusu": "Test"},
        {"MağazaID": 2, "Mağaza": "B Mağazası", "Bölge Sorumlusu": "Test"},
    ])
    dim_unvan = pd.DataFrame([{"UnvanID": "U1", "Unvan": "Kasiyer"}])
    fact_norm = pd.DataFrame([
        {"MağazaID": 1, "UnvanID": "U1", "Norm Kadro": 5},
        {"MağazaID": 2, "UnvanID": "U1", "Norm Kadro": 3},
    ])
    fact_mevcut = pd.DataFrame([
        {"PersonelID": "P1", "MağazaID": 1, "UnvanID": "U1", "İşe Giriş": "2024-01-01", "İşten Çıkış": None},
    ])
    return {
        "Dim_Magaza": dim_magaza, "Dim_Unvan": dim_unvan,
        "Fact_Norm": fact_norm, "Fact_Mevcut": fact_mevcut,
    }


def test_build_powerbi_model_produces_all_expected_tables():
    from services.powerbi_export import build_powerbi_model

    model = build_powerbi_model(_ornek_sheets())
    assert len(model["dim_magaza"]) == 2
    assert len(model["dim_unvan"]) == 1
    assert len(model["fact_norm"]) == 2
    assert len(model["fact_mevcut"]) == 1
    assert len(model["yetim_norm"]) == 0
    assert len(model["yetim_mevcut"]) == 0
    assert len(model["dim_tarih"]) > 0


def test_id_columns_are_normalized_to_consistent_text():
    """Power BI ilişkilerinin en sık kırılma nedeni: '1' (metin) ile 1
    (sayı) FARKLI değer sayılır. Bu, ID sütunlarının HER ZAMAN tutarlı
    bir metin gösterimine sabitlendiğini doğrular."""
    from services.powerbi_export import build_powerbi_model

    sheets = _ornek_sheets()
    # Fact_Norm'da MağazaID'yi KARIŞIK tipte yap (bazı satırlar metin, bazıları sayı).
    sheets["Fact_Norm"]["MağazaID"] = sheets["Fact_Norm"]["MağazaID"].astype(object)
    sheets["Fact_Norm"].loc[0, "MağazaID"] = "1"  # metin
    sheets["Fact_Norm"].loc[1, "MağazaID"] = 2  # sayı

    model = build_powerbi_model(sheets)
    # Her iki taraf da (Dim ve Fact) aynı temsille eşleşmeli -> yetim OLMAMALI.
    assert len(model["yetim_norm"]) == 0
    assert set(model["fact_norm"]["MağazaID"]) == {"1", "2"}
    assert set(model["dim_magaza"]["MağazaID"]) == {"1", "2"}


def test_orphan_fact_records_are_isolated_not_silently_dropped():
    """Dim tablosunda karşılığı olmayan bir MağazaID SESSİZCE atılmamalı
    — ayrı bir 'yetim' tablosunda, NEDENİYLE birlikte raporlanmalı."""
    from services.powerbi_export import build_powerbi_model

    sheets = _ornek_sheets()
    yeni_satir = pd.DataFrame([{"MağazaID": 99, "UnvanID": "U1", "Norm Kadro": 1}])
    sheets["Fact_Norm"] = pd.concat([sheets["Fact_Norm"], yeni_satir], ignore_index=True)

    model = build_powerbi_model(sheets)
    assert len(model["yetim_norm"]) == 1
    assert "99" in str(model["yetim_norm"].iloc[0]["Yetim Nedeni"])
    assert model["yetim_norm"].iloc[0]["Kaynak Sayfa"] == "Fact_Norm"
    # Temiz tabloda yetim satır OLMAMALI.
    assert "99" not in set(model["fact_norm"]["MağazaID"])


def test_duplicate_dim_magaza_rows_are_deduplicated():
    """Aynı MağazaID iki kez girilmişse (kullanıcı hatası), Power BI
    modeli TEKİLLEŞTİRİLMİŞ bir Dim tablosu üretmeli — aksi halde Power
    BI bire-çok ilişki kuramaz (Dim tarafında tekillik şart)."""
    from services.powerbi_export import build_powerbi_model

    sheets = _ornek_sheets()
    tekrar = sheets["Dim_Magaza"].iloc[[0]].copy()
    sheets["Dim_Magaza"] = pd.concat([sheets["Dim_Magaza"], tekrar], ignore_index=True)

    model = build_powerbi_model(sheets)
    assert len(model["dim_magaza"]) == 2  # 3 değil, tekilleştirilmiş
    assert model["dim_magaza"]["MağazaID"].is_unique


def test_relationships_guide_documents_all_join_columns():
    from services.powerbi_export import build_powerbi_model

    model = build_powerbi_model(_ornek_sheets())
    iliskiler = model["iliskiler"]
    ciftler = set(zip(iliskiler["Fact Tablosu"], iliskiler["Dim Tablosu"]))
    assert ("Fact_Norm", "Dim_Magaza") in ciftler
    assert ("Fact_Norm", "Dim_Unvan") in ciftler
    assert ("Fact_Mevcut", "Dim_Magaza") in ciftler
    assert ("Fact_Mevcut", "Dim_Unvan") in ciftler


def test_export_powerbi_workbook_writes_a_real_readable_file(tmp_path):
    from services.powerbi_export import export_powerbi_workbook, OUTPUT_FILE_NAME

    sonuc = export_powerbi_workbook(_ornek_sheets(), tmp_path)

    from pathlib import Path
    assert Path(sonuc["file"]).is_file()
    assert Path(sonuc["file"]).name == OUTPUT_FILE_NAME
    assert sonuc["dim_magaza_sayisi"] == 2
    assert sonuc["yetim_norm_sayisi"] == 0

    # Gerçekten okunabilir bir Excel dosyası mı?
    sheets = pd.read_excel(sonuc["file"], sheet_name=None)
    for beklenen in ("Dim_Magaza", "Dim_Unvan", "Dim_Tarih", "Fact_Norm", "Fact_Mevcut", "Iliskiler_Rehberi"):
        assert beklenen in sheets


def test_export_only_creates_orphan_sheets_when_orphans_exist(tmp_path):
    """Yetim kayıt yoksa (temiz veri), Yetim_Kayitlar_* sayfaları HİÇ
    oluşturulmamalı — gereksiz boş sayfalarla dosyayı kirletmemek için."""
    from services.powerbi_export import export_powerbi_workbook

    sonuc = export_powerbi_workbook(_ornek_sheets(), tmp_path)
    sheets = pd.read_excel(sonuc["file"], sheet_name=None)
    assert "Yetim_Kayitlar_Norm" not in sheets
    assert "Yetim_Kayitlar_Mevcut" not in sheets


def test_id_cell_type_in_excel_is_actually_text_not_number(tmp_path):
    """En kritik doğrulama: Excel'in HAM hücre tipi gerçekten metin (s)
    olmalı — pandas'ın geri okurken sayıya çevirmesi ayrı bir konu,
    Power BI dosyayı AÇARKEN gördüğü ham tip önemlidir."""
    import openpyxl
    from services.powerbi_export import export_powerbi_workbook

    sonuc = export_powerbi_workbook(_ornek_sheets(), tmp_path)
    wb = openpyxl.load_workbook(sonuc["file"])
    ws = wb["Dim_Magaza"]
    for row in range(2, ws.max_row + 1):
        assert ws.cell(row=row, column=1).data_type == "s"


def test_missing_dim_sheets_do_not_crash_gracefully_return_empty():
    """Kaynak input'ta Dim_Magaza/Dim_Unvan hiç yoksa (aşırı uç durum),
    motor çökmemeli — boş Dim tablolarıyla devam etmeli (bu durumda
    doğal olarak TÜM fact kayıtları yetim sayılır)."""
    from services.powerbi_export import build_powerbi_model

    model = build_powerbi_model({"Fact_Norm": pd.DataFrame([{"MağazaID": 1, "UnvanID": "U1", "Norm Kadro": 5}])})
    assert model["dim_magaza"].empty
    assert len(model["yetim_norm"]) == 1
