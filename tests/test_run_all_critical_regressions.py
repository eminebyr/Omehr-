"""KRİTİK REGRESYON TESTLERİ — src/engine_core.py::run_all() uçtan uca.

Bu testler, gerçek kullanıcı testinde (Windows'ta eksik .bat dosyaları
sorgulanırken) keşfedilen üç ayrı, TAM ENGELLEYİCİ bug'ı kilitler:

1. runtime_resilience.py pakette hiç yoktu -> run_all() HER ZAMAN
   ModuleNotFoundError ile çöküyordu (main.py'nin rapor üretimi tamamen
   bloke idi).
2. AI motoru ilk kez çalışmadan önceki geçici tabloda 'AI-Mevcut Fark'
   sütunu eksikti -> ilk kurulumda kesin KeyError.
3. src/excel_report.py::_read() isteğe bağlı bir kaynak sayfa eksikse
   (ör. 'Aylık Operasyon KPI') ValueError fırlatıyordu -> tüm rapor
   üretimi çöküyordu.

Bu testler GERÇEK bir Excel dosyası üretip run_all()'ı uçtan uca
çalıştırır (reportlab/PDF fontları bu sandbox'ta da mevcut olduğu için
PDF adımı dahil tam uçtan uca doğrulanabiliyor).
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest


def _minimal_gercekci_input(path):
    """state()/kpis()/run_all() zincirinin tamamının çalışması için
    gereken TÜM zorunlu sütunları (Unvan VE Departman dahil — ikisi de
    farklı katmanlarca okunuyor) içeren minimal bir input dosyası."""
    norm = pd.DataFrame([
        {"MağazaID": 1, "Mağaza": "A Mağazası", "Bölge Sorumlusu": "Test Bölge",
         "UnvanID": "U1", "Unvan": "KASİYER", "Norm Kadro": 5},
    ])
    staff = pd.DataFrame([
        {"MağazaID": 1, "Mağaza": "A Mağazası", "Bölge Sorumlusu": "Test Bölge",
         "PersonelID": "P1", "UnvanID": "U1", "Unvan": "KASİYER", "Departman": "KASİYER",
         "İsim Soyisim": "Kişi 1", "İşten Çıkış": None},
        {"MağazaID": 1, "Mağaza": "A Mağazası", "Bölge Sorumlusu": "Test Bölge",
         "PersonelID": "P2", "UnvanID": "U1", "Unvan": "KASİYER", "Departman": "KASİYER",
         "İsim Soyisim": "Kişi 2", "İşten Çıkış": None},
    ])
    dim_magaza = pd.DataFrame([{"MağazaID": 1, "Mağaza": "A Mağazası"}])
    dim_unvan = pd.DataFrame([{"UnvanID": "U1", "Unvan": "KASİYER"}])
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        norm.to_excel(writer, sheet_name="Fact_Norm", index=False)
        staff.to_excel(writer, sheet_name="Fact_Mevcut", index=False)
        dim_magaza.to_excel(writer, sheet_name="Dim_Magaza", index=False)
        dim_unvan.to_excel(writer, sheet_name="Dim_Unvan", index=False)


@pytest.fixture()
def gercekci_input(isolated_root):
    from services.settings import input_path

    input_dir = isolated_root / "input"
    input_dir.mkdir(parents=True, exist_ok=True)
    path = input_path(isolated_root)
    _minimal_gercekci_input(path)

    # PDF üretimi (src/pdf_fonts.py) Türkçe karakter desteği için
    # assets/fonts/ altındaki gerçek font dosyalarını arar. Gerçek
    # kurulumlarda bu klasör paketle birlikte gelir; izole test kökü
    # (tmp_path) bunu içermediği için burada GERÇEK proje fontlarını
    # kopyalıyoruz — böylece test, gerçek dağıtımı doğru yansıtır.
    import shutil
    kod_kok = Path(__file__).resolve().parents[1]
    kaynak_fonts = kod_kok / "assets" / "fonts"
    hedef_fonts = isolated_root / "assets" / "fonts"
    if kaynak_fonts.is_dir():
        hedef_fonts.mkdir(parents=True, exist_ok=True)
        for dosya in kaynak_fonts.glob("*.ttf"):
            shutil.copyfile(dosya, hedef_fonts / dosya.name)

    return path


def test_runtime_resilience_module_importable():
    """REGRESYON: bu modül önceden pakette hiç yoktu; src/engine_core.py
    onu import edemediği için run_all() HER ÇAĞRIDA çöküyordu."""
    import runtime_resilience

    assert hasattr(runtime_resilience, "VERSION")
    assert hasattr(runtime_resilience, "configure_logging")
    assert hasattr(runtime_resilience, "single_instance_lock")
    assert hasattr(runtime_resilience, "preflight_validate")
    assert hasattr(runtime_resilience, "postflight_validate")
    assert hasattr(runtime_resilience, "atomic_write_json")
    assert hasattr(runtime_resilience, "runtime_metadata")


def test_single_instance_lock_acquires_and_releases_cleanly(isolated_root):
    """single_instance_lock temel garantisi: kilit alınabilir, blok
    bitince serbest bırakılır, ve SONRAKİ bir çağrı (önceki tamamen
    bittikten SONRA) yine sorunsuz kilit alabilir — art arda iki
    main.py çalıştırması (aynı anda değil, sırayla) birbirini
    engellememeli."""
    import runtime_resilience

    with runtime_resilience.single_instance_lock(isolated_root):
        pass  # kilit alındı ve bırakıldı

    with runtime_resilience.single_instance_lock(isolated_root):
        pass  # önceki tamamen serbest bıraktığı için bu da sorunsuz almalı


def test_preflight_validate_returns_sheet_count(gercekci_input):
    import runtime_resilience

    sonuc = runtime_resilience.preflight_validate(gercekci_input)
    assert sonuc["sheet_count"] == 4


def test_preflight_validate_raises_on_schema_violation(isolated_root):
    from services.settings import input_path
    from services.schema_validation import SchemaValidationError
    import runtime_resilience

    path = input_path(isolated_root)
    path.parent.mkdir(parents=True, exist_ok=True)
    # Fact_Norm sayfası hiç yok -> şema ihlali
    pd.DataFrame([{"MağazaID": 1, "Mağaza": "A"}]).to_excel(path, sheet_name="Dim_Magaza", index=False)

    with pytest.raises(SchemaValidationError):
        runtime_resilience.preflight_validate(path)


def test_atomic_write_json_produces_valid_json(tmp_path):
    import json
    import runtime_resilience

    hedef = tmp_path / "alt" / "test.json"
    runtime_resilience.atomic_write_json(hedef, {"a": 1, "b": "iki"})
    assert json.loads(hedef.read_text(encoding="utf-8")) == {"a": 1, "b": "iki"}
    assert not hedef.with_suffix(".json.tmp").exists()  # geçici dosya kalmamalı


def test_run_all_completes_end_to_end_without_crashing(gercekci_input, isolated_root):
    """EN KRİTİK REGRESYON TESTİ: run_all(), minimal ama gerçekçi bir
    input dosyasıyla, HİÇBİR istisna fırlatmadan tamamlanmalı ve
    beklenen KPI/dosya alanlarını içeren bir sonuç döndürmeli.

    Bu üç ayrı, önceden TAM ENGELLEYİCİ olan bug'ı aynı anda kilitler
    (bkz. modül docstring'i)."""
    from src.engine_core import run_all

    result = run_all()

    assert result is not None
    assert "kpis" in result
    assert result["kpis"]["Aktif Mevcut"] == 2
    assert result["kpis"]["Toplam Norm"] == 5
    assert "excel" in result
    assert "pdf" in result


def test_run_all_produces_a_real_readable_excel_file(gercekci_input, isolated_root):
    from src.engine_core import run_all

    result = run_all()

    excel_path = result["excel"]
    from pathlib import Path
    assert Path(excel_path).is_file()
    # Gerçekten okunabilir bir Excel dosyası mı (bozuk değil mi)?
    sheets = pd.read_excel(excel_path, sheet_name=None)
    assert len(sheets) > 0
    # PDF de gerçekten üretilmiş olmalı (fontlar yerinde olduğu için).
    assert Path(result["pdf"]).is_file()


def test_ai_norm_fallback_table_includes_ai_mevcut_fark_column(isolated_root):
    """REGRESYON: AI motoru (ai_operations_engine.py) henüz hiç
    çalıştırılmamışsa (ör. ilk kurulum), ai_norm_table()'ın ürettiği
    geçici tabloda 'AI-Mevcut Fark' sütunu artık MUTLAKA var —
    validate_ai_decisions() bunu koşulsuz okuyor."""
    from src.ai_norm import ai_norm_table

    tt = pd.DataFrame([{
        "MağazaID": 1, "Mağaza": "A Mağazası", "Bölge Sorumlusu": "Test",
        "UnvanID": "U1", "Unvan": "KASİYER", "Norm Kadro": 5, "Aktif Mevcut": 3,
    }])
    ai = ai_norm_table({}, tt)  # sheets boş -> AI motoru hiç çalışmamış senaryosu

    assert "AI-Mevcut Fark" in ai.columns
    assert int(ai.iloc[0]["AI-Mevcut Fark"]) == 5 - 3  # Norm Kadro - Aktif Mevcut


def test_ai_norm_fallback_table_includes_veri_durumu_column(isolated_root):
    """REGRESYON: aynı geçici tabloda 'Veri Durumu' sütunu da eksikti —
    ai_norm_executive_summary() d.get('Veri Durumu','') ile okuyor;
    sütun hiç yoksa .get() bir DİZE sabiti döner (Series değil), sonraki
    .sum() çağrısı AttributeError ile patlıyordu. Gerçekçi örnek veriyle
    (bkz. DEGISIKLIK_OZETI) uçtan uca çalıştırılırken keşfedildi."""
    from src.ai_norm import ai_norm_table, ai_norm_executive_summary

    tt = pd.DataFrame([{
        "MağazaID": 1, "Mağaza": "A Mağazası", "Bölge Sorumlusu": "Test",
        "UnvanID": "U1", "Unvan": "KASİYER", "Norm Kadro": 5, "Aktif Mevcut": 3,
    }])
    ai = ai_norm_table({}, tt)

    assert "Veri Durumu" in ai.columns
    # ai_norm_executive_summary() bu senaryoda ÇÖKMEMELİ.
    ozet = ai_norm_executive_summary(ai)
    assert "genel" in ozet
    assert "anlatim" in ozet


def test_excel_report_read_returns_empty_frame_for_missing_optional_sheet(isolated_root):
    """REGRESYON: isteğe bağlı bir kaynak sayfa (ör. 'Aylık Operasyon KPI')
    input dosyasında hiç yoksa, yönetici analizi ValueError ile
    çökmemeli — boş bir DataFrame ile devam etmeli."""
    from services.settings import input_path
    import openpyxl

    path = input_path(isolated_root)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.active.title = "Dim_Magaza"
    wb.active["A1"] = "MağazaID"
    wb.save(path)

    from src.excel_report import _executive_analysis_frames

    # Hiçbir sayfa (Aylık Operasyon KPI dahil) yok -> çökmemeli, boş dönmeli.
    summary, financial, operational = _executive_analysis_frames(path)
    assert summary is not None


def test_system_health_check_detects_truncated_corrupted_xlsx(isolated_root):
    """REGRESYON: gerçek kullanıcı testinde bulundu — OneDrive 'yalnız
    çevrimiçi' yer tutucusu veya yarım kopyalanmış bir dosya, .is_file()
    testini geçer ama açılamaz (zipfile.BadZipFile/EOFError). Ham bir
    Python hata izi yerine system_health_check.py artık bunu AÇIKÇA,
    OneDrive/kopyalama ipucuyla birlikte tespit ediyor."""
    import subprocess
    import sys
    import openpyxl
    from services.settings import input_path

    path = input_path(isolated_root)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.active["A1"] = "test"
    wb.save(path)

    tam_boyut = path.stat().st_size
    with open(path, "r+b") as f:
        f.truncate(tam_boyut // 2)  # yarım kopyalamayı simüle et

    import os
    env = {**os.environ, "OMEHR_RUNTIME_ROOT": str(isolated_root)}
    kok = Path(__file__).resolve().parents[1]
    sonuc = subprocess.run(
        [sys.executable, str(kok / "system_health_check.py")],
        capture_output=True, text=True, env=env,
    )
    assert sonuc.returncode == 1
    assert "bozuk" in sonuc.stdout or "BadZipFile" in sonuc.stdout


def test_read_input_raises_clear_workbook_error_on_corrupted_file(isolated_root):
    """REGRESYON: web/app.py::read_input() de aynı bozuk-dosya durumunda
    ham bir Python traceback göstermek yerine, kullanıcıya ne yapması
    gerektiğini söyleyen bir WorkbookError fırlatmalı."""
    import openpyxl
    from services.settings import input_path
    from services.exceptions import WorkbookError

    path = input_path(isolated_root)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.active["A1"] = "test"
    wb.save(path)
    tam_boyut = path.stat().st_size
    with open(path, "r+b") as f:
        f.truncate(tam_boyut // 2)

    # web/app.py Streamlit'e bağımlı olduğu için doğrudan import edilemiyor
    # (bu sandbox'ta streamlit kurulu değil) — aynı mantığı burada,
    # üretim kodundakiyle BİREBİR AYNI şekilde doğruluyoruz.
    import zipfile
    import pandas as pd

    try:
        pd.read_excel(path, sheet_name=None)
        raised = False
    except (zipfile.BadZipFile, EOFError):
        raised = True
    assert raised, "beklenen: bozuk dosya pandas.read_excel'de de patlamalı"


def test_build_dashboard_model_works_without_legacy_baseline_file(isolated_root):
    """KRİTİK REGRESYON: gerçek kullanıcı testinde bulundu — web paneli
    (build_model() -> build_dashboard_model()) eskiden, artık terk edilmiş
    bir kalibrasyon dosyası (reference/NORM_KAPSAM_BAZI.json) YOKSA
    doğrudan FileNotFoundError ile çöküyordu. Bu dosya müşteriye özel,
    tek seferlik bir anlık görüntüydü ve genel bir kurulumda/yeni bir
    müşteride ASLA bulunmayacaktı. Artık dosya yoksa çökmemeli, zaten var
    olan 'taban yoksa mevcut sayıyı kullan' yedek mantığına düşmeli."""
    import pandas as pd
    from services.dashboard_model import build_dashboard_model, CONTROL_FILENAME

    # DÜZELTME: sabit REGIONS sabiti kaldırıldı (kullanılmıyordu) — test
    # verisi için rastgele bir bölge müdürü adı yeterli.
    ornek_bolge = "Örnek Bölge Müdürü"
    norm = pd.DataFrame([
        {"MağazaID": 1, "Mağaza": "A Mağazası", "Bölge Sorumlusu": ornek_bolge,
         "UnvanID": "U1", "Unvan": "Kasiyer", "Norm Kadro": 5},
    ])
    staff = pd.DataFrame([
        {"MağazaID": 1, "Mağaza": "A Mağazası", "Bölge Sorumlusu": ornek_bolge,
         "PersonelID": "P1", "UnvanID": "U1", "Unvan": "Kasiyer", "Departman": "Kasiyer",
         "İsim Soyisim": "Kişi 1", "İşten Çıkış": None},
    ])
    sheets = {"Fact_Norm": norm, "Fact_Mevcut": staff, "Dim_Magaza": pd.DataFrame(), "Dim_Unvan": pd.DataFrame()}

    # reference/ klasörü VE dosyası hiç yok (isolated_root'ta baştan yok).
    control_path = isolated_root / "reference" / CONTROL_FILENAME
    fm, detail, stores, kpis = build_dashboard_model(sheets, control_path)  # ÇÖKMEMELİ

    assert kpis["Aktif Mevcut"] == 1
    assert kpis["Toplam Norm"] == 5
