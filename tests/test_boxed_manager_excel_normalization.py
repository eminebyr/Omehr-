"""src/excel_report.py::build_boxed_manager_excel — kademe birleştirme
tutarlılık regresyon testi.

Kapsam: bu rapor önceden msum/persons eşleştirmesinde DOĞRUDAN Departman'a
bakıyordu — src.state_engine._staff_norm_family'nin (resmi motor, panelin
ve PDF'nin kullandığı) "gerçek unvan UZMAN X/ELİT X ise doğru aileye bağla"
düzeltmesini UYGULAMIYORDU. Somut etki: gerçek unvanı "Uzman Kasiyer" olan
bir personel bu raporda AYRI, norm tanımı olmayan bir satırda (yapay
Fazla=1 olarak) görünürken, aynı kişi resmi motorda KASİYER ailesinin
eksiğini kapatan biri sayılıyordu — iki rapor aynı kişi için çelişen
tablolar üretiyordu. Bu test, iki tarafın artık aynı toplamı ürettiğini
kanıtlar.
"""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.state_engine import state
from src.excel_report import build_boxed_manager_excel


def _frames():
    norm = pd.DataFrame([
        {'MağazaID': '1', 'Mağaza': 'TEST MAĞAZA', 'Bölge Sorumlusu': 'TEST BÖLGE', 'Unvan': 'KASİYER', 'Norm Kadro': 2},
        {'MağazaID': '1', 'Mağaza': 'TEST MAĞAZA', 'Bölge Sorumlusu': 'TEST BÖLGE', 'Unvan': 'KASİYER YARDIMCISI', 'Norm Kadro': 1},
    ])
    staff = pd.DataFrame([
        # Gerçek unvanı "Uzman Kasiyer", Departman alanı da (saha
        # girişinde sıkça olduğu gibi) "UZMAN KASİYER" — KASİYER değil.
        {'MağazaID': '1', 'Mağaza': 'TEST MAĞAZA', 'Bölge Sorumlusu': 'TEST BÖLGE',
         'İsim Soyisim': 'KİŞİ 1', 'Unvan': 'UZMAN KASİYER', 'Departman': 'UZMAN KASİYER'},
    ])
    return norm, staff


def test_boxed_manager_excel_totals_match_official_state_for_uzman_title(tmp_path: Path):
    norm, staff = _frames()
    st, tt = state(norm, staff, {})
    resmi_eksik = int(tt['Norm Eksiği'].sum())
    resmi_fazla = int(tt['Norm Fazlası'].sum())

    out = build_boxed_manager_excel(st, norm, staff, output_path=tmp_path / 'boxed.xlsx')
    wb = load_workbook(out)
    ws = wb[wb.sheetnames[0]]

    # Kart başlık satırındaki toplamları (E/F) oku.
    toplam_satir = None
    for row in ws.iter_rows(values_only=True):
        if row and row[0] and str(row[0]).startswith('MAĞAZA'):
            toplam_satir = row
            break
    assert toplam_satir is not None, "Mağaza toplam satırı bulunamadı"

    rapor_eksik = toplam_satir[7]  # 'E\n<sayı>' hücresi
    rapor_fazla = toplam_satir[9]  # 'F\n<sayı>' hücresi
    rapor_eksik_sayi = int(str(rapor_eksik).split('\n')[-1])
    rapor_fazla_sayi = int(str(rapor_fazla).split('\n')[-1])

    assert rapor_eksik_sayi == resmi_eksik, (
        f"Kutucuklu rapor Eksik={rapor_eksik_sayi}, resmi motor Eksik={resmi_eksik} — tutarsız"
    )
    assert rapor_fazla_sayi == resmi_fazla, (
        f"Kutucuklu rapor Fazla={rapor_fazla_sayi}, resmi motor Fazla={resmi_fazla} — tutarsız"
    )
    # 'Uzman Kasiyer' KASİYER ailesine sayıldığı için raporda AYRI,
    # norm tanımsız bir 'Fazla' satırı olarak görünmemeli.
    hucre_metinleri = [str(c) for row in ws.iter_rows(values_only=True) for c in row if c]
    assert not any('UZMAN KASİYER' in t and 'F\n1' in str(t) for t in hucre_metinleri)
