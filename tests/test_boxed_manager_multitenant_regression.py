from __future__ import annotations

"""KRİTİK ÇOK-KİRACILI REGRESYON TESTİ — Kutucuklu Yönetici Raporu.

src/excel_report.py::build_boxed_manager_excel() önceden 4 SABİT,
orijinal firmanın gerçek bölge sorumlusu ismine göre sayfa oluşturuyordu
(Ali Çelik, Derya Yardımcı, Cüneyt & Ayşe Avcu, Ertan Teki). BAŞKA HER
kiracı için TÜM sayfalar sessizce BOŞ üretiliyordu. Bu test, farklı
(gerçek olmayan) bölge sorumlusu isimleriyle raporun DOLU üretildiğini
doğrulayarak bu hatanın sessizce geri gelmesini engeller.
"""

import pandas as pd


def test_boxed_manager_excel_works_for_any_tenant_region_names(tmp_path):
    from src.excel_report import build_boxed_manager_excel

    st = pd.DataFrame([
        {"Mağaza": "MAĞAZA A", "Bölge Sorumlusu": "TAMAMEN FARKLI MÜDÜR", "Aktif Mevcut": 5,
         "Norm Kadro": 5, "Norm Eksiği": 0, "Norm Fazlası": 0, "Net Fark": 0},
        {"Mağaza": "MAĞAZA B", "Bölge Sorumlusu": "TAMAMEN FARKLI MÜDÜR", "Aktif Mevcut": 3,
         "Norm Kadro": 4, "Norm Eksiği": 1, "Norm Fazlası": 0, "Net Fark": -1},
    ])
    norm = pd.DataFrame([
        {"Mağaza": "MAĞAZA A", "Unvan": "KASİYER", "Norm Kadro": 5},
        {"Mağaza": "MAĞAZA B", "Unvan": "KASİYER", "Norm Kadro": 4},
    ])
    staff = pd.DataFrame([
        {"Mağaza": "MAĞAZA A", "Unvan": "KASİYER", "Departman": "KASİYER", "İsim Soyisim": f"Kişi {i}"}
        for i in range(5)
    ] + [
        {"Mağaza": "MAĞAZA B", "Unvan": "KASİYER", "Departman": "KASİYER", "İsim Soyisim": f"Kişi B{i}"}
        for i in range(3)
    ])

    out = tmp_path / "test_kutucuklu.xlsx"
    build_boxed_manager_excel(st, norm, staff, output_path=out)

    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert wb.sheetnames == ["TAMAMEN FARKLI MÜDÜR"], (
        f"REGRESYON: beklenen sayfa üretilmedi — bulunan: {wb.sheetnames}. "
        "Muhtemelen sabit bir isim listesi geri gelmiş."
    )
    ws = wb["TAMAMEN FARKLI MÜDÜR"]
    dolu_satir = sum(1 for row in ws.iter_rows(min_row=9, values_only=True) if any(row))
    assert dolu_satir > 0, "REGRESYON: sayfa üretildi ama tamamen boş — mağaza verisi eşleşmiyor."
