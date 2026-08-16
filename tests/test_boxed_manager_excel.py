from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from src.excel_report import build_boxed_manager_excel


def test_boxed_manager_excel_has_person_rows_and_expected_sheets(tmp_path: Path):
    st = pd.DataFrame([
        {'Bölge Sorumlusu':'Ali Çelik','Mağaza':'A Mağazası','Aktif Mevcut':2,'Norm Kadro':3,'Norm Eksiği':1,'Norm Fazlası':0,'Net Fark':-1},
        {'Bölge Sorumlusu':'Derya Yardımcı','Mağaza':'B Mağazası','Aktif Mevcut':2,'Norm Kadro':1,'Norm Eksiği':0,'Norm Fazlası':1,'Net Fark':1},
        {'Bölge Sorumlusu':'Cüneyt Çıkrıkçı','Mağaza':'C Mağazası','Aktif Mevcut':1,'Norm Kadro':1,'Norm Eksiği':0,'Norm Fazlası':0,'Net Fark':0},
        {'Bölge Sorumlusu':'Ayşe Avcu','Mağaza':'D Mağazası','Aktif Mevcut':1,'Norm Kadro':1,'Norm Eksiği':0,'Norm Fazlası':0,'Net Fark':0},
        {'Bölge Sorumlusu':'Ertan Teki','Mağaza':'E Mağazası','Aktif Mevcut':1,'Norm Kadro':1,'Norm Eksiği':0,'Norm Fazlası':0,'Net Fark':0},
    ])
    norm = pd.DataFrame([
        {'Mağaza':'A Mağazası','Unvan':'KASİYER','Norm Kadro':3},
        {'Mağaza':'B Mağazası','Unvan':'KASİYER','Norm Kadro':1},
        {'Mağaza':'C Mağazası','Unvan':'REYON','Norm Kadro':1},
        {'Mağaza':'D Mağazası','Unvan':'REYON','Norm Kadro':1},
        {'Mağaza':'E Mağazası','Unvan':'MANAV','Norm Kadro':1},
    ])
    staff = pd.DataFrame([
        {'Mağaza':'A Mağazası','Departman':'KASİYER','Unvan':'KASİYER','İsim Soyisim':'Ali Test','İşe Giriş':'2026-01-01'},
        {'Mağaza':'A Mağazası','Departman':'KASİYER','Unvan':'KASİYER YRD.','İsim Soyisim':'Ayşe Test','İşe Giriş':'2026-02-01'},
        {'Mağaza':'B Mağazası','Departman':'KASİYER','Unvan':'KASİYER','İsim Soyisim':'Biri','İşe Giriş':'2026-01-01'},
        {'Mağaza':'B Mağazası','Departman':'KASİYER','Unvan':'KASİYER','İsim Soyisim':'Fazla Kişi','İşe Giriş':'2026-03-01'},
        {'Mağaza':'C Mağazası','Departman':'REYON','Unvan':'REYON','İsim Soyisim':'C Kişi','İşe Giriş':'2026-01-01'},
        {'Mağaza':'D Mağazası','Departman':'REYON','Unvan':'REYON','İsim Soyisim':'D Kişi','İşe Giriş':'2026-01-01'},
        {'Mağaza':'E Mağazası','Departman':'MANAV','Unvan':'MANAV','İsim Soyisim':'E Kişi','İşe Giriş':'2026-01-01'},
    ])
    out = build_boxed_manager_excel(st, norm, staff, {}, tmp_path/'BASDAS_Kutucuklu_Yonetici_Raporu.xlsx')
    wb = load_workbook(out, read_only=False, data_only=True)
    # DÜZELTME: sayfalar artık TENANT'IN KENDİ verisindeki gerçek "Bölge
    # Sorumlusu" değerlerinden dinamik türetiliyor (bkz. src/excel_report.py
    # üzerindeki düzeltme notu) — sabit bir "Ali Çelik/Derya Yardımcı/
    # Cüneyt & Ayşe Avcu/Ertan Teki" isim listesi YOK. Bu test verisinde
    # "Cüneyt Çıkrıkçı" ve "Ayşe Avcu" AYRI metin değerleri olduğu için artık
    # AYRI sayfalar üretir (eskiden sabit kodla TEK sayfada zorla
    # birleştiriliyorlardı — gerçek üretim verisinde ortak yöneticiler zaten
    # TEK birleşik metin olarak saklanır, ör. "CÜNEYT ÇIKRIKÇI - AYŞE AVCU").
    assert set(wb.sheetnames) == {'Ali Çelik', 'Derya Yardımcı', 'Cüneyt Çıkrıkçı', 'Ayşe Avcu', 'Ertan Teki'}
    ali = wb['Ali Çelik']
    values = [c.value for row in ali.iter_rows() for c in row if c.value is not None]
    assert 'A MAĞAZASI' in values
    assert 'GERÇEK UNVAN' in values and 'AD SOYAD' in values
    assert 'Ali Test' in values and 'Ayşe Test' in values
    assert 'BOŞ POZİSYON' not in values  # iki mevcut, bir eksik aynı unvanda ilk satırda E=1
    derya = wb['Derya Yardımcı']
    dvals = [c.value for row in derya.iter_rows() for c in row if c.value is not None]
    assert 'Fazla Kişi' in dvals
