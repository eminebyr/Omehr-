from __future__ import annotations

import json
import math
import re
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from scipy import stats
from services.runtime_paths import runtime_root
from services.settings import input_path
from services.version import MODEL_VERSION as _MODEL_VERSION
from services.safe_exec import log_swallowed


def _input(): return input_path(runtime_root())
def _output(): return runtime_root() / "output"
def _ai_file(): return _output() / "V19_AI_Norm_Sonuclari.xlsx"
def _analytics_file(): return _output() / "V19_Istatistik_ML_Operasyon_Analizi.xlsx"
def _json_file(): return _output() / "V19_Model_Karti.json"


def _canon(value) -> str:
    table = str.maketrans("ÇĞİÖŞÜ", "CGIOSU")
    return re.sub(r"[^A-Z0-9]+", "", str(value or "").upper().translate(table))


def _active(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    if "Durum" in result and result["Durum"].notna().any():
        result = result[result["Durum"].astype(str).str.strip().str.casefold().eq("aktif")]
    elif "İşten Çıkış" in result:
        result = result[result["İşten Çıkış"].isna()]
    return result


def _number(frame: pd.DataFrame, *names: str) -> pd.Series:
    wanted = {_canon(name) for name in names}
    for column in frame.columns:
        if _canon(column) in wanted or any(name in _canon(column) for name in wanted):
            return pd.to_numeric(frame[column], errors="coerce")
    return pd.Series(np.nan, index=frame.index, dtype=float)


def _latest(frame: pd.DataFrame) -> pd.DataFrame:
    for column in frame.columns:
        if _canon(column) in {"AY", "DONEM", "TARIH"}:
            parsed = pd.to_datetime(frame[column], errors="coerce")
            if parsed.notna().any():
                return frame.loc[parsed.eq(parsed.max())].copy()
            values = frame[column].dropna().astype(str)
            if not values.empty:
                return frame.loc[frame[column].astype(str).eq(values.max())].copy()
    return frame.copy()


def _operation_features(sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    raw=pd.read_excel(_input(),sheet_name="Aylık Operasyon KPI",header=None)
    best_row=0
    best_score=-1
    wanted={"MAGAZAID","MAGAZA","AYLIKCIRO","AYLIKFIS","ORTSEPET","ONLINESIPARIS","MALKABUL"}
    for row_index in range(min(12,len(raw))):
        score=sum(_canon(value) in wanted for value in raw.iloc[row_index].dropna())
        if score>best_score:
            best_row,best_score=row_index,score
    source = _latest(pd.read_excel(_input(),sheet_name="Aylık Operasyon KPI",header=best_row))
    if source.empty:
        return pd.DataFrame(columns=["MağazaID"])
    result = pd.DataFrame()
    for column in source.columns:
        if _canon(column) == "MAGAZAID":
            result["MağazaID"] = source[column].astype(str).str.strip()
            break
    if "MağazaID" not in result:
        return pd.DataFrame(columns=["MağazaID"])
    mapping = {
        "Aylık Ciro": ("Aylık Ciro", "Ciro"),
        "Aylık Fiş": ("Aylık Fiş", "Fiş"),
        "Ortalama Sepet": ("Ort. Sepet", "Ortalama Sepet"),
        "Online Sipariş": ("Online Sipariş",),
        "Mal Kabul": ("Mal Kabul",),
        "Fazla Mesai": ("Fazla Mesai", "Mesai"),
        "Devamsızlık": ("Devamsızlık",),
        "Fire Oranı": ("Fire Oranı", "Fire"),
        "Performans": ("Performans", "PBC"),
    }
    for target, aliases in mapping.items():
        result[target] = _number(source, *aliases).fillna(0)
    return result.groupby("MağazaID", as_index=False).sum(numeric_only=True)


def _workload_model(sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    activity = sheets["Gunluk_Aktivite_Hacmi"].copy()
    # SAVUNMACI ERİŞİM: "Veri Durumu" sütunu (ör. kullanıcı tarafından
    # silinirse) yoksa KeyError ile çökmek yerine boş kabul edilir —
    # güven skoru hesaplaması etkilenir ama sistem AYAKTA kalır.
    if "Veri Durumu" not in activity.columns:
        activity["Veri Durumu"] = ""
    capacity = sheets["Kapasite_Parametreleri"].copy()
    minimum = sheets["Minimum_Kadro_Kurallari"].copy()
    peak = sheets["Vardiya_Pik_Saat"].copy()
    calibration = sheets["Kalibrasyon"].copy()
    norm = sheets["Fact_Norm"].copy()
    staff = _active(sheets["Fact_Mevcut"])

    # İş Yükü sütunu Excel formülüdür. Excel/LibreOffice yeniden hesaplama
    # yapmadan kaydedilmişse pandas formül önbelleğini NaN okuyabilir. Bu durum
    # gerçek iş yükünü 0'a, bütün güven skorlarını da tabana indiriyordu.
    cached_workload = pd.to_numeric(activity["İş Yükü (Dk)"], errors="coerce")
    amount = pd.to_numeric(activity["Aktivite Miktarı"], errors="coerce")
    standard = pd.to_numeric(activity["Standart Süre (Dk)"], errors="coerce")
    calibration_factor = pd.to_numeric(activity["Kalibrasyon Katsayısı"], errors="coerce").fillna(1)
    derived_workload = amount * standard * calibration_factor
    usable_cached = cached_workload.notna() & cached_workload.ge(0)
    activity["İş Yükü (Dk)"] = cached_workload.where(usable_cached, derived_workload).fillna(0)
    activity["_İş Yükü Kaynağı"] = np.where(usable_cached, "Excel hesaplanmış değer", "Bileşenlerden yeniden hesaplandı")
    workload = (
        activity.groupby(["MağazaID", "Mağaza", "Bölge", "UnvanID", "Unvan"], dropna=False)
        .agg(
            **{
                "Toplam İş Yükü (Dk)": ("İş Yükü (Dk)", "sum"),
                "Aktivite Sayısı": ("AktiviteID", "nunique"),
                "Kaynak Satır": ("AktiviteID", "size"),
                "Gerçek Kaynak Satır": (
                    "Veri Durumu",
                    lambda x: int(x.astype(str).str.contains("Gerçek", case=False, na=False).sum()),
                ),
                "Yeniden Hesaplanan Satır": (
                    "_İş Yükü Kaynağı",
                    lambda x: int(x.astype(str).str.contains("yeniden", case=False, na=False).sum()),
                ),
            }
        )
        .reset_index()
    )
    # SAVUNMACI ERİŞİM: kullanıcı Kapasite_Parametreleri'nden "Veri Durumu"
    # sütununu (ör. "dummy görünmesin" diye) silerse, sistem KeyError ile
    # çökmek yerine bu bilgiyi sessizce eksik sayar — hesaplama akışı bozulmaz.
    if "Veri Durumu" not in capacity.columns:
        capacity["Veri Durumu"] = ""
    capacity = capacity[["UnvanID", "Net Üretken Dakika", "Veri Durumu"]].copy()
    capacity.columns = ["UnvanID", "Net Üretken Dk", "Kapasite Veri Durumu"]
    capacity["Net Üretken Dk"] = pd.to_numeric(capacity["Net Üretken Dk"], errors="coerce")
    workload = workload.merge(capacity, on="UnvanID", how="left")

    minimum = minimum[minimum["Aktif Mi"].astype(str).str.casefold().isin({"evet", "e", "1", "true"})]
    minimum = minimum.groupby("UnvanID", as_index=False)["Minimum Kişi"].max()
    workload = workload.merge(minimum, on="UnvanID", how="left")

    peak["Pik Katsayısı"] = pd.to_numeric(peak["Pik Katsayısı"], errors="coerce")
    peak = peak.groupby("MağazaID", as_index=False)["Pik Katsayısı"].max()
    workload = workload.merge(peak, on="MağazaID", how="left")

    calibration["Güven Skoru"] = pd.to_numeric(calibration["Güven Skoru"], errors="coerce")
    calibration["Gözlem Sayısı"] = pd.to_numeric(calibration["Gözlem Sayısı"], errors="coerce")
    cal = calibration.groupby(["MağazaID", "UnvanID"], as_index=False).agg(
        **{"Kalibrasyon Güveni": ("Güven Skoru", "mean"), "Gözlem Sayısı": ("Gözlem Sayısı", "sum")}
    )
    workload = workload.merge(cal, on=["MağazaID", "UnvanID"], how="left")

    norm["Yönetim Normu"] = pd.to_numeric(norm["Norm Kadro"], errors="coerce").fillna(0)
    norm = norm.groupby(["MağazaID", "UnvanID"], as_index=False)["Yönetim Normu"].sum()
    current = staff.groupby(["MağazaID", "Departman"], dropna=False).size().reset_index(name="Aktif Mevcut")
    current["_Departman"] = current["Departman"].map(_canon)
    title_map = sheets["Fact_Norm"][["UnvanID", "Unvan"]].drop_duplicates()
    title_map["_Departman"] = title_map["Unvan"].map(_canon)
    current = current.merge(title_map[["UnvanID", "_Departman"]], on="_Departman", how="left")
    current = current.groupby(["MağazaID", "UnvanID"], as_index=False)["Aktif Mevcut"].sum()

    result = workload.merge(norm, on=["MağazaID", "UnvanID"], how="outer")
    result = result.merge(current, on=["MağazaID", "UnvanID"], how="outer")
    truth = sheets["Fact_Norm"][["MağazaID", "Mağaza", "Bölge Sorumlusu", "UnvanID", "Unvan"]].drop_duplicates(
        ["MağazaID", "UnvanID"]
    )
    result = result.merge(truth, on=["MağazaID", "UnvanID"], how="left", suffixes=("", "_Gerçek"))
    for column in ["Mağaza", "Unvan"]:
        result[column] = result[f"{column}_Gerçek"].fillna(result.get(column))
    result["Bölge"] = result["Bölge Sorumlusu"].fillna(result.get("Bölge"))
    for column in ["Yönetim Normu", "Aktif Mevcut", "Toplam İş Yükü (Dk)", "Minimum Kişi"]:
        result[column] = pd.to_numeric(result[column], errors="coerce").fillna(0)
    result["Net Üretken Dk"] = pd.to_numeric(result["Net Üretken Dk"], errors="coerce").fillna(405)
    result["Pik Katsayısı"] = pd.to_numeric(result["Pik Katsayısı"], errors="coerce").fillna(1).clip(1, 1.35)
    result["İş Yükü FTE"] = result["Toplam İş Yükü (Dk)"] / result["Net Üretken Dk"].replace(0, np.nan)
    result["İş Yükü FTE"] = result["İş Yükü FTE"].fillna(0)
    result["AI Ham İş Yükü Normu"] = np.maximum(
        result["Minimum Kişi"], np.ceil(result["İş Yükü FTE"] * result["Pik Katsayısı"])
    ).astype(int)
    # İş yükü kaydı olmayan unvanlarda yönetim normu güvenli referanstır.
    no_workload = result["Toplam İş Yükü (Dk)"].le(0)
    result.loc[no_workload, "AI Ham İş Yükü Normu"] = result.loc[no_workload, "Yönetim Normu"].round().astype(int)
    real_share = result["Gerçek Kaynak Satır"].fillna(0) / result["Kaynak Satır"].replace(0, np.nan)
    cal_conf = result["Kalibrasyon Güveni"].fillna(0)
    # Kaynak 0-1 veya 0-100 ölçeğinde gelebilir.
    cal_conf = np.where(cal_conf > 1, cal_conf / 100, cal_conf)
    cal_conf = pd.Series(cal_conf, index=result.index).clip(0, 1)
    sample_conf = np.minimum(1, np.log1p(result["Gözlem Sayısı"].fillna(0)) / np.log(501))
    workload_completeness = result["Toplam İş Yükü (Dk)"].gt(0).astype(float)
    # Güven bir "başarı oranı" değildir: kalibrasyon, örneklem yeterliliği,
    # gerçek veri payı ve hesaplanabilirlik bileşimidir.
    result["Güven Skoru"] = (
        100
        * (
            0.40 * cal_conf
            + 0.25 * sample_conf
            + 0.20 * real_share.fillna(0)
            + 0.15 * workload_completeness
        )
    ).clip(20, 95).round(1)
    result.loc[no_workload, "Güven Skoru"] = (
        100 * (0.40 * cal_conf[no_workload] + 0.25 * sample_conf[no_workload])
    ).clip(20, 65).round(1)
    # Yönetim normu resmi ankrajdır. İş yükü normunun ağırlığı veri güvenine
    # bağlıdır ve en fazla %35 olabilir; böylece tek dönemli/dummy veri ani
    # kadro sıçraması üretmez.
    result["AI İş Yükü Ağırlığı"] = (0.35 * result["Güven Skoru"] / 100).clip(0, 0.35)
    blended = (
        (1 - result["AI İş Yükü Ağırlığı"]) * result["Yönetim Normu"]
        + result["AI İş Yükü Ağırlığı"] * result["AI Ham İş Yükü Normu"]
    )
    result["AI Önerilen Norm"] = np.floor(blended + 0.5).astype(int).clip(lower=0)
    result.loc[no_workload, "AI Önerilen Norm"] = result.loc[no_workload, "Yönetim Normu"].round().astype(int)
    # Son operasyon güvenlik tavanı: AI, tek dönemde yönetim normunu sınırsız
    # büyütemez. Normu olan satırlarda en fazla %20 (yuvarlama nedeniyle
    # küçük normlarda +1, büyük normlarda çoğunlukla +1/+2); normsuz yeni
    # pozisyonda ancak yüksek güven VE ağırlıklı gerçek saha verisi birlikte
    # varsa en fazla 1 kişi önerilir. Dummy/tahmini veri yeni kadro açamaz.
    management_norm = result["Yönetim Normu"].round().astype(int).clip(lower=0)
    result["AI Norm Üst Sınırı"] = np.ceil(management_norm * 1.20).astype(int)
    zero_norm = management_norm.eq(0)
    new_position_evidence = result["Güven Skoru"].ge(85) & real_share.fillna(0).ge(0.70)
    result.loc[zero_norm, "AI Norm Üst Sınırı"] = np.where(
        new_position_evidence.loc[zero_norm], 1, 0
    )
    result["AI Önerilen Norm"] = np.minimum(
        result["AI Önerilen Norm"], result["AI Norm Üst Sınırı"]
    ).astype(int)
    result["AI-Mevcut Fark"] = result["AI Önerilen Norm"] - result["Aktif Mevcut"].round().astype(int)
    result["Kapasite Açığı/Fazlası (Dk)"] = (
        result["AI Önerilen Norm"] - result["Aktif Mevcut"]
    ) * result["Net Üretken Dk"]
    result["Veri Kalitesi Uyarısı"] = np.where(
        result["Güven Skoru"].lt(50),
        "Bu mağazada veri kalitesi düşük, AI önerisi dikkatli kullanılmalı",
        "",
    )
    result["Veri Durumu"] = np.select(
        [real_share.fillna(0).ge(0.70), real_share.fillna(0).ge(0.25)],
        ["Ağırlıklı gerçek veri", "Karma veri"],
        "Dummy/saha etüdü gerekli",
    )
    return result.drop(columns=[c for c in result if c.endswith("_Gerçek")], errors="ignore")


def _statistics(model: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    tests: list[dict] = []
    groups = [
        group["İş Yükü FTE"].dropna().values
        for _, group in model.groupby("Bölge")
        if group["İş Yükü FTE"].notna().sum() >= 2
    ]
    if len(groups) >= 2:
        f_stat, p_value = stats.f_oneway(*groups)
        grand = np.concatenate(groups).mean()
        ss_between = sum(len(group) * (group.mean() - grand) ** 2 for group in groups)
        ss_total = sum(((group - grand) ** 2).sum() for group in groups)
        tests.append(
            {
                "Test": "Tek Yönlü ANOVA",
                "Hipotez": "Bölgelerin ortalama iş yükü FTE değerleri eşittir",
                "İstatistik": f_stat,
                "p-değeri": p_value,
                "Etki Büyüklüğü": ss_between / ss_total if ss_total else 0,
                "Sonuç": "Anlamlı fark var" if p_value < 0.05 else "Anlamlı fark kanıtlanmadı",
            }
        )
        lev, lev_p = stats.levene(*groups, center="median")
        tests.append(
            {
                "Test": "Levene",
                "Hipotez": "Bölge varyansları homojendir",
                "İstatistik": lev,
                "p-değeri": lev_p,
                "Etki Büyüklüğü": np.nan,
                "Sonuç": "Varyanslar farklı" if lev_p < 0.05 else "Homojenlik reddedilmedi",
            }
        )
        kw, kw_p = stats.kruskal(*groups)
        tests.append(
            {
                "Test": "Kruskal-Wallis",
                "Hipotez": "Bölge dağılımları aynıdır",
                "İstatistik": kw,
                "p-değeri": kw_p,
                "Etki Büyüklüğü": np.nan,
                "Sonuç": "Anlamlı fark var" if kw_p < 0.05 else "Anlamlı fark kanıtlanmadı",
            }
        )
    contingency = pd.crosstab(model["Bölge"], np.where(model["AI-Mevcut Fark"] > 0, "AI Açık", "Açık Yok"))
    chi_summary = {}
    if contingency.shape[0] >= 2 and contingency.shape[1] >= 2:
        chi, p_value, dof, expected = stats.chi2_contingency(contingency)
        n = contingency.values.sum()
        cramers_v = math.sqrt(chi / (n * max(min(contingency.shape) - 1, 1)))
        tests.append(
            {
                "Test": "Ki-Kare Bağımsızlık",
                "Hipotez": "Bölge ile AI açık durumu bağımsızdır",
                "İstatistik": chi,
                "p-değeri": p_value,
                "Etki Büyüklüğü": cramers_v,
                "Sonuç": "İlişki var" if p_value < 0.05 else "İlişki kanıtlanmadı",
            }
        )
        chi_summary = {"chi2": chi, "p_value": p_value, "dof": dof, "cramers_v": cramers_v}
    return pd.DataFrame(tests), chi_summary


def _machine_learning(model: pd.DataFrame, operation: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    data = model.merge(operation, on="MağazaID", how="left")
    numeric_features = [
        "Aylık Ciro",
        "Aylık Fiş",
        "Ortalama Sepet",
        "Online Sipariş",
        "Mal Kabul",
        "Fazla Mesai",
        "Devamsızlık",
        "Fire Oranı",
        "Performans",
        "Pik Katsayısı",
        "Net Üretken Dk",
    ]
    for column in numeric_features:
        values=data[column] if column in data else pd.Series(0,index=data.index,dtype=float)
        data[column] = pd.to_numeric(values, errors="coerce").fillna(0)
    target = data["İş Yükü FTE"].astype(float)
    valid = target.notna() & np.isfinite(target)
    data, target = data.loc[valid].reset_index(drop=True), target.loc[valid].reset_index(drop=True)
    if len(data) < 20 or target.nunique() < 3:
        return pd.DataFrame(), pd.DataFrame(), {"status": "SKIPPED", "reason": "Yeterli gözlem yok"}
    try:
        from sklearn.compose import ColumnTransformer
        from sklearn.ensemble import ExtraTreesRegressor, RandomForestRegressor
        from sklearn.impute import SimpleImputer
        from sklearn.linear_model import ElasticNet, Ridge
        from sklearn.metrics import precision_score, recall_score, f1_score, accuracy_score
        from sklearn.model_selection import GroupKFold, KFold, StratifiedGroupKFold, StratifiedKFold, cross_val_predict, cross_validate
        from sklearn.pipeline import Pipeline
        from sklearn.preprocessing import OneHotEncoder, StandardScaler
    except Exception as exc:
        log_swallowed("ai_operations_engine._machine_learning: beklenmeyen hata", exc)
        return pd.DataFrame(), pd.DataFrame(), {"status": "SKIPPED", "reason": str(exc)}

    categorical = ["UnvanID", "Bölge"]
    preprocessor = ColumnTransformer(
        [
            ("num", Pipeline([("impute", SimpleImputer(strategy="median")), ("scale", StandardScaler())]), numeric_features),
            ("cat", OneHotEncoder(handle_unknown="ignore"), categorical),
        ]
    )
    # KARAR: model_benchmark.py'nin geniş kapsamlı GroupKFold karşılaştırmasında
    # (14 regresyon algoritması) Extra Trees, Random Forest'ı MAE'de geride
    # bırakıyor (0.384 vs 0.401) ve aşırı öğrenme riski aynı derecede düşük
    # ("DÜŞÜK — genelleme iyi"). "Tahmin motorları en yüksek doğrulukla
    # çalışmalı" kararına göre üretim aday havuzuna eklenmiştir. KNN ve
    # AdaBoost ise aynı karşılaştırmada belirgin şekilde zayıf/riskli
    # çıktığı için (KNN: yüksek aşırı öğrenme; AdaBoost: sınıflandırmada
    # F1=0.38, neredeyse rastgele) üretim aday havuzuna hiç alınmamıştır.
    candidates = {
        "Ridge Regresyon": Ridge(alpha=1.0),
        "ElasticNet": ElasticNet(alpha=0.02, l1_ratio=0.35, max_iter=20000),
        "Random Forest": RandomForestRegressor(
            n_estimators=350, min_samples_leaf=3, max_features=0.75, random_state=42, n_jobs=-1
        ),
        "Extra Trees": ExtraTreesRegressor(
            n_estimators=350, min_samples_leaf=3, max_features=0.85, random_state=42, n_jobs=-1
        ),
    }
    folds = min(5, max(3, len(data) // 60))
    # DÜZELTME (P1 — reviewer önerisi): Sıradan KFold, AYNI mağazanın farklı
    # unvan satırlarının train/test arasında BÖLÜNMESİNE izin verir — bu,
    # model AYNI mağazanın diğer özelliklerini eğitimde "görüp" test
    # setindeki satırını tahmin ederken avantaj kazanabileceği için VERİ
    # SIZINTISIDIR ve CV skorlarını olduğundan iyi gösterir. GroupKFold,
    # AYNI mağazanın TÜM satırlarını ya tamamen eğitimde ya tamamen testte
    # tutar — bu, model_benchmark.py'nin geniş karşılaştırmasında zaten
    # kullanılan, gerçek genelleme performansını yansıtan doğru yöntemdir.
    grup_sayisi = data["MağazaID"].nunique() if "MağazaID" in data.columns else 0
    if grup_sayisi >= folds:
        cv = GroupKFold(n_splits=folds)
        cv_gruplar = data["MağazaID"]
    else:
        # Çok az benzersiz mağaza varsa (ör. küçük test verisi) GroupKFold
        # kuramaz — bu durumda KFold'a düşer ama bunu AÇIKÇA loglar.
        from services.safe_exec import log_swallowed
        log_swallowed(
            f"_workload_model: GroupKFold için yeterli benzersiz mağaza yok "
            f"({grup_sayisi} < {folds} kat) — sıradan KFold'a düşüldü, sonuçlar iyimser olabilir",
            ValueError("yetersiz mağaza grubu"), level="WARNING",
        )
        cv = KFold(n_splits=folds, shuffle=True, random_state=42)
        cv_gruplar = None
    rows = []
    for name, estimator in candidates.items():
        pipe = Pipeline([("prepare", preprocessor), ("model", estimator)])
        scores = cross_validate(
            pipe,
            data[numeric_features + categorical],
            target,
            cv=cv,
            groups=cv_gruplar,
            scoring={"mae": "neg_mean_absolute_error", "rmse": "neg_root_mean_squared_error", "r2": "r2"},
        )
        rows.append(
            {
                "Model": name,
                "CV MAE": -scores["test_mae"].mean(),
                "CV RMSE": -scores["test_rmse"].mean(),
                "CV R²": scores["test_r2"].mean(),
                "Kat Sayısı": folds,
            }
        )
    comparison = pd.DataFrame(rows).sort_values(["CV MAE", "CV RMSE"]).reset_index(drop=True)
    best_name = comparison.iloc[0]["Model"]
    best_pipe = Pipeline([("prepare", preprocessor), ("model", candidates[best_name])])
    predictions = cross_val_predict(best_pipe, data[numeric_features + categorical], target, cv=cv, groups=cv_gruplar)
    # AŞIRI ÖĞRENME (OVERFITTING) TEŞHİSİ: seçilen en iyi modeli TÜM veriyle
    # eğitip AYNI veri üzerinde ölçmek (eğitim skoru), gerçek genelleme
    # performansını yansıtan mağaza-dışı CV skoruyla karşılaştırılır. Aradaki
    # fark büyükse model ezberlemiş demektir.
    from sklearn.base import clone
    from sklearn.metrics import mean_absolute_error as _mae_fn, r2_score as _r2_fn
    _egitim_pipe = clone(best_pipe).fit(data[numeric_features + categorical], target)
    _egitim_tahmin = _egitim_pipe.predict(data[numeric_features + categorical])
    _egitim_mae = _mae_fn(target, _egitim_tahmin)
    _egitim_r2 = _r2_fn(target, _egitim_tahmin)
    _cv_r2 = comparison.iloc[0]["CV R²"]
    _r2_farki = _egitim_r2 - _cv_r2
    if _r2_farki > 0.15:
        _asiri_ogrenme_durumu = "YÜKSEK RİSK — aşırı öğrenme olası"
    elif _r2_farki > 0.07:
        _asiri_ogrenme_durumu = "ORTA — izlenmeli"
    else:
        _asiri_ogrenme_durumu = "DÜŞÜK — genelleme iyi"
    comparison["Eğitim MAE"] = None
    comparison["Eğitim R²"] = None
    comparison["R² Farkı (Eğitim-CV)"] = None
    comparison["Aşırı Öğrenme Durumu"] = None
    comparison.loc[0, "Eğitim MAE"] = _egitim_mae
    comparison.loc[0, "Eğitim R²"] = _egitim_r2
    comparison.loc[0, "R² Farkı (Eğitim-CV)"] = _r2_farki
    comparison.loc[0, "Aşırı Öğrenme Durumu"] = _asiri_ogrenme_durumu
    data["ML Tahmini İş Yükü FTE"] = np.maximum(predictions, 0)
    data["ML Tahmini Norm"] = np.maximum(
        data["Minimum Kişi"].fillna(0),
        np.ceil(data["ML Tahmini İş Yükü FTE"] * data["Pik Katsayısı"]),
    ).astype(int)

    labels = (data["AI Önerilen Norm"] > data["Aktif Mevcut"]).astype(int)
    class_summary = {"status": "SKIPPED", "reason": "Tek sınıf"}
    if labels.nunique() == 2 and labels.value_counts().min() >= 3:
        class_features = numeric_features + ["Yönetim Normu", "Aktif Mevcut"]
        class_preprocessor = ColumnTransformer(
            [
                ("num", Pipeline([("impute", SimpleImputer(strategy="median")), ("scale", StandardScaler())]), class_features),
                ("cat", OneHotEncoder(handle_unknown="ignore"), categorical),
            ]
        )
        class_folds = min(5, int(labels.value_counts().min()))
        # DÜZELTME (P1): StratifiedKFold sınıf dengesini korur ama AYNI
        # mağazanın satırlarının train/test arasında bölünmesine izin
        # verirdi (veri sızıntısı). StratifiedGroupKFold hem sınıf
        # dengesini HEM mağaza ayrımını aynı anda sağlar.
        class_grup_sayisi = data["MağazaID"].nunique() if "MağazaID" in data.columns else 0
        if class_grup_sayisi >= class_folds:
            class_cv = StratifiedGroupKFold(class_folds, shuffle=True, random_state=42)
            class_cv_gruplar = data["MağazaID"]
        else:
            from services.safe_exec import log_swallowed
            log_swallowed(
                f"_machine_learning (sınıflandırma): StratifiedGroupKFold için yeterli benzersiz "
                f"mağaza yok ({class_grup_sayisi} < {class_folds} kat) — StratifiedKFold'a düşüldü",
                ValueError("yetersiz mağaza grubu"), level="WARNING",
            )
            class_cv = StratifiedKFold(class_folds, shuffle=True, random_state=42)
            class_cv_gruplar = None
        # KARAR: model_benchmark.py'nin sınıflandırma karşılaştırmasında
        # Gradient Boosting (F1=0.951) Random Forest'ı (F1=0.926) geride
        # bırakıyor ve aşırı öğrenme riski düşük çıkıyor. "En yüksek doğruluk"
        # kararına göre üretimde de bu kullanılır.
        from sklearn.ensemble import GradientBoostingClassifier
        classifier = Pipeline(
            [
                ("prepare", class_preprocessor),
                ("model", GradientBoostingClassifier(n_estimators=250, learning_rate=0.035, max_depth=3, random_state=42)),
            ]
        )
        class_pred = cross_val_predict(
            classifier,
            data[class_features + categorical],
            labels,
            cv=class_cv,
            groups=class_cv_gruplar,
        )
        class_summary = {
            "status": "SUCCESS",
            "model": "GradientBoostingClassifier",
            # Model versiyonu, uygulama/kod sürümünden BİLEREK ayrı tutulur —
            # hiperparametreler değiştiğinde bu versiyon da değişmeli, kod
            # sürümü (APP_VERSION) değişmeden model davranışı sessizce
            # değişmemeli. Artık services/version.py'deki TEK kaynaktan
            # okunur (P2 — sürüm standardı merkezileştirmesi).
            "model_version": _MODEL_VERSION,
            "library": "scikit-learn",
            "parameters": {
                "n_estimators": 250,
                "learning_rate": 0.035,
                "max_depth": 3,
                "random_state": 42,
            },
            "precision": precision_score(labels, class_pred, zero_division=0),
            "recall": recall_score(labels, class_pred, zero_division=0),
            "f1": f1_score(labels, class_pred, zero_division=0),
            "accuracy": accuracy_score(labels, class_pred),
            "positive_class": "AI normuna göre açık var",
        }
    return comparison, data, {"status": "SUCCESS", "best_model": best_name, "overfitting_status": _asiri_ogrenme_durumu, "classification": class_summary, "comparison_df": comparison}



def _apply_model_selected_ai_norm(model: pd.DataFrame, ml_data: pd.DataFrame, model_card: dict) -> pd.DataFrame:
    """Seçilen en iyi mağaza-dışı CV modelini güvenli karar zincirine bağlar.

    Resmî Fact_Norm ankrajdır. AI önerisi yalnız gerçek/karma operasyon verisi,
    yeterli güven ve kabul edilebilir GroupKFold başarımı olduğunda yayımlanır.
    Dummy/saha etüdü bekleyen kayıtlar sayısal kadro artışı üretemez.
    """
    result = model.copy()
    result["Formül Tabanlı Aday Norm"] = pd.to_numeric(result["AI Önerilen Norm"], errors="coerce").fillna(result["Yönetim Normu"]).round().astype(int)

    comparison = model_card.get("comparison_df")
    best_name = str(model_card.get("best_model") or "")
    best_mae = float("nan")
    best_r2 = float("nan")
    overfit = str(model_card.get("overfitting_status") or "")
    if isinstance(comparison, pd.DataFrame) and not comparison.empty:
        row = comparison.iloc[0]
        best_mae = pd.to_numeric(row.get("CV MAE"), errors="coerce")
        best_r2 = pd.to_numeric(row.get("CV R²"), errors="coerce")

    model_ok = bool(best_name) and np.isfinite(best_mae) and np.isfinite(best_r2) and best_r2 >= 0.25 and "YÜKSEK" not in overfit.upper()

    if not ml_data.empty and {"MağazaID","UnvanID","ML Tahmini İş Yükü FTE"}.issubset(ml_data.columns):
        preds = ml_data[["MağazaID","UnvanID","ML Tahmini İş Yükü FTE"]].drop_duplicates(["MağazaID","UnvanID"], keep="last")
        result = result.merge(preds, on=["MağazaID","UnvanID"], how="left")
    else:
        result["ML Tahmini İş Yükü FTE"] = np.nan

    confidence = pd.to_numeric(result["Güven Skoru"], errors="coerce").fillna(0).clip(0,100)
    data_state = result.get("Veri Durumu", pd.Series("", index=result.index)).astype(str)
    real_data = data_state.str.contains("Ağırlıklı gerçek", case=False, na=False)
    mixed_data = data_state.str.contains("Karma", case=False, na=False)
    workload_ok = pd.to_numeric(result["Toplam İş Yükü (Dk)"], errors="coerce").fillna(0).gt(0)
    global_gate = result.get("Global Veri Kapısı", pd.Series("KAPALI", index=result.index)).astype(str).eq("AÇIK")
    row_publish = global_gate & model_ok & workload_ok & ((real_data & confidence.ge(60)) | (mixed_data & confidence.ge(75)))

    formula_fte = pd.to_numeric(result["İş Yükü FTE"], errors="coerce").fillna(0)
    ml_fte = pd.to_numeric(result["ML Tahmini İş Yükü FTE"], errors="coerce")
    # Model güveni: R² yükseldikçe ML payı artar; hiçbir zaman formül/FTE denetimini tamamen devralmaz.
    model_share = float(np.clip(best_r2 if np.isfinite(best_r2) else 0, 0, 0.70))
    selected_fte = formula_fte.where(ml_fte.isna(), (1-model_share)*formula_fte + model_share*ml_fte)
    peak = pd.to_numeric(result["Pik Katsayısı"], errors="coerce").fillna(1).clip(1,1.35)
    minimum = pd.to_numeric(result["Minimum Kişi"], errors="coerce").fillna(0)
    workload_candidate = np.maximum(minimum, selected_fte * peak)

    management = pd.to_numeric(result["Yönetim Normu"], errors="coerce").fillna(0).clip(lower=0)
    ai_weight = (0.35 * confidence / 100).clip(0,0.35)
    blended = management*(1-ai_weight) + workload_candidate*ai_weight
    candidate = np.floor(blended + 0.5).clip(lower=0).astype(int)

    # Yönetim normu 0 olan yeni kadro ancak çok güçlü gerçek veriyle açılabilir.
    zero_norm = management.eq(0)
    new_position_ok = zero_norm & real_data & confidence.ge(85) & model_ok
    candidate = pd.Series(candidate, index=result.index)
    candidate.loc[zero_norm & ~new_position_ok] = 0
    candidate.loc[new_position_ok] = candidate.loc[new_position_ok].clip(upper=1)

    result["Seçilen Model FTE"] = selected_fte.round(3)
    result["Model Seçim Durumu"] = np.where(model_ok, f"{best_name} | GroupKFold CV", "Model yayıma uygun değil")
    result["AI Yayın Durumu"] = np.where(row_publish, "YAYINLANDI", "YAYINLANMADI — yönetim normu korundu")
    result["AI Önerilen Norm"] = np.where(row_publish, candidate, management.round().astype(int)).astype(int)
    result["AI-Mevcut Fark"] = result["AI Önerilen Norm"] - pd.to_numeric(result["Aktif Mevcut"], errors="coerce").fillna(0).round().astype(int)
    result["Kapasite Açığı/Fazlası (Dk)"] = (result["AI Önerilen Norm"] - pd.to_numeric(result["Aktif Mevcut"], errors="coerce").fillna(0)) * pd.to_numeric(result["Net Üretken Dk"], errors="coerce").fillna(0)
    result["AI Karar Gerekçesi"] = np.where(
        row_publish,
        "Ciro/fiş/online/operasyon metrikleri + seçilen en iyi model + iş yükü FTE + güven ağırlığı",
        "Standart süre doğrulaması, gerçek veri, güven veya model başarımı yayın eşiğini geçmedi; Fact_Norm korundu",
    )
    return result

def _explanations(model: pd.DataFrame, ml_data: pd.DataFrame, model_card: dict) -> pd.DataFrame:
    result = model.copy()
    if not ml_data.empty:
        predictions = ml_data[["MağazaID", "UnvanID", "ML Tahmini İş Yükü FTE", "ML Tahmini Norm"]]
        result = result.merge(predictions, on=["MağazaID", "UnvanID"], how="left")
    else:
        result["ML Tahmini İş Yükü FTE"] = np.nan
        result["ML Tahmini Norm"] = np.nan
    result["AI Modeli"] = model_card.get("best_model", "İş yükü/FTE formülü")

    def action(row) -> str:
        gap = int(row["AI-Mevcut Fark"])
        warning = "DÜŞÜK VERİ KALİTESİ — " if float(row["Güven Skoru"]) < 50 else ""
        if gap > 0:
            return f"{warning}{gap} kişilik kapasite açığı: önce yakın mağazadan transfer, kalan için işe alım"
        if gap < 0:
            return f"{warning}{abs(gap)} kişi AI normuna göre transfer havuzunda değerlendirilmeli"
        return f"{warning}Mevcut kadro korunmalı; operasyon göstergeleri izlenmeli"

    def explanation(row) -> str:
        confidence = float(row["Güven Skoru"])
        source_note = (
            "Saha verisi güçlü"
            if confidence >= 75
            else ("Karma veri; karar öncesi yönetici kontrolü" if confidence >= 50 else "Dummy veri ağırlıklı; saha zaman etüdü gerekli")
        )
        gap = int(row["AI-Mevcut Fark"])
        direction = f"{gap} kişi açık" if gap > 0 else (f"{abs(gap)} kişi fazla" if gap < 0 else "dengede")
        return (
            f"{row['Mağaza']} / {row['Unvan']}: {row['Toplam İş Yükü (Dk)']:.0f} dk iş yükü, "
            f"{row['Net Üretken Dk']:.0f} dk kişi kapasitesi ve {row['Pik Katsayısı']:.2f} pik katsayısı "
            f"{row['İş Yükü FTE']:.2f} FTE ihtiyacı oluşturdu. Yönetim normu {int(row['Yönetim Normu'])}, "
            f"aktif mevcut {int(row['Aktif Mevcut'])}, ham iş yükü normu {int(row['AI Ham İş Yükü Normu'])}, "
            f"güven ağırlıklı AI önerisi {int(row['AI Önerilen Norm'])}; sonuç {direction}. "
            f"Güven %{confidence:.0f}. {source_note}. "
            f"{row.get('Veri Kalitesi Uyarısı','')}"
        )

    result["Önerilen Aksiyon"] = result.apply(action, axis=1)
    result["Yönetici Açıklaması"] = result.apply(explanation, axis=1)
    result["Öncelik Seviyesi"] = pd.cut(
        result["AI-Mevcut Fark"].clip(lower=0) * result["Güven Skoru"] / 100,
        [-1, 0, 1.5, 3, np.inf],
        labels=["Düşük", "Orta", "Yüksek", "Kritik"],
    ).astype(str)
    ordered = [
        "MağazaID", "Mağaza", "Bölge", "UnvanID", "Unvan", "Yönetim Normu", "Aktif Mevcut",
        "Toplam İş Yükü (Dk)", "Net Üretken Dk", "İş Yükü FTE", "Minimum Kişi", "Pik Katsayısı",
        "AI Ham İş Yükü Normu", "Formül Tabanlı Aday Norm", "Seçilen Model FTE",
        "Doğrulanmış Standart Süre Payı", "Global Veri Kapısı",
        "AI İş Yükü Ağırlığı", "AI Norm Üst Sınırı", "Model Seçim Durumu", "AI Yayın Durumu",
        "AI Önerilen Norm", "AI-Mevcut Fark", "AI Karar Gerekçesi",
        "Kapasite Açığı/Fazlası (Dk)", "Güven Skoru", "Veri Kalitesi Uyarısı",
        "Veri Durumu", "AI Modeli", "ML Tahmini İş Yükü FTE", "ML Tahmini Norm",
        "Öncelik Seviyesi", "Önerilen Aksiyon", "Yönetici Açıklaması",
    ]
    return result[[column for column in ordered if column in result.columns]].rename(columns={"Minimum Kişi": "Minimum Kadro"})


def _format_analytics_workbook(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook=load_workbook(path)
    for sheet in workbook.worksheets:
        sheet.freeze_panes="A2"
        sheet.auto_filter.ref=sheet.dimensions
        for cell in sheet[1]:
            cell.fill=PatternFill("solid",fgColor="102F64")
            cell.font=Font(color="FFFFFF",bold=True)
            cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        sheet.row_dimensions[1].height=32
        for column_cells in sheet.columns:
            letter=column_cells[0].column_letter
            sample=[str(cell.value or "") for cell in list(column_cells)[:150]]
            sheet.column_dimensions[letter].width=min(55,max(11,max(map(len,sample),default=10)+2))
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment=Alignment(vertical="top",wrap_text=cell.column>15)

    if "Model_Karsilastirma" in workbook.sheetnames:
        sheet=workbook["Model_Karsilastirma"]
        chart=BarChart()
        chart.type="bar"
        chart.style=10
        chart.title="Model Karşılaştırması — CV MAE"
        chart.x_axis.title="Ortalama Mutlak Hata"
        chart.y_axis.title="Model"
        chart.add_data(Reference(sheet,min_col=2,min_row=1,max_row=sheet.max_row),titles_from_data=True)
        chart.set_categories(Reference(sheet,min_col=1,min_row=2,max_row=sheet.max_row))
        chart.height=7
        chart.width=14
        sheet.add_chart(chart,"H2")
    if "Siniflandirma_Metrikleri" in workbook.sheetnames:
        sheet=workbook["Siniflandirma_Metrikleri"]
        chart=BarChart()
        chart.style=10
        chart.title="Sınıflandırma Modeli Performansı (Gradient Boosting)"
        chart.y_axis.title="Skor (0–1)"
        chart.add_data(Reference(sheet,min_col=2,min_row=1,max_row=sheet.max_row),titles_from_data=True)
        chart.set_categories(Reference(sheet,min_col=1,min_row=2,max_row=sheet.max_row))
        chart.height=7
        chart.width=14
        sheet.add_chart(chart,"E2")
    workbook.save(path)


def run() -> dict:
    _output().mkdir(parents=True, exist_ok=True)
    sheets = pd.read_excel(_input(), sheet_name=None)
    model = _workload_model(sheets)
    # GLOBAL VERİ KAPISI: standart sürelerin çoğu saha etüdüyle doğrulanmadan
    # model karşılaştırması çalışabilir, fakat kadro önerisi yayımlanamaz.
    std = sheets.get("Standart_Sure_Kutuphanesi", pd.DataFrame())
    if not std.empty and "Kaynak" in std.columns:
        source = std["Kaynak"].astype(str)
        verified = ~source.str.contains("bekleniyor|dummy|test|varsay", case=False, na=False)
        verified_share = float(verified.mean())
    else:
        verified_share = 0.0
    model["Doğrulanmış Standart Süre Payı"] = verified_share
    model["Global Veri Kapısı"] = np.where(verified_share >= 0.70, "AÇIK", "KAPALI")
    operation = _operation_features(sheets)
    tests, chi_summary = _statistics(model)
    comparison, ml_data, model_card = _machine_learning(model, operation)
    model = _apply_model_selected_ai_norm(model, ml_data, model_card)
    result = _explanations(model, ml_data, model_card)

    regression_rows = []
    usable = model[["İş Yükü FTE", "Toplam İş Yükü (Dk)", "Pik Katsayısı", "Net Üretken Dk"]].dropna()
    if len(usable) >= 10:
        y = usable["İş Yükü FTE"].to_numpy(float)
        x = usable[["Toplam İş Yükü (Dk)", "Pik Katsayısı", "Net Üretken Dk"]].to_numpy(float)
        x = np.column_stack([np.ones(len(x)), x])
        beta = np.linalg.lstsq(x, y, rcond=None)[0]
        prediction = x @ beta
        residual = y - prediction
        r2 = 1 - (residual ** 2).sum() / max(((y - y.mean()) ** 2).sum(), 1e-9)
        regression_rows = [
            {
                "Hedef": "İş Yükü FTE",
                "N": len(y),
                "R²": r2,
                "RMSE": math.sqrt(np.mean(residual ** 2)),
                "MAE": np.mean(np.abs(residual)),
                "Sabit": beta[0],
                "İş Yükü Dk Katsayısı": beta[1],
                "Pik Katsayısı Katsayısı": beta[2],
                "Net Üretken Dk Katsayısı": beta[3],
            }
        ]

    class_metrics = model_card.get("classification", {})
    metrics = pd.DataFrame(
        [
            {"Metrik": "Precision", "Değer": class_metrics.get("precision"), "Açıklama": "Açık uyarılarının ne kadarı doğru"},
            {"Metrik": "Recall", "Değer": class_metrics.get("recall"), "Açıklama": "Gerçek açıkların ne kadarını yakalıyor"},
            {"Metrik": "F1", "Değer": class_metrics.get("f1"), "Açıklama": "Precision ve recall dengesi"},
            {"Metrik": "Accuracy", "Değer": class_metrics.get("accuracy"), "Açıklama": "Toplam doğru sınıflandırma oranı"},
        ]
    )
    # CEO raporlarında teknik MağazaID tek başına bırakılmaz. Operasyon tablosuna
    # Dim_Magaza / sonuç tablosundaki mağaza adı eklenir ve MağazaID'nin hemen
    # yanına yerleştirilir.
    if 'MağazaID' in operation.columns:
        store_map=pd.DataFrame()
        dim_store=sheets.get('Dim_Magaza',pd.DataFrame()) if isinstance(sheets,dict) else pd.DataFrame()
        if not dim_store.empty and {'MağazaID','Mağaza'}.issubset(dim_store.columns):
            store_map=dim_store[['MağazaID','Mağaza']].drop_duplicates('MağazaID',keep='last')
        elif {'MağazaID','Mağaza'}.issubset(result.columns):
            store_map=result[['MağazaID','Mağaza']].drop_duplicates('MağazaID',keep='last')
        if 'Mağaza' not in operation.columns:
            operation=operation.merge(store_map,on='MağazaID',how='left') if not store_map.empty else operation.assign(Mağaza='')
        elif not store_map.empty:
            operation=operation.merge(store_map,on='MağazaID',how='left',suffixes=('','_Dim'))
            operation['Mağaza']=operation['Mağaza'].fillna(operation.get('Mağaza_Dim'))
            operation=operation.drop(columns=['Mağaza_Dim'],errors='ignore')
        op_cols=list(operation.columns)
        if 'Mağaza' in op_cols:
            op_cols.remove('Mağaza')
            op_cols.insert(op_cols.index('MağazaID')+1,'Mağaza')
            operation=operation[op_cols]

    with pd.ExcelWriter(_ai_file(), engine="openpyxl") as writer:
        result.to_excel(writer, sheet_name="AI_Norm_Sonuclari", index=False)
    with pd.ExcelWriter(_analytics_file(), engine="openpyxl") as writer:
        tests.to_excel(writer, sheet_name="Hipotez_Testleri", index=False)
        pd.DataFrame(regression_rows).to_excel(writer, sheet_name="Regresyon", index=False)
        comparison.to_excel(writer, sheet_name="Model_Karsilastirma", index=False)
        metrics.to_excel(writer, sheet_name="Siniflandirma_Metrikleri", index=False)
        result.to_excel(writer, sheet_name="AI_Norm_ve_Aksiyon", index=False)
        operation.to_excel(writer, sheet_name="Operasyon_Metrikleri", index=False)
    _format_analytics_workbook(_ai_file())
    _format_analytics_workbook(_analytics_file())
    payload = {
        "version": "V19.0",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "rows": int(len(result)),
        "best_model": model_card.get("best_model"),
        "model_comparison": comparison.replace({np.nan: None}).to_dict("records"),
        "classification": class_metrics,
        "chi_square": chi_summary,
        "official_kpi_source": "Fact_Mevcut + Fact_Norm",
        "ai_decision_source": "İş yükü dakikası + kapasite + pik + minimum kadro + operasyon modeli",
        "warning": "AI önerisi karar desteğidir; yönetim normunu otomatik olarak değiştirmez. Dummy işaretli girdiler saha etüdüyle değiştirilmelidir.",
    }
    _json_file().write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return payload


if __name__ == "__main__":
    print(json.dumps(run(), ensure_ascii=False, indent=2, default=str))
