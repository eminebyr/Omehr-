from __future__ import annotations

import json
import math
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from services.runtime_paths import runtime_root
from services.settings import input_path
from services.model_lifecycle import assess_model_maturity, independent_classification_target, temporal_workload_backtest
from services.safe_exec import log_swallowed

def _output(): return runtime_root() / "output"
def _ai_file(): return _output() / "V19_AI_Norm_Sonuclari.xlsx"
def _analytics_file(): return _output() / "V19_Istatistik_ML_Operasyon_Analizi.xlsx"
def _benchmark_file(): return _output() / "V19_1_Derin_Model_Karsilastirmasi.xlsx"
def _model_card(): return _output() / "V19_1_Derin_Model_Karti.json"


def _dataset() -> tuple[pd.DataFrame, list[str], list[str]]:
    ai = pd.read_excel(_ai_file(), sheet_name="AI_Norm_Sonuclari")
    operation = pd.read_excel(_analytics_file(), sheet_name="Operasyon_Metrikleri")
    data = ai.merge(operation, on="MağazaID", how="left")
    # Her iki kaynakta bulunan açıklama sütunları merge sonrası _x/_y adını
    # alabilir. Model ayrıntı raporunun kullandığı kanonik adları geri kur.
    for column in ("Mağaza", "Unvan"):
        if column not in data.columns:
            left = data.get(f"{column}_x")
            right = data.get(f"{column}_y")
            if left is not None and right is not None:
                data[column] = left.combine_first(right)
            elif left is not None:
                data[column] = left
            elif right is not None:
                data[column] = right
            else:
                data[column] = ""
    numeric = [
        "Aylık Ciro", "Aylık Fiş", "Ortalama Sepet", "Online Sipariş", "Mal Kabul",
        "Fazla Mesai", "Devamsızlık", "Fire Oranı", "Performans", "Pik Katsayısı",
        "Net Üretken Dk", "Yönetim Normu", "Aktif Mevcut",
    ]
    categorical = ["UnvanID", "Bölge"]
    for column in numeric:
        data[column] = pd.to_numeric(data.get(column, 0), errors="coerce").fillna(0)
    data["İş Yükü FTE"] = pd.to_numeric(data["İş Yükü FTE"], errors="coerce")
    data = data[data["İş Yükü FTE"].notna()].reset_index(drop=True)
    data["AI Açık"] = (pd.to_numeric(data["AI-Mevcut Fark"], errors="coerce").fillna(0) > 0).astype(int)
    return data, numeric, categorical


def _bootstrap_interval(values: np.ndarray, statistic=np.mean, repeats: int = 2000) -> tuple[float, float]:
    values = np.asarray(values, dtype=float)
    if len(values) < 2:
        return (float("nan"), float("nan"))
    rng = np.random.default_rng(42)
    estimates = np.empty(repeats)
    for index in range(repeats):
        estimates[index] = statistic(rng.choice(values, size=len(values), replace=True))
    return tuple(np.quantile(estimates, [0.025, 0.975]).tolist())


def _pipelines(numeric: list[str], categorical: list[str]):
    from sklearn.compose import ColumnTransformer
    from sklearn.dummy import DummyClassifier, DummyRegressor
    from sklearn.ensemble import (
        AdaBoostClassifier, AdaBoostRegressor, ExtraTreesClassifier, ExtraTreesRegressor,
        GradientBoostingClassifier, GradientBoostingRegressor, HistGradientBoostingClassifier,
        HistGradientBoostingRegressor, RandomForestClassifier, RandomForestRegressor,
    )
    from sklearn.impute import SimpleImputer
    from sklearn.linear_model import ElasticNet, HuberRegressor, Lasso, LinearRegression, LogisticRegression, Ridge
    from sklearn.neighbors import KNeighborsClassifier, KNeighborsRegressor
    from sklearn.pipeline import Pipeline
    from sklearn.preprocessing import OneHotEncoder, StandardScaler
    from sklearn.svm import SVC, SVR
    from sklearn.tree import DecisionTreeClassifier, DecisionTreeRegressor

    prepare = ColumnTransformer(
        [
            ("numeric", Pipeline([("impute", SimpleImputer(strategy="median")), ("scale", StandardScaler())]), numeric),
            ("categorical", OneHotEncoder(handle_unknown="ignore", sparse_output=False), categorical),
        ],
        sparse_threshold=0,
    )

    def pipeline(model):
        return Pipeline([("prepare", prepare), ("model", model)])

    regression = {
        "Naif Ortalama": pipeline(DummyRegressor(strategy="mean")),
        "Doğrusal Regresyon": pipeline(LinearRegression()),
        "Ridge": pipeline(Ridge(alpha=1.0)),
        "Lasso": pipeline(Lasso(alpha=0.01, max_iter=30000)),
        "ElasticNet": pipeline(ElasticNet(alpha=0.02, l1_ratio=0.35, max_iter=30000)),
        "Huber Robust": pipeline(HuberRegressor(max_iter=3000)),
        "KNN": pipeline(KNeighborsRegressor(n_neighbors=12, weights="distance")),
        "SVR RBF": pipeline(SVR(C=10, epsilon=0.1, gamma="scale")),
        "Karar Ağacı": pipeline(DecisionTreeRegressor(max_depth=8, min_samples_leaf=5, random_state=42)),
        # Cross-validation katları aşağıda kontrollü ve sıralı çalışır.
        # Ağaçların ayrıca n_jobs=-1 ile iç içe süreç açması Railway CPU'sunu
        # aşırı paylaştırıyor ve sklearn/joblib paralellik uyarısını her ağaç
        # görevinde tekrar üretiyordu. Tek kat/tek model düzeyinde n_jobs=1,
        # aynı random_state ile aynı sonucu daha düşük süreç maliyetiyle verir.
        "Random Forest": pipeline(RandomForestRegressor(n_estimators=450, min_samples_leaf=3, max_features=0.75, random_state=42, n_jobs=1)),
        "Extra Trees": pipeline(ExtraTreesRegressor(n_estimators=450, min_samples_leaf=3, max_features=0.85, random_state=42, n_jobs=1)),
        "Gradient Boosting": pipeline(GradientBoostingRegressor(n_estimators=250, learning_rate=0.035, max_depth=3, loss="huber", random_state=42)),
        "HistGradientBoosting": pipeline(HistGradientBoostingRegressor(max_iter=300, learning_rate=0.05, max_leaf_nodes=20, l2_regularization=1.0, random_state=42)),
        "AdaBoost": pipeline(AdaBoostRegressor(n_estimators=250, learning_rate=0.04, loss="square", random_state=42)),
    }
    classification = {
        "Naif Çoğunluk": pipeline(DummyClassifier(strategy="prior")),
        "Lojistik Regresyon": pipeline(LogisticRegression(max_iter=5000, class_weight="balanced")),
        "KNN": pipeline(KNeighborsClassifier(n_neighbors=15, weights="distance")),
        "SVC RBF": pipeline(SVC(C=5, probability=True, class_weight="balanced", random_state=42)),
        "Karar Ağacı": pipeline(DecisionTreeClassifier(max_depth=8, min_samples_leaf=5, class_weight="balanced", random_state=42)),
        "Random Forest": pipeline(RandomForestClassifier(n_estimators=450, min_samples_leaf=3, max_features=0.75, class_weight="balanced", random_state=42, n_jobs=1)),
        "Extra Trees": pipeline(ExtraTreesClassifier(n_estimators=450, min_samples_leaf=3, max_features=0.85, class_weight="balanced", random_state=42, n_jobs=1)),
        "Gradient Boosting": pipeline(GradientBoostingClassifier(n_estimators=250, learning_rate=0.035, max_depth=3, random_state=42)),
        "HistGradientBoosting": pipeline(HistGradientBoostingClassifier(max_iter=300, learning_rate=0.05, max_leaf_nodes=20, l2_regularization=1.0, random_state=42)),
        "AdaBoost": pipeline(AdaBoostClassifier(n_estimators=250, learning_rate=0.04, random_state=42)),
    }
    return regression, classification


def _regression_benchmark(data, features, groups, models, cv):
    from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
    from sklearn.base import clone

    y = data["İş Yükü FTE"].to_numpy(float)
    x = data[features]
    rows, predictions = [], {}
    for name, model in models.items():
        # Önceden cross_validate + cross_val_predict aynı modeli aynı beş
        # katta iki kez eğitiyordu. Tek kontrollü CV döngüsü hem grup-dışı
        # tahminleri hem kat metriklerini üretir; eğitim sayısı 10'dan 5'e iner.
        pred = np.empty(len(data), dtype=float)
        fold_r2 = []
        for train_index, test_index in cv.split(x, y, groups):
            fold_model = clone(model).fit(x.iloc[train_index], y[train_index])
            fold_prediction = fold_model.predict(x.iloc[test_index])
            pred[test_index] = fold_prediction
            fold_r2.append(r2_score(y[test_index], fold_prediction))
        predictions[name] = pred
        abs_error = np.abs(y - pred)
        mae_low, mae_high = _bootstrap_interval(abs_error)
        # AŞIRI ÖĞRENME (OVERFITTING) TEŞHİSİ: Modeli TÜM veriyle eğitip AYNI veri
        # üzerinde ölçmek (eğitim skoru), gerçek genelleme performansını gösteren
        # GroupKFold (mağaza-dışı) skoruyla karşılaştırılır. Eğitim skoru CV
        # skorundan ÇOK yüksekse (büyük "fark"), model eğitim verisini ezberlemiş
        # (overfit) demektir — sadece tek bir yüksek R²'ye bakmak yanıltıcıdır.
        fitted = clone(model).fit(x, y)
        train_pred = fitted.predict(x)
        train_mae = mean_absolute_error(y, train_pred)
        train_r2 = r2_score(y, train_pred)
        cv_mae = mean_absolute_error(y, pred)
        cv_r2 = r2_score(y, pred)
        mae_farki = cv_mae - train_mae
        r2_farki = train_r2 - cv_r2
        if r2_farki > 0.15:
            durum = "YÜKSEK RİSK — aşırı öğrenme olası"
        elif r2_farki > 0.07:
            durum = "ORTA — izlenmeli"
        else:
            durum = "DÜŞÜK — genelleme iyi"
        rows.append(
            {
                "Model": name,
                "GroupKFold MAE": cv_mae,
                "MAE %95 GA Alt": mae_low,
                "MAE %95 GA Üst": mae_high,
                "GroupKFold RMSE": math.sqrt(mean_squared_error(y, pred)),
                "GroupKFold R²": cv_r2,
                "Kat Ort. R²": float(np.mean(fold_r2)),
                "Kat R² Std": float(np.std(fold_r2, ddof=1)),
                "Naif Modele Göre MAE İyileşmesi %": np.nan,
                "Eğitim MAE": train_mae,
                "Eğitim R²": train_r2,
                "MAE Farkı (CV-Eğitim)": mae_farki,
                "R² Farkı (Eğitim-CV)": r2_farki,
                "Aşırı Öğrenme Durumu": durum,
            }
        )
    result = pd.DataFrame(rows)
    baseline = float(result.loc[result["Model"].eq("Naif Ortalama"), "GroupKFold MAE"].iloc[0])
    result["Naif Modele Göre MAE İyileşmesi %"] = (baseline - result["GroupKFold MAE"]) / baseline * 100
    result = result.sort_values(["GroupKFold MAE", "GroupKFold RMSE"]).reset_index(drop=True)
    best = result.loc[~result["Model"].eq("Naif Ortalama")].iloc[0]["Model"]
    detail = data[["MağazaID", "Mağaza", "Bölge", "UnvanID", "Unvan", "İş Yükü FTE"]].copy()
    detail["En İyi Model"] = best
    detail["Grup-Dışı Tahmin"] = predictions[best]
    detail["Artık"] = detail["İş Yükü FTE"] - detail["Grup-Dışı Tahmin"]
    detail["Mutlak Hata"] = detail["Artık"].abs()
    return result, detail, best


def _classification_benchmark(data, features, groups, models, cv):
    from sklearn.metrics import (
        accuracy_score, balanced_accuracy_score, f1_score, precision_score, recall_score, roc_auc_score,
    )
    from sklearn.base import clone

    y = data["AI Açık"].to_numpy(int)
    x = data[features]
    rows = []
    for name, model in models.items():
        # predict ve predict_proba için iki ayrı CV eğitimi yerine her katı
        # bir kez eğitip iki çıktıyı aynı fitted modelden al.
        pred = np.empty(len(data), dtype=int)
        probability = np.full(len(data), np.nan, dtype=float)
        for train_index, test_index in cv.split(x, y, groups):
            fold_model = clone(model).fit(x.iloc[train_index], y[train_index])
            pred[test_index] = fold_model.predict(x.iloc[test_index])
            if hasattr(fold_model, "predict_proba"):
                proba = fold_model.predict_proba(x.iloc[test_index])
                classes = list(fold_model.classes_)
                if 1 in classes:
                    probability[test_index] = proba[:, classes.index(1)]
        try:
            auc = roc_auc_score(y, probability) if np.isfinite(probability).all() else np.nan
        except Exception as _exc:
            log_swallowed("model_benchmark._classification_benchmark: beklenmeyen hata", _exc)
            auc = np.nan
        # AŞIRI ÖĞRENME TEŞHİSİ: aynı mantık — modeli tüm veriyle eğitip aynı
        # veride ölçmek (eğitim F1) ile mağaza-dışı GroupKFold F1 karşılaştırılır.
        cv_f1 = f1_score(y, pred, zero_division=0)
        try:
            fitted = clone(model).fit(x, y)
            train_pred = fitted.predict(x)
            train_f1 = f1_score(y, train_pred, zero_division=0)
        except Exception as _exc:
            log_swallowed("model_benchmark._classification_benchmark: beklenmeyen hata", _exc)
            train_f1 = np.nan
        f1_farki = (train_f1 - cv_f1) if pd.notna(train_f1) else np.nan
        if pd.isna(f1_farki):
            durum = "Hesaplanamadı"
        elif f1_farki > 0.15:
            durum = "YÜKSEK RİSK — aşırı öğrenme olası"
        elif f1_farki > 0.07:
            durum = "ORTA — izlenmeli"
        else:
            durum = "DÜŞÜK — genelleme iyi"
        rows.append(
            {
                "Model": name,
                "Precision": precision_score(y, pred, zero_division=0),
                "Recall": recall_score(y, pred, zero_division=0),
                "F1": cv_f1,
                "Accuracy": accuracy_score(y, pred),
                "Balanced Accuracy": balanced_accuracy_score(y, pred),
                "ROC AUC": auc,
                "Eğitim F1": train_f1,
                "F1 Farkı (Eğitim-CV)": f1_farki,
                "Aşırı Öğrenme Durumu": durum,
                "Pozitif Sınıf": "AI normuna göre kapasite açığı",
            }
        )
    result = pd.DataFrame(rows).sort_values(["F1", "Balanced Accuracy", "Precision"], ascending=False).reset_index(drop=True)
    best = result.loc[~result["Model"].eq("Naif Çoğunluk")].iloc[0]["Model"]
    return result, best


def _importance(data, features, groups, best_model, model):
    from sklearn.inspection import permutation_importance
    from sklearn.model_selection import GroupKFold

    train, test = next(GroupKFold(n_splits=5).split(data[features], data["İş Yükü FTE"], groups))
    model.fit(data.iloc[train][features], data.iloc[train]["İş Yükü FTE"])
    score = permutation_importance(
        model, data.iloc[test][features], data.iloc[test]["İş Yükü FTE"],
        scoring="neg_mean_absolute_error", n_repeats=20, random_state=42, n_jobs=1,
    )
    return pd.DataFrame(
        {"Değişken": features, "Permütasyon Önemi": score.importances_mean, "Önem Std": score.importances_std, "Model": best_model}
    ).sort_values("Permütasyon Önemi", ascending=False)


def _format(path: Path):
    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = load_workbook(path)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="102F64")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.row_dimensions[1].height = 34
        for cells in sheet.columns:
            values = [str(cell.value or "") for cell in list(cells)[:100]]
            sheet.column_dimensions[cells[0].column_letter].width = min(44, max(12, max(map(len, values), default=10) + 2))
    for sheet_name, value_column, title in [
        ("Regresyon_Model_Karsilastirma", 2, "Regresyon Modelleri — Mağaza Dışı MAE"),
        ("Siniflandirma_Karsilastirma", 4, "Sınıflandırma Modelleri — F1"),
        ("Degisken_Onemi", 2, "En İyi Model — Permütasyon Değişken Önemi"),
    ]:
        sheet = workbook[sheet_name]
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = title
        chart.add_data(Reference(sheet, min_col=value_column, min_row=1, max_row=min(sheet.max_row, 16)), titles_from_data=True)
        chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=min(sheet.max_row, 16)))
        chart.height, chart.width = 8, 15
        sheet.add_chart(chart, "K2")
    workbook.save(path)


def run() -> dict:
    from sklearn.model_selection import GroupKFold

    data, numeric, categorical = _dataset()
    source_sheets = pd.read_excel(input_path(runtime_root()), sheet_name=None)
    # AI operasyon adımı aynı motor çalıştırmasında bu backtesti zaten üretir.
    # Railway CPU'sunu iki kez harcamamak için önce o sonucu yeniden kullan.
    try:
        temporal_summary = pd.read_excel(_analytics_file(), sheet_name="Zamansal_Backtest_Ozet")
    except Exception:
        temporal_detail, temporal_summary = temporal_workload_backtest(source_sheets)
    else:
        try:
            temporal_detail = pd.read_excel(_analytics_file(), sheet_name="Zamansal_Backtest_Detay")
        except Exception:
            temporal_detail = pd.DataFrame()
    lifecycle = assess_model_maturity(source_sheets, backtest_summary=temporal_summary)
    features = numeric + categorical
    groups = data["MağazaID"].astype(str)
    unique_groups = groups.nunique()
    folds = min(5, unique_groups)
    cv = GroupKFold(n_splits=folds)
    regression_models, _classification_models = _pipelines(numeric, categorical)
    regression, residuals, best_regression = _regression_benchmark(data, features, groups, regression_models, cv)
    independent_target, classification_reason = independent_classification_target(source_sheets)
    # AI Açık, AI-Mevcut Fark'tan türetildiği için bağımsız hedef değildir.
    # Bu hedefle çok yüksek F1 üretmek yerine sınıflandırma kanıtını, sonradan
    # gerçekleşen/onaylanan ihtiyaç etiketi gelene kadar açıkça bekletiriz.
    classification = pd.DataFrame([{
        "Model": "YAYINLANMADI",
        "Durum": "Bağımsız hedef bulundu" if not independent_target.empty else "Bağımsız hedef bekleniyor",
        "Açıklama": classification_reason,
    }])
    best_classification = None
    importance = _importance(data, features, groups, best_regression, regression_models[best_regression])
    residual_summary = (
        residuals.groupby(["Bölge", "Unvan"], as_index=False)
        .agg(**{"Gözlem": ("Mutlak Hata", "size"), "MAE": ("Mutlak Hata", "mean"), "Bias": ("Artık", "mean")})
        .sort_values("MAE", ascending=False)
    )
    limitations = pd.DataFrame(
        [
            {"Konu": "Test tasarımı", "Değerlendirme": f"{folds} katlı GroupKFold; aynı mağaza eğitim ve testte birlikte bulunmaz.", "Durum": "Güçlü"},
            {"Konu": "Zaman ayrımı", "Değerlendirme": "Birden fazla gerçek dönem bulunmadığı için ileri zaman testi yapılamadı.", "Durum": "Sınırlılık"},
            {"Konu": "Hedef", "Değerlendirme": "İş yükü FTE, aktivite hacmi ve standart sürelerden türetilmiştir.", "Durum": "Türetilmiş hedef"},
            {"Konu": "Veri kalitesi", "Değerlendirme": "Dummy/saha etüdü bekleyen girdiler güven skorunda cezalandırılır.", "Durum": "Açık caveat"},
            {"Konu": "Nedensellik", "Değerlendirme": "Model ilişkisel tahmin üretir; nedensel personel etkisi iddia etmez.", "Durum": "Karar desteği"},
        ]
    )
    with pd.ExcelWriter(_benchmark_file(), engine="openpyxl") as writer:
        regression.to_excel(writer, sheet_name="Regresyon_Model_Karsilastirma", index=False)
        classification.to_excel(writer, sheet_name="Siniflandirma_Karsilastirma", index=False)
        lifecycle.to_frame().to_excel(writer, sheet_name="Model_Yasam_Dongusu", index=False)
        temporal_summary.to_excel(writer, sheet_name="Zamansal_Backtest_Ozet", index=False)
        if not temporal_detail.empty:
            temporal_detail.to_excel(writer, sheet_name="Zamansal_Backtest_Detay", index=False)
        importance.to_excel(writer, sheet_name="Degisken_Onemi", index=False)
        residual_summary.to_excel(writer, sheet_name="Segment_Hata_Analizi", index=False)
        residuals.sort_values("Mutlak Hata", ascending=False).to_excel(writer, sheet_name="Grup_Disi_Tahminler", index=False)
        limitations.to_excel(writer, sheet_name="Model_Sinirliliklari", index=False)
    _format(_benchmark_file())
    best_row = regression.loc[regression["Model"].eq(best_regression)].iloc[0]
    payload = {
        "version": "V19.1",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "validation": "GroupKFold by MağazaID",
        "folds": folds,
        "stores": int(unique_groups),
        "rows": int(len(data)),
        "regression_models_tested": int(len(regression)),
        "classification_models_tested": int(len(classification)),
        "best_regression_model": best_regression,
        "best_regression": best_row.replace({np.nan: None}).to_dict(),
        "best_regression_overfitting_status": best_row.get("Aşırı Öğrenme Durumu"),
        "best_classification_model": best_classification,
        "best_classification": None,
        "best_classification_overfitting_status": None,
        "classification_status": classification_reason,
        "model_lifecycle": lifecycle.to_dict(),
        "temporal_backtest": temporal_summary.replace({np.nan: None}).to_dict("records"),
        "share_status": "Production decision support" if lifecycle.release_allowed else "Experimental / observation only",
        "required_caveat": lifecycle.reason,
    }
    _model_card().write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return payload


if __name__ == "__main__":
    print(json.dumps(run(), ensure_ascii=False, indent=2, default=str))
