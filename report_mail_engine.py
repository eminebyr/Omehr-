from __future__ import annotations

import json
import hashlib
import re
import shutil
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd

from services.region_access import is_global_scope, load_contacts, region_report_paths, safe_text, yes
from services.message_personalization import is_executive_audience, product_label, report_scope_text, salutation
from services.runtime_paths import runtime_root
from services.settings import input_path
from services.safe_exec import log_swallowed

def _input_file():
    return input_path(runtime_root())


def _output_dir():
    return runtime_root() / "output"


def _log_dir():
    return runtime_root() / "logs"


def _log_file():
    return _log_dir() / "OMEHR_Outlook_Gonderim_Log.json"


def _staging_dir():
    return _output_dir() / "Outlook_Hazir"


STAGING_RETENTION_SAAT = 48


def _eski_staging_dosyalarini_temizle() -> int:
    """DÜZELTME: Outlook_Hazir klasöründeki (her gönderim için üretilen,
    SHA-256 doğrulanmış, kaynak raporun tek kullanımlık kopyaları)
    dosyalar için hiçbir temizleme mekanizması YOKTU — zamanla disk
    alanı sınırsız büyüyebilirdi. Artık her gönderim çalıştırmasının
    SONUNDA, STAGING_RETENTION_SAAT'ten (varsayılan 48 saat) eski
    dosyalar otomatik silinir. Aynı gün içindeki yeniden denemeler
    (ör. hata sonrası) etkilenmez; yalnız gerçekten eski, artık
    ihtiyaç duyulmayan kopyalar temizlenir. Silinen dosya sayısını
    döner (loglama/test amaçlı)."""
    if not _staging_dir().is_dir():
        return 0
    esik = datetime.now().timestamp() - STAGING_RETENTION_SAAT * 3600
    silinen = 0
    for dosya in _staging_dir().iterdir():
        try:
            if dosya.is_file() and dosya.stat().st_mtime < esik:
                dosya.unlink()
                silinen += 1
        except Exception as _exc:
            log_swallowed(f"_eski_staging_dosyalarini_temizle: '{dosya}' silinemedi", _exc, level="INFO")
    return silinen



def _ascii_name(value: str) -> str:
    table = str.maketrans({"ç":"c","Ç":"C","ğ":"g","Ğ":"G","ı":"i","İ":"I","ö":"o","Ö":"O","ş":"s","Ş":"S","ü":"u","Ü":"U"})
    text = unicodedata.normalize("NFKD", str(value).translate(table))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", text).strip("_")
    return text or "rapor"


def _validate_attachment(path: Path) -> None:
    if not path.is_file() or path.stat().st_size <= 0:
        raise FileNotFoundError(f"Ek dosyası bulunamadı veya boş: {path}")
    suffix = path.suffix.lower()
    with path.open("rb") as stream:
        header = stream.read(8)
    if suffix == ".pdf" and not header.startswith(b"%PDF-"):
        raise ValueError(f"Geçersiz PDF dosyası: {path.name}")
    if suffix == ".xlsx" and not header.startswith(b"PK"):
        raise ValueError(f"Geçersiz Excel dosyası: {path.name}")
    if suffix == ".pdf":
        from pypdf import PdfReader
        reader=PdfReader(str(path))
        if not reader.pages:
            raise ValueError(f"PDF sayfası bulunamadı: {path.name}")
    if suffix == ".xlsx":
        from openpyxl import load_workbook
        workbook=load_workbook(path,read_only=True,data_only=True)
        if not workbook.sheetnames:
            raise ValueError(f"Excel sayfası bulunamadı: {path.name}")


def _fresh_outlook_copy(path: Path, region: str) -> Path:
    """Kaynak raporu değiştirmeden, Outlook için doğrulanmış yeni bir kopya üretir."""
    _validate_attachment(path)
    _staging_dir().mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    scope = _ascii_name(region or "Genel")
    target = _staging_dir() / f"{scope}_{stamp}_{_ascii_name(path.stem)}{path.suffix.lower()}"
    shutil.copyfile(path, target)
    _validate_attachment(target)
    source_hash=hashlib.sha256(path.read_bytes()).hexdigest()
    target_hash=hashlib.sha256(target.read_bytes()).hexdigest()
    if source_hash != target_hash:
        raise IOError(f"Outlook ek kopyası kaynak dosyayla aynı değil: {path.name}")
    return target

def _file_info(path: Path) -> dict:
    stat=path.stat()
    return {"name":path.name,"full_path":str(path.resolve()),"size_bytes":stat.st_size,
            "modified_at":datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds")}


def _attachments(row: pd.Series) -> list[Path]:
    region=safe_text(row.get("Bölge"))
    send_type=safe_text(row.get("Gönderim Tipi"))
    add_pdf=yes(row.get("PDF Ekle","Evet")); add_excel=yes(row.get("Excel Ekle","Evet"))
    if is_executive_audience(row):
        files=[
            _output_dir()/"OMEHR_Admin_Yonetici_Ozeti.pdf",
            _output_dir()/"OMEHR_Yonetici_Raporu.pdf",
            _output_dir()/"OMEHR_Kutucuklu_Yonetici_Raporu.xlsx",
            _output_dir()/"OMEHR_AI_Karar_Analizi.pdf",
            _output_dir()/"OMEHR_Operasyon_Verimlilik_Analizi.pdf",
            _output_dir()/"OMEHR_Maliyet_Analizi.pdf",
            _output_dir()/"OMEHR_Admin_Norm_ve_Aksiyonlar.xlsx",
            _output_dir()/"OMEHR_Admin_Maliyet_ve_Operasyon.xlsx",
            _output_dir()/"OMEHR_Executive_Data.xlsx",
        ]
    elif is_global_scope(region,send_type):
        files=[]
        if add_pdf: files.append(_output_dir()/"OMEHR_Yonetici_Raporu.pdf")
        # Şirket geneli/yönetici gönderiminde PDF'nin Excel karşılığı zorunludur.
        files.append(_output_dir()/"OMEHR_Kutucuklu_Yonetici_Raporu.xlsx")
        # Global İK kullanıcılarına ana raporlar gider; bölge raporlarının tümü ayrıca eklenmez.
    else:
        files=region_report_paths(region,add_pdf,add_excel)
        # Aktif ve gerçek adresli yönetici için özel bölge dosyası yoksa rapor
        # gönderimini düşürmek yerine şirket geneli resmi norm raporunu kullan.
        if not files:
            if add_pdf: files.append(_output_dir()/"OMEHR_Yonetici_Raporu.pdf")
            files.append(_output_dir()/"OMEHR_Kutucuklu_Yonetici_Raporu.xlsx")
    out=[]; seen=set()
    for p in files:
        p=Path(p).resolve()
        key=str(p).casefold()
        if key not in seen and p.is_file() and p.stat().st_size>0:
            seen.add(key); out.append(p)
    return out


def _body(row: pd.Series, attachments: list[Path]) -> str:
    names="\n".join(f"- {p.name}" for p in attachments)
    content = "karar odaklı yönetici özeti, norm/aksiyon ve maliyet-operasyon" if is_executive_audience(row) else ("güncel norm kadro, AI öneri, aksiyon ve transfer" if "AI " in product_label() else "güncel yönetim normu, kadro ve transfer")
    return (f"{salutation(row)},\n\n{report_scope_text(row).capitalize()} kapsamına ait {content} raporları ekte bilgilerinize sunulmuştur.\n\n"
            f"Rapor tarihi: {datetime.now():%d.%m.%Y %H:%M}\nEk dosya sayısı: {len(attachments)}\n{names}\n\n"
            "Değerlendirmenize arz eder, iyi çalışmalar dileriz.\n\nİnsan Kaynakları Direktörlüğü")


def _write(records: list[dict]) -> None:
    _log_dir().mkdir(parents=True,exist_ok=True)
    _log_file().write_text(json.dumps(records,ensure_ascii=False,indent=2),encoding="utf-8")


def send_reports_via_outlook(input_file: Path | None = None, display_only: bool=False) -> dict:
    if input_file is None:
        input_file = _input_file()
    df=load_contacts(input_file); result={"status":"SKIPPED","sent":0,"failed":0,"log":str(_log_file())}; records=[]
    if df.empty: result["reason"]="Mail_Listesi okunamadı veya boş."; return result
    if "Aktif" not in df.columns: result["reason"]="Mail_Listesi içinde Aktif sütunu yok."; return result
    active=df[df["Aktif"].map(yes)].copy()
    active=active[~active.get("E-posta",pd.Series("",index=active.index)).astype(str).str.contains("dummy.omehr.local",case=False,na=False)]
    if active.empty: result["reason"]="Aktif alıcı yok."; return result
    # DÜZELTME (Madde 30-31 — mail_router.py bağlandı): bu akışta ÖNCEDEN
    # yalnız "Aktif" sütunu kontrol ediliyordu, hiçbir ABONELİK (Norm_Genel/
    # Norm_Bolge vb.) filtresi UYGULANMIYORDU — services/mail_router.py'nin
    # asıl kattığı değer buydu (var ama hiçbir akışa bağlı değildi). Şimdi
    # her satırın kapsamına (şirket geneli vs bölgesel) göre doğru olay
    # türü ile abonelik kontrolü uygulanır. Abonelik sütunu YOKSA (mevcut
    # çoğu kurulumda olduğu gibi) davranış TAMAMEN değişmez.
    from services.mail_router import _apply_subscription_filter
    _sirket_geneli_alicilar = [safe_text(r.get("E-posta")).strip().casefold() for _, r in active.iterrows() if is_global_scope(safe_text(r.get("Bölge")), "")]
    _bolgesel_alicilar = [safe_text(r.get("E-posta")).strip().casefold() for _, r in active.iterrows() if not is_global_scope(safe_text(r.get("Bölge")), "")]
    _abone_kalanlar = set(_apply_subscription_filter(_sirket_geneli_alicilar, df, "COMPANY_NORM_REPORT")) | set(_apply_subscription_filter(_bolgesel_alicilar, df, "REGION_NORM_REPORT"))
    _oncesi_sayi = len(active)
    active = active[active.get("E-posta", pd.Series("", index=active.index)).astype(str).str.strip().str.casefold().isin(_abone_kalanlar)]
    if active.empty: result["reason"]="Abonelik filtresi sonrası aktif alıcı kalmadı."; return result
    if display_only:
        result["reason"]="Taslak önizleme yalnız eski Outlook motorunda desteklenir; güvenli adaptörde gönderim kullanılmalıdır."
        return result
    # TEKRAR GÖNDERİM KORUMASI (idempotency): Her alıcıya, o günün TAM İÇERİĞİYLE
    # (ekler dahil) daha önce başarıyla gönderilip gönderilmediği merkezi
    # services/mail_idempotency.py katmanından kontrol edilir. Aynı gün içinde
    # (ör. hem zamanlanmış görev hem kullanıcının main.py'yi manuel tekrar
    # çalıştırması yüzünden) ikinci bir çağrı olursa, HER ALICI için ayrı ayrı
    # "zaten gönderildi mi" kontrolü yapılır — sadece kaba bir "bugün gönderildi"
    # işaretçisi değil, alıcı+ek-dosya-içeriği bazında gerçek bir garanti.
    from services.mail_idempotency import send_idempotent
    run_id = datetime.now().strftime("%Y-%m-%d")
    for _,row in active.iterrows():
        to=safe_text(row.get("E-posta")); rec={"time":datetime.now().isoformat(timespec="seconds"),"to":to,"region":safe_text(row.get("Bölge"))}
        try:
            if "@" not in to: raise ValueError("Geçerli e-posta adresi yok.")
            source_att=_attachments(row)
            if not source_att: raise FileNotFoundError("Bu kapsam için rapor bulunamadı. Önce ana motoru çalıştırın.")
            att=[_fresh_outlook_copy(path, rec["region"]) for path in source_att]
            transport=send_idempotent(
                "DAILY_REGION_REPORT",
                f"{product_label()} - {rec['region'] or 'Genel'} Raporu - {datetime.now():%d.%m.%Y}",
                _body(row,att),
                [to],
                att,
                run_id=run_id,
            )
            if transport.startswith("SKIPPED"):
                rec.update(status="SKIPPED",transport=transport)
            elif not transport.startswith("SENT"):
                raise RuntimeError(transport)
            else:
                result["sent"]+=1; rec.update(status="SENT",transport=transport,attachments=[_file_info(x) for x in att])
        except Exception as exc:
            log_swallowed("report_mail_engine.send_reports_via_outlook: beklenmeyen hata", exc)
            result["failed"]+=1; rec.update(status="FAILED",error_type=type(exc).__name__,error=str(exc))
        records.append(rec)
    result["status"]="SUCCESS" if result["failed"]==0 else ("PARTIAL" if result["sent"] else "FAILED")
    errors=[
        {"to":r.get("to",""),"error":r.get("error","")}
        for r in records if r.get("status")=="FAILED"
    ]
    if errors:
        result["errors"]=errors
    _write(records)
    _eski_staging_dosyalarini_temizle()
    return result

def send_reports(display_only: bool=False) -> dict:
    return send_reports_via_outlook(_input_file(), display_only=display_only)


if __name__=="__main__":
    import sys
    preview = any(arg in sys.argv for arg in ("--display", "--preview", "--taslak"))
    print(json.dumps(send_reports_via_outlook(display_only=preview),ensure_ascii=False,indent=2))
