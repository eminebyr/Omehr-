from __future__ import annotations

import hashlib
import json
import os
import smtplib
import sqlite3
import urllib.request
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path
from typing import Any

import pandas as pd
from services.message_personalization import is_company_owner, product_label, salutation
from services.security import authenticate as secure_authenticate
from services.outlook_adapter import send_outlook
from services.runtime_paths import runtime_root
from services.settings import input_path
from services.safe_exec import log_swallowed
from services.pdf_compat import make_outlook_safe_pdf

def _data_dir():
    return runtime_root() / "data"


def _output_dir():
    return runtime_root() / "output"


def _db_path():
    return _data_dir() / "v16_management.db"


def _config_path():
    return runtime_root() / "config_web.json"


def _input_path():
    return input_path(runtime_root())


def default_config() -> dict[str, Any]:
    return {
        "company": {"name": "Omehr Marketler", "logo_path": ""},
        "security": {
            "password_storage": "data/security.db",
            "algorithm": "PBKDF2-HMAC-SHA256",
            "iterations": 600000,
            "lock_policy": "5 hatalı giriş / 15 dakika",
        },
        "power_bi": {"enabled": False, "report_url": "", "label": "Power BI Yönetim Raporu"},
        "notifications": {
            "critical_deficit_threshold": 5,
            "teams_webhook_url": "",
            "smtp": {"enabled": False, "host": "", "port": 587, "username": "", "password": "", "from": "", "to": "", "use_tls": True},
        },
        "approval": {"allowed_statuses": ["Beklemede", "Onaylandı", "Reddedildi", "Revizyon İstendi"]},
        "backup": {"max_backups": 20},
    }


def _users_from_input() -> list[dict[str, Any]]:
    """Yalnız kullanıcı yetki metadatasını okur; parola verisini asla döndürmez."""
    if not _input_path().exists():
        return []
    try:
        from services.cached_excel_reader import read_sheet_cached
        df = read_sheet_cached(_input_path(), "Mail_Listesi")
    except Exception as _exc:
        log_swallowed("services.management_center._users_from_input: beklenmeyen hata", _exc)
        return []
    required = {"Web Kullanıcı"}
    if not required.issubset(df.columns):
        return []
    users: list[dict[str, Any]] = []
    for _, row in df.iterrows():
        username = str(row.get("Web Kullanıcı", "") or "").strip()
        active = str(row.get("Aktif", "Evet") or "").strip().casefold()
        if not username or active not in {"evet", "yes", "1", "true"}:
            continue
        region = str(row.get("Bölge", "") or "").strip()
        role = str(row.get("Rol", "") or "").strip().upper()
        scope = str(row.get("Yetki Kapsamı", "") or "").strip()
        # DÜZELTME (çok kiracılı SaaS): "feyzi omehr" adı özel durumu
        # kaldırıldı — ÇOK KİRACILI bir SaaS'ta bu, HER firmanın CEO
        # rolündeki kullanıcısını (kendi gerçek adı ne olursa olsun)
        # yanlışlıkla bu belirli kişi olarak GÖSTERİRDİ. Yalnız genel,
        # kiracıdan bağımsız alan kontrolü (Bölge/Sorumlu=="CEO") kalır.
        is_owner = (
            region.casefold() == "ceo"
            or str(row.get("Sorumlu", "") or "").strip().casefold() == "ceo"
        )
        if is_owner:
            role = "GM"
            scope = "ALL"
        if not role:
            role = "HR_DIRECTOR" if region.upper() in {"TÜMÜ", "TUMU", "ALL"} else "REGION"
        if not scope:
            scope = "ALL" if role in {"HR_DIRECTOR", "GM", "ADMIN", "MANAGEMENT"} else region
        users.append({
            "username": username,
            "name": str(row.get("Sorumlu", username) or username).strip(),
            "role": role,
            "scope": scope,
            "email": str(row.get("E-posta", "") or "").strip(),
            "approval_level": int(float(row.get("Onay Seviyesi", 2 if role in {"HR_DIRECTOR","GM","ADMIN","MANAGEMENT"} else 1) or 1)),
        })
    return users


def ensure_config() -> dict[str, Any]:
    """Yapılandırmayı yükler ve yerel giriş hesaplarını güvenli biçimde yeniler."""
    defaults = default_config()
    _config_path().parent.mkdir(parents=True, exist_ok=True)

    if _config_path().exists():
        try:
            cfg = json.loads(_config_path().read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            cfg = defaults
    else:
        cfg = defaults

    # Mevcut kullanıcı ayarlarını koru; yalnızca eksik alanları güvenli varsayılanlarla tamamla.
    cfg["security"] = defaults["security"]

    cfg.setdefault("power_bi", defaults["power_bi"])
    cfg.setdefault("notifications", defaults["notifications"])
    cfg.setdefault("approval", defaults["approval"])
    cfg.setdefault("company", defaults["company"])
    cfg.setdefault("backup", defaults["backup"])

    _config_path().write_text(
        json.dumps(cfg, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return cfg


def connect() -> sqlite3.Connection:
    _data_dir().mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(_db_path(), timeout=30)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=30000")
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    ensure_config()
    with connect() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS approvals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                proposal_key TEXT NOT NULL,
                scenario TEXT NOT NULL,
                personnel_id TEXT,
                personnel_name TEXT,
                source_store TEXT,
                target_store TEXT,
                current_title TEXT,
                target_title TEXT,
                status TEXT NOT NULL DEFAULT 'Beklemede',
                note TEXT,
                decided_by TEXT,
                decided_at TEXT,
                created_at TEXT NOT NULL,
                UNIQUE(proposal_key, scenario)
            );
            CREATE TABLE IF NOT EXISTS alerts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                alert_key TEXT NOT NULL UNIQUE,
                severity TEXT NOT NULL,
                region TEXT,
                store TEXT,
                title TEXT,
                deficit REAL,
                message TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'Açık',
                created_at TEXT NOT NULL,
                closed_at TEXT
            );
            CREATE TABLE IF NOT EXISTS audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT,
                action TEXT NOT NULL,
                details TEXT,
                created_at TEXT NOT NULL
            );
            -- DEĞİŞTİRİLEMEZ DENETİM KAYDI: bkz. services/web_runtime.py
            -- action_log_no_update/no_delete ile aynı gerekçe.
            CREATE TRIGGER IF NOT EXISTS audit_log_no_update
                BEFORE UPDATE ON audit_log
                BEGIN SELECT RAISE(ABORT, 'audit_log değiştirilemez: UPDATE reddedildi'); END;
            CREATE TRIGGER IF NOT EXISTS audit_log_no_delete
                BEFORE DELETE ON audit_log
                BEGIN SELECT RAISE(ABORT, 'audit_log değiştirilemez: DELETE reddedildi'); END;
            CREATE TABLE IF NOT EXISTS transfer_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_key TEXT NOT NULL UNIQUE,
                personnel_id TEXT NOT NULL,
                personnel_name TEXT,
                source_store TEXT NOT NULL,
                target_store TEXT NOT NULL,
                current_title TEXT,
                target_title TEXT,
                source_region TEXT,
                target_region TEXT,
                requested_by TEXT NOT NULL,
                requested_by_name TEXT,
                requested_at TEXT NOT NULL,
                planned_date TEXT,
                reason TEXT,
                requester_note TEXT,
                hr_status TEXT NOT NULL DEFAULT 'İK Onayı Bekliyor',
                hr_note TEXT,
                hr_decided_by TEXT,
                hr_decided_at TEXT,
                fact_status TEXT NOT NULL DEFAULT 'Bekleniyor',
                fact_store TEXT,
                fact_title TEXT,
                fact_checked_at TEXT,
                completed_at TEXT,
                outlook_sent TEXT NOT NULL DEFAULT 'Hayır',
                outlook_sent_at TEXT,
                outlook_message TEXT,
                updated_at TEXT NOT NULL
            );
            """
        )
        existing = {r[1] for r in conn.execute("PRAGMA table_info(transfer_requests)").fetchall()}
        for col, ddl in {
            "outlook_sent": "TEXT NOT NULL DEFAULT 'Hayır'",
            "outlook_sent_at": "TEXT",
            "outlook_message": "TEXT",
        }.items():
            if col not in existing:
                conn.execute(f"ALTER TABLE transfer_requests ADD COLUMN {col} {ddl}")


def authenticate(username: str, password: str) -> dict[str, Any] | None:
    ok, _, _ = secure_authenticate(username, password)
    if not ok:
        return None
    for user in _users_from_input():
        if user.get("username", "").casefold() == username.strip().casefold():
            log_action(user["username"], "LOGIN", "Başarılı güvenli giriş")
            return {k: user.get(k) for k in ("username", "name", "role", "scope", "approval_level", "email")}
    return None


def log_action(username: str, action: str, details: str = "") -> None:
    init_db()
    with connect() as conn:
        conn.execute(
            "INSERT INTO audit_log(username,action,details,created_at) VALUES(?,?,?,?)",
            (username, action, details, datetime.now().isoformat(timespec="seconds")),
        )


def proposal_key(row: pd.Series | dict[str, Any]) -> str:
    vals = [str(row.get(k, "")) for k in ("PersonelID", "Kaynak Mağaza", "Hedef Mağaza", "İhtiyaç Unvanı")]
    return hashlib.sha256("|".join(vals).encode("utf-8")).hexdigest()[:24]


def upsert_proposals(df: pd.DataFrame, scenario: str) -> None:
    if df.empty:
        return
    init_db()
    now = datetime.now().isoformat(timespec="seconds")
    with connect() as conn:
        for _, row in df.iterrows():
            key = proposal_key(row)
            conn.execute(
                """
                INSERT OR IGNORE INTO approvals(
                    proposal_key,scenario,personnel_id,personnel_name,source_store,target_store,
                    current_title,target_title,status,created_at
                ) VALUES(?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    key, scenario, str(row.get("PersonelID", "")), str(row.get("İsim Soyisim", "")),
                    str(row.get("Kaynak Mağaza", "")), str(row.get("Hedef Mağaza", "")),
                    str(row.get("Mevcut Unvan", "")), str(row.get("İhtiyaç Unvanı", "")), "Beklemede", now,
                ),
            )


def list_approvals(scenario: str | None = None) -> pd.DataFrame:
    init_db()
    query = "SELECT * FROM approvals"
    params: tuple[Any, ...] = ()
    if scenario:
        query += " WHERE scenario=?"
        params = (scenario,)
    query += " ORDER BY created_at DESC, id DESC"
    with connect() as conn:
        return pd.read_sql_query(query, conn, params=params)


def set_approval(proposal_key_value: str, scenario: str, status: str, note: str, username: str) -> None:
    cfg = ensure_config()
    allowed = cfg.get("approval", {}).get("allowed_statuses", [])
    if status not in allowed:
        raise ValueError(f"Geçersiz onay durumu: {status}")
    init_db()
    now = datetime.now().isoformat(timespec="seconds")
    with connect() as conn:
        conn.execute(
            """UPDATE approvals SET status=?,note=?,decided_by=?,decided_at=?
               WHERE proposal_key=? AND scenario=?""",
            (status, note, username, now, proposal_key_value, scenario),
        )
    log_action(username, "TRANSFER_DECISION", f"{scenario}/{proposal_key_value}: {status}")


def scan_alerts(store_df: pd.DataFrame, title_df: pd.DataFrame) -> pd.DataFrame:
    init_db()
    cfg = ensure_config()
    threshold = float(cfg.get("notifications", {}).get("critical_deficit_threshold", 5))
    now = datetime.now().isoformat(timespec="seconds")
    generated: list[dict[str, Any]] = []
    critical = store_df[pd.to_numeric(store_df.get("Norm Eksiği", 0), errors="coerce").fillna(0) >= threshold]
    with connect() as conn:
        for _, row in critical.iterrows():
            region, store, deficit = str(row.get("Bölge Sorumlusu", "")), str(row.get("Mağaza", "")), float(row.get("Norm Eksiği", 0))
            key = hashlib.sha256(f"STORE|{region}|{store}|{deficit}".encode()).hexdigest()[:24]
            msg = f"{store} mağazasında {int(deficit)} kişilik kritik norm açığı bulunuyor."
            conn.execute(
                "INSERT OR IGNORE INTO alerts(alert_key,severity,region,store,title,deficit,message,status,created_at) VALUES(?,?,?,?,?,?,?,?,?)",
                (key, "Kritik", region, store, "", deficit, msg, "Açık", now),
            )
            generated.append({"alert_key": key, "severity": "Kritik", "region": region, "store": store, "deficit": deficit, "message": msg})
    return pd.DataFrame(generated)


def list_alerts(open_only: bool = True) -> pd.DataFrame:
    init_db()
    q = "SELECT * FROM alerts"
    if open_only:
        q += " WHERE status='Açık'"
    q += " ORDER BY deficit DESC, created_at DESC"
    with connect() as conn:
        return pd.read_sql_query(q, conn)


def close_alert(alert_id: int, username: str) -> None:
    init_db()
    with connect() as conn:
        conn.execute("UPDATE alerts SET status='Kapalı',closed_at=? WHERE id=?", (datetime.now().isoformat(timespec="seconds"), alert_id))
    log_action(username, "ALERT_CLOSE", str(alert_id))


def send_teams(message: str) -> tuple[bool, str]:
    cfg = ensure_config()
    url = os.getenv("OMEHR_TEAMS_WEBHOOK") or cfg.get("notifications", {}).get("teams_webhook_url", "")
    if not url:
        return False, "Teams webhook tanımlı değil."
    req = urllib.request.Request(url, data=json.dumps({"text": message}).encode("utf-8"), headers={"Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            return 200 <= resp.status < 300, f"Teams yanıtı: {resp.status}"
    except Exception as exc:
        log_swallowed("services.management_center.send_teams: beklenmeyen hata", exc)
        return False, str(exc)


def send_email(subject: str, body: str) -> tuple[bool, str]:
    cfg = ensure_config().get("notifications", {}).get("smtp", {})
    if not cfg.get("enabled"):
        return False, "SMTP bildirimi etkin değil."
    msg = EmailMessage(); msg["Subject"] = subject; msg["From"] = cfg.get("from"); msg["To"] = cfg.get("to"); msg.set_content(body)
    try:
        with smtplib.SMTP(cfg.get("host"), int(cfg.get("port", 587)), timeout=20) as server:
            if cfg.get("use_tls", True): server.starttls()
            if cfg.get("username"): server.login(cfg.get("username"), cfg.get("password"))
            server.send_message(msg)
        return True, "E-posta gönderildi."
    except Exception as exc:
        log_swallowed("services.management_center.send_email: beklenmeyen hata", exc)
        return False, str(exc)


def simulate(store_df: pd.DataFrame, hires: int = 0, hire_store: str = "", closures: list[str] | None = None) -> dict[str, Any]:
    closures = closures or []
    df = store_df.copy()
    for c in ("Aktif Mevcut", "Norm Kadro", "Norm Eksiği", "Norm Fazlası"):
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    base = {
        "Aktif Mevcut": int(df["Aktif Mevcut"].sum()),
        "Toplam Norm": int(df["Norm Kadro"].sum()),
        "Norm Eksiği": int(df["Norm Eksiği"].sum()),
        "Norm Fazlası": int(df["Norm Fazlası"].sum()),
    }
    sim = df[~df["Mağaza"].astype(str).isin(closures)].copy()
    if hires > 0 and hire_store and hire_store in sim["Mağaza"].astype(str).values:
        idx = sim.index[sim["Mağaza"].astype(str) == hire_store][0]
        sim.loc[idx, "Aktif Mevcut"] += int(hires)
    sim["Norm Eksiği"] = (sim["Norm Kadro"] - sim["Aktif Mevcut"]).clip(lower=0)
    sim["Norm Fazlası"] = (sim["Aktif Mevcut"] - sim["Norm Kadro"]).clip(lower=0)
    # What-if değişikliklerinden sonra eski Excel değerini taşımak yerine
    # Net Fark'ı yeniden hesapla: pozitif değer norm fazlasını, negatif değer açığı gösterir.
    sim["Net Fark"] = sim["Aktif Mevcut"] - sim["Norm Kadro"]
    result = {
        "Aktif Mevcut": int(sim["Aktif Mevcut"].sum()),
        "Toplam Norm": int(sim["Norm Kadro"].sum()),
        "Norm Eksiği": int(sim["Norm Eksiği"].sum()),
        "Norm Fazlası": int(sim["Norm Fazlası"].sum()),
    }
    result["Net İhtiyaç"] = result["Norm Eksiği"] - result["Norm Fazlası"]
    base["Net İhtiyaç"] = base["Norm Eksiği"] - base["Norm Fazlası"]
    return {"base": base, "result": result, "detail": sim}


if __name__ == "__main__":
    init_db()
    print("OMEHR yönetim merkezi veritabanı ve yapılandırması hazır.")

# =========================================================
# WEB TRANSFER TALEP VE FACT_MEVCUT DOĞRULAMA SÜRECİ
# =========================================================

def _transfer_request_key(personnel_id: str, source_store: str, target_store: str, requested_at: str) -> str:
    raw = f"{personnel_id}|{source_store}|{target_store}|{requested_at}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:24]


def create_transfer_request(
    *, personnel_id: str, personnel_name: str, source_store: str, target_store: str,
    current_title: str, target_title: str, source_region: str, target_region: str,
    requested_by: str, requested_by_name: str, planned_date: str = "",
    reason: str = "", requester_note: str = "",
) -> str:
    init_db()
    now = datetime.now().isoformat(timespec="seconds")
    key = _transfer_request_key(personnel_id, source_store, target_store, now)
    with connect() as conn:
        conn.execute(
            """INSERT INTO transfer_requests(
                request_key,personnel_id,personnel_name,source_store,target_store,current_title,target_title,
                source_region,target_region,requested_by,requested_by_name,requested_at,planned_date,reason,
                requester_note,hr_status,fact_status,updated_at
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (key, personnel_id, personnel_name, source_store, target_store, current_title, target_title,
             source_region, target_region, requested_by, requested_by_name, now, planned_date, reason,
             requester_note, "İK Onayı Bekliyor", "Bekleniyor", now),
        )
    log_action(requested_by, "TRANSFER_REQUEST_CREATE", f"{key}: {personnel_name} / {source_store} -> {target_store}")
    ok, msg = _sync_transfer_requests_to_excel()
    log_action(requested_by, "TRANSFER_EXCEL_SYNC" if ok else "TRANSFER_EXCEL_SYNC_FAILED", msg)
    return key


def list_transfer_requests(user: dict[str, Any] | None = None, status: str | None = None) -> pd.DataFrame:
    init_db()
    q = "SELECT * FROM transfer_requests"
    clauses: list[str] = []
    params: list[Any] = []
    if user:
        role = str(user.get("role", "")).upper()
        scope = str(user.get("scope", ""))
        global_user = role in {"HR_DIRECTOR","GM","ADMIN","MANAGEMENT"} and scope.upper() == "ALL"
        if not global_user:
            clauses.append("(source_region=? OR target_region=? OR requested_by=?)")
            params.extend([scope, scope, str(user.get("username", ""))])
    if status:
        clauses.append("hr_status=?")
        params.append(status)
    if clauses:
        q += " WHERE " + " AND ".join(clauses)
    q += " ORDER BY requested_at DESC, id DESC"
    with connect() as conn:
        return pd.read_sql_query(q, conn, params=tuple(params))


def _sync_transfer_requests_to_excel() -> tuple[bool, str]:
    """SQLite taleplerini input içindeki görünür Transfer_Talepleri sayfasına yazar."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        df = list_transfer_requests({"role":"HR_DIRECTOR","scope":"ALL","username":"system"})
        headers = ["Talep No","Talep Tarihi","Kaynak Mağaza","Hedef Mağaza","Personel ID","Personel","Unvan","Hedef Unvan","Bölge Müdürü","Kaynak Bölge","Hedef Bölge","Durum","İK Onayı","İK Notu","Onay Tarihi","Fact_Mevcut Durumu","Fact_Mevcut Kontrol Tarihi","Tamamlanma Tarihi","Outlook Gönderildi mi","Outlook Gönderim Tarihi","Son Güncelleme"]
        keys = ["request_key","requested_at","source_store","target_store","personnel_id","personnel_name","current_title","target_title","requested_by_name","source_region","target_region","hr_status","hr_decided_by","hr_note","hr_decided_at","fact_status","fact_checked_at","completed_at","outlook_sent","outlook_sent_at","updated_at"]
        targets = [_input_path(), runtime_root() / "tenants" / "OMEHR" / "input" / _input_path().name]
        updated = 0
        for target in targets:
            if not target.exists():
                continue
            wb = load_workbook(target)
            ws = wb["Transfer_Talepleri"] if "Transfer_Talepleri" in wb.sheetnames else wb.create_sheet("Transfer_Talepleri")
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row-1)
            for c,h in enumerate(headers,1):
                cell=ws.cell(1,c,h); cell.fill=PatternFill("solid",fgColor="102F64"); cell.font=Font(color="FFFFFF",bold=True); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            for _,r in df.iterrows():
                ws.append([str(r.get(k,"") or "") for k in keys])
            ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:U{max(1,ws.max_row)}"
            wb.save(target); updated += 1
        return True, f"{updated} input dosyası güncellendi."
    except PermissionError:
        return False, "Input Excel açık olduğu için Transfer_Talepleri sayfası yazılamadı. Excel'i kapatıp tekrar deneyin."
    except Exception as exc:
        log_swallowed("services.management_center._sync_transfer_requests_to_excel: beklenmeyen hata", exc)
        return False, str(exc)


def _export_transfer_tracking_pdf() -> Path | None:
    """Tüm transfer taleplerini output klasöründe Türkçe karakterli PDF olarak üretir."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from src.pdf_fonts import font as _omehr_font
        # NOT (FONT_TURKCE_DOGRULAMA.md): burada eskiden ayrı, Windows'a özgü
        # bir font arama zinciri vardı (WINDIR/Fonts, matplotlib paketi,
        # /usr/share/fonts...) ve TÜRKÇE GLİF DOĞRULAMASI yapmıyordu — bulduğu
        # ilk dosyayı doğrulamadan kaydedip kullanıyordu. Artık engine_core'un
        # TEK Türkçe-doğrulamalı font kayıt noktası (src/pdf_fonts.font)
        # kullanılıyor: font paket dışı bir kaynağa asla düşmez, gerekli
        # Türkçe glifler (ÇĞİÖŞÜçğıöşü₺) yoksa PDF üretimi RuntimeError ile
        # açıkça durur (sessizce None dönüp kullanıcıyı yanıltmaz).
        font_regular = _omehr_font(bold=False)
        font_bold = _omehr_font(bold=True)
        _output_dir().mkdir(parents=True, exist_ok=True)
        out = _output_dir() / "OMEHR_Transfer_Takip_Raporu.pdf"
        df = list_transfer_requests({"role":"HR_DIRECTOR","scope":"ALL","username":"system"})
        doc = SimpleDocTemplate(str(out), pagesize=landscape(A4), leftMargin=8*mm, rightMargin=8*mm, topMargin=8*mm, bottomMargin=8*mm)
        head = ParagraphStyle("head", fontName=font_bold, fontSize=6.2, leading=7, textColor=colors.white, leftIndent=0, rightIndent=0, firstLineIndent=0, alignment=1)
        body = ParagraphStyle("body", fontName=font_regular, fontSize=5.8, leading=7, leftIndent=0, rightIndent=0, firstLineIndent=0)
        title = ParagraphStyle("title", fontName=font_bold, fontSize=14, textColor=colors.HexColor("#102F64"), alignment=1)
        story=[Paragraph("OMEHR TRANSFER TALEPLERİ TAKİP RAPORU", title), Spacer(1,3*mm)]
        cols=["Talep Tarihi","Talep Eden","Bölge","Personel","Kaynak","Hedef","İK Durumu","Fact_Mevcut Durumu","İK Notu"]
        keys=["requested_at","requested_by_name","source_region","personnel_name","source_store","target_store","hr_status","fact_status","hr_note"]
        rows=[[Paragraph(c,head) for c in cols]]
        if df.empty:
            rows.append([Paragraph("Henüz transfer talebi bulunmuyor.",body)]+[""]*(len(cols)-1))
        else:
            for _,r in df.iterrows(): rows.append([Paragraph(str(r.get(k,"") or ""),body) for k in keys])
        t=Table(rows,colWidths=[24*mm,28*mm,30*mm,32*mm,30*mm,30*mm,31*mm,40*mm,48*mm],repeatRows=1,hAlign="LEFT")
        t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#102F64")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),font_bold),("GRID",(0,0),(-1,-1),.35,colors.HexColor("#A6A6A6")),("VALIGN",(0,0),(-1,-1),"TOP"),("LEFTPADDING",(0,0),(-1,-1),3),("RIGHTPADDING",(0,0),(-1,-1),3),("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4)]))
        story.append(t)
        doc.build(story)
        make_outlook_safe_pdf(out)
        return out
    except Exception as exc:
        log_swallowed("services.management_center._export_transfer_tracking_pdf: beklenmeyen hata", exc)
        log_action("system", "TRANSFER_PDF_FAILED", str(exc))
        return None


def _transfer_notification_recipients(
    contacts: pd.DataFrame,
    source_region: str,
    target_region: str,
) -> list[str]:
    selected = _transfer_notification_contacts(contacts, source_region, target_region)
    return selected["E-posta"].astype(str).tolist() if not selected.empty else []


def _transfer_notification_contacts(
    contacts: pd.DataFrame,
    source_region: str,
    target_region: str,
) -> pd.DataFrame:
    """Transfer alıcı satırlarını kişiselleştirme bilgileriyle, şirket sahibi hariç döndürür."""
    if contacts.empty or "E-posta" not in contacts.columns:
        return pd.DataFrame(columns=contacts.columns)

    active_values = contacts.get("Aktif", pd.Series("Evet", index=contacts.index))
    active = contacts[active_values.astype(str).str.strip().str.casefold().isin({"evet", "yes", "1", "true"})].copy()
    source_region = str(source_region or "").strip()
    target_region = str(target_region or "").strip()
    management_roles = {"HR_DIRECTOR", "GM", "ADMIN", "MANAGEMENT"}
    selected_indices: list[Any] = []
    for _,r in active.iterrows():
        role=str(r.get("Rol","") or "").strip().upper()
        scope=str(r.get("Yetki Kapsamı",r.get("Bölge","")) or "").strip()
        region=str(r.get("Bölge","") or "").strip()
        email=str(r.get("E-posta","") or "").strip()
        is_management = (
            role in management_roles
            or scope.upper() in {"TÜMÜ","TUMU","ALL"}
        )
        is_related_region = (
            role == "REGION"
            and any(
                candidate
                and (
                    scope.casefold() == candidate.casefold()
                    or region.casefold() == candidate.casefold()
                )
                for candidate in (source_region, target_region)
            )
        )
        if "@" in email and (is_management or is_related_region) and not is_company_owner(r):
            selected_indices.append(r.name)
    if not selected_indices:
        return active.iloc[0:0].copy()
    selected=active.loc[selected_indices].copy()
    selected["_email_key"]=selected["E-posta"].astype(str).str.strip().str.casefold()
    return selected.drop_duplicates("_email_key",keep="first").drop(columns="_email_key")


def _send_transfer_outlook_notification(request_key: str, status: str, note: str) -> tuple[bool,str]:
    """İK kararı sonrası platforma uygun e-posta adaptörüyle bildirim gönderir."""
    try:
        with connect() as conn:
            req = conn.execute("SELECT * FROM transfer_requests WHERE request_key=?", (request_key,)).fetchone()
        if req is None: return False, "Talep bulunamadı."
        from services.cached_excel_reader import read_sheet_cached
        contacts = read_sheet_cached(_input_path(), "Mail_Listesi")
        source_region = str(req["source_region"] or "").strip()
        target_region = str(req["target_region"] or "").strip()
        recipient_contacts = _transfer_notification_contacts(
            contacts,
            source_region,
            target_region,
        )
        if recipient_contacts.empty: return False, "Uygun e-posta alıcısı bulunamadı."
        report=_export_transfer_tracking_pdf()
        main_pdf=_output_dir()/"OMEHR_Yonetici_Raporu.pdf"
        sent=0
        failures=[]
        for _,recipient in recipient_contacts.iterrows():
            try:
                fact_text='Fact_Mevcut güncellemesi bekleniyor' if status=='İK Onayladı' else 'İşlem beklemiyor'
                body=(
                    f"{salutation(recipient)},\n\n"
                    f"Aşağıdaki transfer talebine ilişkin İnsan Kaynakları kararı bilgilerinize sunulmuştur.\n\n"
                    f"Karar: {status}\n"
                    f"Personel: {req['personnel_name']}\n"
                    f"Mevcut unvan: {req['current_title']}\n"
                    f"Hedef unvan: {req['target_title']}\n"
                    f"Kaynak mağaza: {req['source_store']}\n"
                    f"Hedef mağaza: {req['target_store']}\n"
                    f"Planlanan transfer tarihi: {req['planned_date']}\n"
                    f"İK açıklaması: {note or '-'}\n"
                    f"Sistem durumu: {fact_text}\n\n"
                    "Bilgilerinize sunar, gereğini rica ederiz.\n\n"
                    "İnsan Kaynakları Direktörlüğü"
                )
                attachments=[p for p in (report,main_pdf) if p and p.exists()]
                result=send_outlook(
                    f"{product_label()} - Transfer Kararı - {req['personnel_name']} - {status}",
                    body,
                    [str(recipient.get("E-posta","")).strip()],
                    attachments,
                )
                if result.startswith("SENT"): sent+=1
                else: failures.append(f"{recipient.get('E-posta','')}: {result}")
            except Exception as recipient_exc:
                log_swallowed("services.management_center._send_transfer_outlook_notification: beklenmeyen hata", recipient_exc)
                failures.append(f"{recipient.get('E-posta','')}: {recipient_exc}")
        if failures:
            return False, f"{sent} alıcıya gönderildi; {len(failures)} gönderim başarısız: {' | '.join(failures)}"
        return True, f"Outlook bildirimi kişiselleştirilerek {sent} alıcıya gönderildi."
    except Exception as exc:
        log_swallowed("services.management_center._send_transfer_outlook_notification: beklenmeyen hata", exc)
        return False, str(exc)


def decide_transfer_request(request_key: str, status: str, note: str, username: str) -> None:
    allowed = {"İK Onayı Bekliyor", "İK İncelemesinde", "İK Onayladı", "Reddedildi", "Revizyon İstendi", "İptal Edildi"}
    if status not in allowed:
        raise ValueError(f"Geçersiz transfer durumu: {status}")
    init_db()
    now = datetime.now().isoformat(timespec="seconds")
    fact_status = "Fact_Mevcut Güncellemesi Bekleniyor" if status == "İK Onayladı" else "Bekleniyor"
    with connect() as conn:
        conn.execute(
            """UPDATE transfer_requests SET hr_status=?,hr_note=?,hr_decided_by=?,hr_decided_at=?,
               fact_status=?,updated_at=? WHERE request_key=?""",
            (status, note, username, now, fact_status, now, request_key),
        )
    log_action(username, "TRANSFER_REQUEST_DECISION", f"{request_key}: {status}")
    _export_transfer_tracking_pdf()
    if status in {"İK Onayladı", "Reddedildi", "Revizyon İstendi"}:
        ok, message = _send_transfer_outlook_notification(request_key, status, note)
        sent_at = datetime.now().isoformat(timespec="seconds") if ok else None
        with connect() as conn:
            conn.execute("UPDATE transfer_requests SET outlook_sent=?,outlook_sent_at=?,outlook_message=?,updated_at=? WHERE request_key=?", ("Evet" if ok else "Hayır", sent_at, message, datetime.now().isoformat(timespec="seconds"), request_key))
        log_action(username, "TRANSFER_OUTLOOK_SENT" if ok else "TRANSFER_OUTLOOK_FAILED", message)
    sync_ok, sync_msg = _sync_transfer_requests_to_excel()
    log_action(username, "TRANSFER_EXCEL_SYNC" if sync_ok else "TRANSFER_EXCEL_SYNC_FAILED", sync_msg)


def reconcile_transfer_requests(fact_mevcut: pd.DataFrame) -> dict[str, int]:
    """İK onaylı transferleri (transfers tablosu) Fact_Mevcut ile karşılaştırır;
    hedefe henüz taşınmamışsa personnel_exit.update_personnel ile OTOMATİK uygular.

    DÜZELTME (2026-08-27 — kritik şema hatası): Bu fonksiyon önceden web
    panelin FİİLEN KULLANDIĞI 'transfers' tablosunu değil, hiçbir yerden
    INSERT edilmeyen eski/ölü 'transfer_requests' tablosunu sorguluyordu —
    yani panel üzerinden verilen hiçbir İK onayı bu fonksiyon tarafından
    asla görülemezdi. Ayrıca fonksiyonun hiçbir çağıranı yoktu (bkz.
    web/tab_modules/onaylar.py). Artık: (1) doğru tablo ('transfers')
    sorgulanıyor, (2) kimlik eşleştirmesi PersonelID varsa onunla, yoksa
    İsim Soyisim ile yapılıyor, (3) hedefe taşınmamışsa personnel_exit.
    update_personnel ile Fact_Mevcut'taki Mağaza/Unvan alanı OTOMATİK
    hedefe taşınıyor (kilitli + denetim izli yazma yolu — bkz.
    services/personnel_exit.py:update_personnel), (4) bu fonksiyon artık
    web/tab_modules/onaylar.py içinden İK onayı kaydedildikten hemen sonra
    çağrılıyor.
    """
    if fact_mevcut is None or fact_mevcut.empty:
        return {"checked": 0, "completed": 0, "waiting": 0, "mismatch": 0, "applied": 0, "failed": 0}

    def canon_col(*names: str) -> str | None:
        mapping = {str(c).strip().casefold().replace("ı", "i"): c for c in fact_mevcut.columns}
        for name in names:
            key = name.strip().casefold().replace("ı", "i")
            if key in mapping:
                return mapping[key]
        return None

    pid_col = canon_col("PersonelID", "Personel ID", "Sicil No", "Sicil")
    name_col = canon_col("İsim Soyisim", "Isim Soyisim", "Ad Soyad")
    store_col = canon_col("Mağaza", "Magaza")
    title_col = canon_col("Unvan")
    if not store_col or (not pid_col and not name_col):
        return {"checked": 0, "completed": 0, "waiting": 0, "mismatch": 0, "applied": 0, "failed": 0}

    current_by_pid: dict[str, tuple[str, str]] = {}
    current_by_name: dict[str, tuple[str, str]] = {}
    for _, row in fact_mevcut.iterrows():
        store_val = str(row.get(store_col, "")).strip()
        title_val = str(row.get(title_col, "")).strip() if title_col else ""
        if pid_col:
            pid_key = str(row.get(pid_col, "")).strip().casefold()
            if pid_key:
                current_by_pid[pid_key] = (store_val, title_val)
        if name_col:
            name_key = str(row.get(name_col, "")).strip().casefold()
            if name_key:
                current_by_name[name_key] = (store_val, title_val)

    counts = {"checked": 0, "completed": 0, "waiting": 0, "mismatch": 0, "applied": 0, "failed": 0}
    now = datetime.now().isoformat(timespec="seconds")
    with connect() as conn:
        pending = conn.execute(
            "SELECT * FROM transfers WHERE status='İK Onayladı' "
            "AND (fact_status='Fact_Mevcut Güncellemesi Bekleniyor' OR fact_status IS NULL)"
        ).fetchall()
        for req in pending:
            counts["checked"] += 1
            pid = str(req["person_id"] or "").strip()
            pname = str(req["person_name"] or "").strip()
            target_store = str(req["target_store"] or "")
            target_title = str(req["target_title"] or "")

            actual_store, actual_title = "", ""
            if pid and pid.casefold() in current_by_pid:
                actual_store, actual_title = current_by_pid[pid.casefold()]
            elif pname and pname.casefold() in current_by_name:
                actual_store, actual_title = current_by_name[pname.casefold()]

            if actual_store and actual_store.casefold() == target_store.casefold() and (
                not target_title or not actual_title or actual_title.casefold() == target_title.casefold()
            ):
                # Zaten hedefte (elle taşınmış ya da daha önceki bir çalıştırmada uygulanmış).
                status = "Tamamlandı"
                completed_at = now
                counts["completed"] += 1
            elif actual_store and actual_store.casefold() != str(req["source_store"] or "").casefold():
                # Beklenmedik biçimde ne kaynakta ne hedefte; elle kontrol gerektirir.
                status = "Fact_Mevcut ile Uyumsuz"
                completed_at = None
                counts["mismatch"] += 1
            else:
                # Hâlâ kaynak mağazada: OTOMATİK UYGULA — personeli Fact_Mevcut'ta
                # hedef mağaza/unvana taşı (personnel_exit.update_personnel ile,
                # add_personnel/update_personnel'in kullandığı aynı kilitli,
                # denetim-izli yazma yolunu kullanarak).
                try:
                    from services.personnel_exit import load_personnel_view, update_personnel
                    staff, _, _, _ = load_personnel_view(_input_path())
                    eslesen = staff.iloc[0:0]
                    if pname:
                        _isim_norm = pname.strip().casefold()
                        eslesen = staff[staff["İsim Soyisim"].astype(str).str.strip().str.casefold() == _isim_norm]
                    if len(eslesen) == 1:
                        idx = eslesen.index[0]
                        guncellemeler = {"Mağaza": target_store}
                        if target_title:
                            guncellemeler["Unvan"] = target_title
                        update_personnel(
                            input_path=_input_path(), root=runtime_root(), staff=staff,
                            index=idx, guncellemeler=guncellemeler, username="system_rotasyon",
                        )
                        status = "Tamamlandı"
                        completed_at = now
                        counts["completed"] += 1
                        counts["applied"] += 1
                    else:
                        # Belirsiz eşleşme (0 veya birden fazla aynı isim) — güvenlik
                        # nedeniyle otomatik uygulama atlanır, İK'nın elle çözmesi gerekir.
                        status = "Fact_Mevcut Güncellemesi Bekleniyor"
                        completed_at = None
                        counts["waiting"] += 1
                        counts["failed"] += 1
                        log_swallowed(
                            f"reconcile_transfer_requests: '{pname}' için Fact_Mevcut'ta {len(eslesen)} eşleşme bulundu (1 bekleniyordu), otomatik uygulama atlandı.",
                            ValueError(f"{len(eslesen)} eşleşme bulundu, 1 bekleniyordu"), level="WARNING",
                        )
                except Exception as exc:
                    status = "Fact_Mevcut Güncellemesi Bekleniyor"
                    completed_at = None
                    counts["waiting"] += 1
                    counts["failed"] += 1
                    log_swallowed(f"reconcile_transfer_requests: '{pname}' otomatik uygulama hatası", exc, level="ERROR")

            conn.execute(
                "UPDATE transfers SET fact_status=?,completed_at=?,updated_at=? WHERE id=? AND status='İK Onayladı'",
                (status, completed_at, now, int(req["id"])),
            )
        conn.commit()
    return counts
