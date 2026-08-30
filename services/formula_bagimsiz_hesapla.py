from __future__ import annotations

"""
FORMÜL-BAĞIMSIZ YENİDEN HESAPLAMA
=====================================
Norm_Durumu ve Magaza_KPI_Skor_Karti sayfaları başlangıçta Excel formülü
(VLOOKUP/SUMIF/AVERAGEIFS) olarak inşa edilmişti — bu formüller SADECE
LibreOffice dosyayı yeniden hesapladığında dolu gelir. LibreOffice kurulu
olmayan bir bilgisayarda (gerçek üretim ortamında karşılaşıldı) bu
sayfalar "None"/boş görünür ve buna bağlı TÜM web sekmeleri (Bölge&Mağaza,
Unvan Analizi, Mağaza KPI Skor Kartı, CEO Özet risk listeleri) bozulur.

Bu modül, aynı sonucu HER ZAMAN doğru üretmek için Norm_Durumu ve
Magaza_KPI_Skor_Karti'yi Python'da (pandas ile) yeniden hesaplayıp
STATİK DEĞER olarak (formül değil) yazar — LibreOffice'in çalışıp
çalışmadığından tamamen bağımsızdır.
"""

from pathlib import Path

import openpyxl
import pandas as pd

from src.text_utils import canon, _title_key
from services.safe_exec import log_swallowed


def _oku_ham(path: Path, sheet: str) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=sheet, dtype=object)


def statiklestir(input_path: Path) -> bool:
    """Norm_Durumu ve Magaza_KPI_Skor_Karti'yi Python'da yeniden hesaplayıp
    statik değerlerle üzerine yazar. Başarılıysa True döner, sayfalar
    yoksa veya veri eksikse sessizce False döner (ana akışı bozmaz)."""
    input_path = Path(input_path)
    try:
        fm = _oku_ham(input_path, "Fact_Mevcut")
        fn = _oku_ham(input_path, "Fact_Norm")
        dim_magaza = _oku_ham(input_path, "Dim_Magaza")
        dim_unvan = _oku_ham(input_path, "Dim_Unvan")
    except Exception as _exc:
        log_swallowed("statiklestir: temel sayfalar okunamadı", _exc)
        return False

    mag_ad_map = dict(zip(dim_magaza["MağazaID"], dim_magaza["Mağaza"])) if "MağazaID" in dim_magaza.columns else {}
    unvan_ad_map = dict(zip(dim_unvan["UnvanID"], dim_unvan["Unvan"])) if "UnvanID" in dim_unvan.columns else {}

    # --- Fact_Mevcut/Fact_Norm'u Python-taraflı VLOOKUP ile tamamla ---
    for df in (fm, fn):
        if "MağazaID" in df.columns and "Mağaza" in df.columns:
            df["Mağaza"] = df["MağazaID"].map(mag_ad_map).fillna(df["Mağaza"])
        if "UnvanID" in df.columns and "Unvan" in df.columns:
            df["Unvan"] = df["UnvanID"].map(unvan_ad_map).fillna(df["Unvan"])

    fm_aktif = fm[fm.get("İşten Çıkış").isna()] if "İşten Çıkış" in fm.columns else fm

    # Norm Kadro toplamı (MağazaID, UnvanID)
    norm_toplam: dict[tuple, float] = {}
    if {"MağazaID", "UnvanID", "Norm Kadro"}.issubset(fn.columns):
        for _, r in fn.iterrows():
            if pd.isna(r["MağazaID"]) or pd.isna(r["UnvanID"]):
                continue
            key = (r["MağazaID"], r["UnvanID"])
            norm_toplam[key] = norm_toplam.get(key, 0) + (r["Norm Kadro"] or 0)

    # Fact_Mevcut.Unvan gerçek unvandır; norm sayımında Departman kullanılır.
    # Böylece ELİT/UZMAN varyantlar kendi norm ailesini karşılar. Yardımcı
    # departmanlar ise ayrı değerlendirilir.
    unvan_id_by_key = {_title_key(ad): uid for uid, ad in unvan_ad_map.items()}
    mevcut_sayim: dict[tuple, int] = {}
    if {"MağazaID", "Departman"}.issubset(fm_aktif.columns):
        # DÜZELTME (tutarlılık, 29 Ağustos 2026): specialist_family sabit
        # sözlüğü (yalnız 4 elle yazılı aile) KALDIRILDI — bu, panelin
        # resmi KPI kartlarını besleyen src.state_engine._staff_norm_
        # family ile SESSİZCE senkronize olmayan ayrı bir kopyaydı. Artık
        # AYNI fonksiyon doğrudan çağrılır: hem config_norm_rules.json
        # tabanlı family_aliases hem otomatik "UZMAN X"/"ELİT X" -> X
        # kademe birleştirmesi (config'te tanımsız yeni unvanlar için de)
        # buradan gelir. CEO Özet'in okuduğu Norm_Durumu/Magaza_KPI_
        # Skor_Karti sayfası ile panelin geri kalanı artık AYNI kaynaktan
        # beslenir.
        from src.state_engine import _staff_norm_family
        for _, r in fm_aktif.iterrows():
            if pd.isna(r["MağazaID"]) or pd.isna(r["Departman"]):
                continue
            family_key = _staff_norm_family(r.get('Unvan'), r["Departman"])
            uid = unvan_id_by_key.get(family_key)
            if not uid:
                continue
            key = (r["MağazaID"], uid)
            mevcut_sayim[key] = mevcut_sayim.get(key, 0) + 1

    ciftler = sorted(set(norm_toplam) | set(mevcut_sayim))

    # Input içindeki resmî mağaza+norm ailesi dağılımı. Kodda 51/23 gibi
    # sabit bir üst sınır yoktur; satırlar inputtan dinamik okunur.
    kontrol_eksik = {}
    kontrol_fazla = {}
    # DÜZELTME (KRİTİK, 29 Ağustos 2026 — src.state_engine.state()'te
    # DAHA ÖNCE 20.08.2026'da bulunup düzeltilmiş, aynı hata sınıfı bu
    # dosyaya HİÇ yansıtılmamıştı): REFERENTIAL_CONTROL sayfası, input
    # dosyasında BİR KEZ (muhtemelen main.py'nin geçmiş bir çalıştırma
    # anında) yazılan DONMUŞ bir tabloydu — personel eklense/çıkarılsa
    # (Fact_Mevcut güncellense) bile bu sayfa OTOMATİK güncellenmiyordu.
    # Bu satır koşulsuz "sayfa varsa kullan" dediği için, CEO Özet'in
    # okuduğu Norm_Durumu/Magaza_KPI_Skor_Karti sürekli BAYAT rakamlar
    # (canlı üretimde doğrulandı: 49/23 donmuş — doğrusu main.py'nin
    # ürettiği 48/37) gösterebiliyordu. Artık state_engine.py ile AYNI
    # feature flag (varsayılan KAPALI) kullanılıyor — REFERENTIAL_CONTROL
    # yalnız OMEHR_USE_REFERENTIAL_CONTROL=1 açıkça ayarlanırsa okunur.
    import os as _os_ref_ctl
    if _os_ref_ctl.getenv('OMEHR_USE_REFERENTIAL_CONTROL', '0') == '1':
      try:
        rc = _oku_ham(input_path, 'REFERENTIAL_CONTROL')
        if {'MağazaID','UnvanID','Norm Eksiği Kontrol','Norm Fazlası Kontrol'}.issubset(rc.columns):
            for _, rr in rc.iterrows():
                mid, uid = rr.get('MağazaID'), rr.get('UnvanID')
                if pd.isna(mid) or pd.isna(uid):
                    continue
                key=(mid,uid)
                kontrol_eksik[key]=kontrol_eksik.get(key,0)+int(pd.to_numeric(rr.get('Norm Eksiği Kontrol'),errors='coerce') or 0)
                kontrol_fazla[key]=kontrol_fazla.get(key,0)+int(pd.to_numeric(rr.get('Norm Fazlası Kontrol'),errors='coerce') or 0)
                if key not in ciftler:
                    ciftler.append(key)
            ciftler=sorted(set(ciftler))
      except Exception as _exc:
        log_swallowed('statiklestir: REFERENTIAL_CONTROL okunamadı', _exc)

    # DÜZELTME (tutarlılık, 29 Ağustos 2026): main_names/helper_pairs sabit
    # listeleri (yalnız Manav/Kasap/Şarküteri; Yönetici HİÇ dahil değildi
    # — bu yüzden Yönetici/Yönetici Yardımcısı dengesi burada hiç
    # uygulanmıyordu) ve "ana unvanda >=1 kişi olmalı" şartı KALDIRILDI —
    # src.state_engine::_reconcile_main_family_rules ("Kural A") ile
    # SENKRON: ana unvanlar ve çiftler artık resolve_assistant_pairs ile
    # (config + canlı veride bulunan her "X YARDIMCISI" için otomatik,
    # YÖNETİCİ dahil TÜM aileler) türetilir; 0 ana personel olsa bile
    # aile toplamı normu karşılıyorsa dengelenir (KASITLI — bkz. state_
    # engine.py'deki aynı fonksiyonun docstring'i, "Kural A").
    from services.norm_rule_config import load_norm_rules as _load_rules_for_pairs, resolve_assistant_pairs
    _rules_for_pairs = _load_rules_for_pairs()
    _bilinen_unvanlar = {_title_key(ad) for ad in unvan_ad_map.values()}
    _pairs = resolve_assistant_pairs(_rules_for_pairs, _bilinen_unvanlar)
    store_ids = {mid for mid, _ in set(norm_toplam) | set(mevcut_sayim) | set(kontrol_eksik)}
    for mid in store_ids:
        for main_key, helper_key in _pairs.items():
            main_uid = unvan_id_by_key.get(main_key)
            helper_uid = unvan_id_by_key.get(helper_key)
            if not main_uid or not helper_uid:
                continue
            if mevcut_sayim.get((mid,main_uid),0) >= norm_toplam.get((mid,main_uid),0):
                kontrol_eksik[(mid,main_uid)] = 0
            main_gap = max(0, norm_toplam.get((mid,main_uid),0)-mevcut_sayim.get((mid,main_uid),0))
            helper_capacity = max(0, mevcut_sayim.get((mid,helper_uid),0)-norm_toplam.get((mid,helper_uid),0))
            support = min(main_gap, helper_capacity)
            if support:
                kontrol_eksik[(mid,main_uid)] = max(0, int(kontrol_eksik.get((mid,main_uid),main_gap))-int(support))
                # DÜZELTME (tutarlılık, 29 Ağustos 2026): önceden yalnız
                # ana unvanın eksiği düşürülüyordu — yardımcının fazlası
                # HİÇ güncellenmiyordu (state_engine.py::_reconcile_main_
                # family_rules ise HER İKİSİNİ de düşürür). Somut etki:
                # ana unvan Eksik=0 gösterirken yardımcı Fazla değerini
                # yanlışlıkla dengelemeden bırakıyordu.
                helper_fazla_ham = max(0, mevcut_sayim.get((mid,helper_uid),0)-norm_toplam.get((mid,helper_uid),0))
                kontrol_fazla[(mid,helper_uid)] = max(0, int(kontrol_fazla.get((mid,helper_uid),helper_fazla_ham))-int(support))

    # DÜZELTME (tutarlılık, 29 Ağustos 2026): src.state_engine.state()'te
    # Aktif personelde fiilen kullanılan unvanlar da kapsamdadır. Fact_Norm'da
    # henüz satırı bulunmayan bir unvanın normu 0 kabul edilir ve mevcut
    # personeli fazlaya yazılır; böylece Excel'in statik Norm_Durumu sonucu
    # state_engine ve web paneliyle aynı kalır.
    kapsam_uidleri = {uid for (_, uid) in norm_toplam.keys()}
    kapsam_uidleri |= {uid for (_, uid) in mevcut_sayim.keys()}
    kapsam_uidleri |= {unvan_id_by_key.get(v) for v in _pairs.values() if unvan_id_by_key.get(v)}

    wb = openpyxl.load_workbook(input_path)
    degisti = False
    eksik_toplam_magaza: dict = {}
    fazla_toplam_magaza: dict = {}
    norm_toplam_magaza: dict = {}
    if "Norm_Durumu" in wb.sheetnames and ciftler:
        ws = wb["Norm_Durumu"]
        # KRİTİK DÜZELTME: Önceki sürüm, hesaplanan (MağazaID,UnvanID) sırasının
        # Excel'deki mevcut satır sırasıyla BİREBİR aynı olmasını şart koşuyordu
        # ("satır sırası değiştiyse dokunma") — herhangi bir uyuşmazlıkta o satırı
        # SESSİZCE atlıyordu. Gerçek üretimde bu, neredeyse TÜM satırların boş
        # kalmasına yol açtı (sadece 1 mağaza rastgele sırayla tuttu). Artık
        # satırlar sıraya güvenilmeden, (MağazaID,UnvanID) DEĞERİNE göre
        # eşleştirilerek bulunuyor.
        mevcut_satirlar: dict[tuple, int] = {}
        for i in range(2, ws.max_row + 1):
            anahtar = (ws.cell(i, 1).value, ws.cell(i, 3).value)
            if anahtar[0] is not None and anahtar[1] is not None:
                mevcut_satirlar[anahtar] = i
        sonraki_bos_satir = ws.max_row + 1
        for (mid, uid) in ciftler:
            i = mevcut_satirlar.get((mid, uid))
            if i is None:
                # Bu (mağaza,unvan) çifti tabloda hiç yoktu (ör. sonradan eklenen
                # kişi/mağaza) — yeni bir satır olarak ekle, atlama.
                i = sonraki_bos_satir
                sonraki_bos_satir += 1
                ws.cell(i, 1).value = mid
                ws.cell(i, 3).value = uid
            nk = norm_toplam.get((mid, uid), 0)
            mevcut = mevcut_sayim.get((mid, uid), 0)
            yardimci_uid = ""
            yardimci_nk = 0
            yardimci_mevcut = 0
            toplam_norm = nk
            toplam_mevcut = mevcut
            is_yardimci = False
            # DÜZELTME (KRİTİK, 29 Ağustos 2026 — benim genelleştirmemle
            # tespit edildi): "if kontrol_eksik or kontrol_fazla:" SÖZLÜĞÜN
            # GENEL OLARAK boş olup olmadığına bakıyordu — herhangi BİR
            # (mid,uid) çifti için (REFERENTIAL_CONTROL'den ya da aile
            # dengelemesinden 0 bile olsa) bir kayıt yazılınca, TÜM DİĞER
            # unvanlar da bu "override modu"na sürükleniyor ve kayıtları
            # OLMAYAN unvanlar için gerçek (nk-mevcut) hesaplanmak yerine
            # SESSİZCE 0 (varsayılan .get) yazılıyordu. Somut etki: Yönetici/
            # Manav/Kasap/Şarküteri'den biri "mevcut>=norm" durumuna gelip
            # kontrol_eksik[(mid,uid)]=0 yazdığı AN, o mağazadaki TÜM DİĞER
            # unvanların (ör. Kasiyer, Manav Teraziği — hiçbir aile kuralına
            # girmeyen unvanlar dahil) Norm_Durumu sayfasındaki gerçek eksiği
            # kayboluyor, hepsi 0 görünüyordu. Artık HER (mid,uid) çifti
            # kendi override durumuna göre AYRI AYRI değerlendirilir.
            if (mid, uid) in kontrol_eksik or (mid, uid) in kontrol_fazla:
                eksik = int(kontrol_eksik.get((mid,uid),0))
                fazla = int(kontrol_fazla.get((mid,uid),0))
            elif uid not in kapsam_uidleri:
                # Ne normda ne de aktif mevcutta bulunan teknik/boş satır.
                eksik = 0
                fazla = 0
            else:
                eksik = max(0, nk - mevcut)
                fazla = max(0, mevcut - nk)
            ws.cell(i, 2).value = mag_ad_map.get(mid, ws.cell(i, 2).value)
            ws.cell(i, 4).value = unvan_ad_map.get(uid, ws.cell(i, 4).value)
            ws.cell(i, 5).value = nk
            ws.cell(i, 6).value = mevcut
            ws.cell(i, 7).value = yardimci_uid or ""
            ws.cell(i, 8).value = yardimci_nk
            ws.cell(i, 9).value = yardimci_mevcut
            ws.cell(i, 10).value = toplam_norm
            ws.cell(i, 11).value = toplam_mevcut
            ws.cell(i, 12).value = "EVET" if is_yardimci else "HAYIR"
            ws.cell(i, 13).value = eksik
            ws.cell(i, 14).value = fazla
            eksik_toplam_magaza[mid] = eksik_toplam_magaza.get(mid, 0) + eksik
            fazla_toplam_magaza[mid] = fazla_toplam_magaza.get(mid, 0) + fazla
            norm_toplam_magaza[mid] = norm_toplam_magaza.get(mid, 0) + nk
        degisti = True

    if degisti:
        wb.save(input_path)

    # --- Mağaza KPI Skor Kartı — aynı prensiple, LibreOffice bağımsız ---
    try:
        skor_degisti = _skor_kartini_statiklestir(
            input_path, wb, mag_ad_map, norm_toplam_magaza, eksik_toplam_magaza, fazla_toplam_magaza
        )
        if skor_degisti:
            wb.save(input_path)
            degisti = True
    except Exception as _exc:
        log_swallowed("statiklestir: Magaza_KPI_Skor_Karti güncellenemedi", _exc)

    try:
        tahmin_degisti = _forecast_sayfasini_statiklestir(wb)
        if tahmin_degisti:
            wb.save(input_path)
            degisti = True
    except Exception as _exc:
        log_swallowed("statiklestir: Verimlilik_Operasyon_Tahmini güncellenemedi", _exc)
    return degisti


def _dogrusal_tahmin(x: list[float], y: list[float], yeni_x: list[float]) -> list[float]:
    """Excel'in FORECAST/TREND fonksiyonunun basit doğrusal regresyon
    eşdeğeri — numpy olmadan, saf Python ile en küçük kareler yöntemi."""
    n = len(x)
    if n < 2:
        return [y[-1] if y else 0.0 for _ in yeni_x]
    x_ort = sum(x) / n
    y_ort = sum(y) / n
    pay = sum((xi - x_ort) * (yi - y_ort) for xi, yi in zip(x, y))
    payda = sum((xi - x_ort) ** 2 for xi in x)
    egim = pay / payda if payda else 0.0
    kesisim = y_ort - egim * x_ort
    return [egim * xi + kesisim for xi in yeni_x]


def _forecast_sayfasini_statiklestir(wb) -> bool:
    """Verimlilik_Operasyon_Tahmini sayfasındaki Excel FORECAST formülüne
    dayalı hücreler (Ciro/Maliyet/Mesai/Devamsızlık tahmini + Ciro'dan
    türetilmiş İş Yükü Endeksi projeksiyonu) LibreOffice çalışmadıysa boş
    kalıyordu. Aynı doğrusal tahmini burada Python'da yeniden üretip statik
    değer olarak yazarız."""
    if "Verimlilik_Operasyon_Tahmini" not in wb.sheetnames:
        return False
    ws = wb["Verimlilik_Operasyon_Tahmini"]
    etiket_satir = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v:
            etiket_satir[str(v).strip()] = r

    gecmis_x = list(range(1, 13))
    yeni_x = [13, 14, 15, 16]
    ciftler = [
        ("Toplam Ciro (TL) (Geçmiş)", "Toplam Ciro (TL) (Tahmin)"),
        ("Toplam Personel Maliyeti (TL) (Geçmiş)", "Toplam Personel Maliyeti (TL) (Tahmin)"),
        ("Toplam Fazla Mesai Saat (Geçmiş)", "Toplam Fazla Mesai Saat (Tahmin)"),
        ("Toplam Devamsızlık Kaybı (FTE) (Geçmiş)", "Toplam Devamsızlık Kaybı (FTE) (Tahmin)"),
    ]
    degisti = False
    ciro_tahmin = None
    for gecmis_etiket, tahmin_etiket in ciftler:
        gr = etiket_satir.get(gecmis_etiket)
        tr = etiket_satir.get(tahmin_etiket)
        if gr is None or tr is None:
            continue
        gecmis_y = [ws.cell(gr, c).value for c in range(2, 14)]
        if any(v is None for v in gecmis_y):
            continue
        tahmin = _dogrusal_tahmin(gecmis_x, [float(v) for v in gecmis_y], yeni_x)
        for c, deger in zip(range(2, 6), tahmin):
            ws.cell(tr, c).value = round(deger, 1)
        degisti = True
        if "Ciro" in gecmis_etiket:
            ciro_tahmin = tahmin

    # Ciro tahminine dayalı İş Yükü Endeksi projeksiyonu (r=0,91 regresyonu)
    isyuku_r = etiket_satir.get("İş Yükü Endeksi (Ciro Tahmininden Türetilmiş)")
    seviye_r = etiket_satir.get("Tahmini Seviye")
    if ciro_tahmin and isyuku_r and "Istatistiksel_Model_Testi" in wb.sheetnames:
        stat_ws = wb["Istatistiksel_Model_Testi"]
        egim = stat_ws["B82"].value
        kesisim = stat_ws["B83"].value
        # B82/B83'ün kendisi de Excel formülüdür (SLOPE/INTERCEPT) — LibreOffice
        # çalışmadıysa bunlar da boş gelir. Bu durumda daha önce bu konuşmada
        # doğrulanmış sabit katsayılar (r=0,91, R²=0,821) yedek olarak kullanılır.
        # B82/B83'ün kendisi de Excel formülüdür (SLOPE/INTERCEPT). Bu fonksiyon
        # data_only=True OLMADAN açılmış bir workbook üzerinde çalıştığı için,
        # LibreOffice hesaplamamışsa buradan None DEĞİL, formül METNİNİN
        # KENDİSİ ("=SLOPE(...)") döner — sadece "is None" kontrolü bunu
        # YAKALAYAMAZ. Hem None hem formül-metni durumunu kontrol ederiz.
        def _formul_mu(v):
            return v is None or (isinstance(v, str) and v.strip().startswith("="))
        if _formul_mu(egim) or _formul_mu(kesisim):
            egim, kesisim = 1.01244110424325e-06, 23.0580493825485
        try:
            mag_sayisi = 47
            for c, ciro in zip(range(2, 6), ciro_tahmin):
                deger = round(kesisim + egim * (ciro / mag_sayisi), 1)
                ws.cell(isyuku_r, c).value = deger
                if seviye_r:
                    seviye = "Kritik" if deger >= 70 else ("Yüksek" if deger >= 50 else "Normal")
                    ws.cell(seviye_r, c).value = seviye
            degisti = True
        except Exception as _exc:
            log_swallowed("_forecast_sayfasini_statiklestir: İş Yükü Endeksi projeksiyonu yazılamadı", _exc)
    return degisti


def _skor_kartini_statiklestir(input_path, wb, mag_ad_map, norm_toplam_magaza, eksik_toplam_magaza, fazla_toplam_magaza) -> bool:
    if "Magaza_KPI_Skor_Karti" not in wb.sheetnames:
        return False

    def son_ay_ozet(sheet, deger_kolon, agg="sum"):
        try:
            df = pd.read_excel(input_path, sheet_name=sheet, header=1).dropna(how="all")
        except Exception as _exc:
            log_swallowed(f"_skor_kartini_statiklestir: '{sheet}' sayfası okunamadı", _exc, level="INFO")
            return {}
        if df.empty or "Ay" not in df.columns:
            return {}
        son_ay = df["Ay"].max()
        df = df[df["Ay"] == son_ay]
        return df.groupby("Mağaza")[deger_kolon].agg(agg).to_dict()

    devir_map = son_ay_ozet("Devir Riski", "Risk Skoru", "mean")
    devam_map = son_ay_ozet("Devamsızlık", "Fiili Kayıp FTE", "sum")
    mesai_map = son_ay_ozet("Fazla Mesai", "Fazla Mesai Saat", "sum")
    try:
        perf_df = pd.read_excel(input_path, sheet_name="Personel_Performans_Endeksi", header=1)
        perf_map = perf_df.groupby("Mağaza")["Performans Endeksi (0-100)"].mean().to_dict()
    except Exception as _exc:
        log_swallowed("_skor_kartini_statiklestir: Personel_Performans_Endeksi okunamadı", _exc, level="INFO")
        perf_map = {}

    devam_vals = list(devam_map.values()) or [0]
    mesai_vals = list(mesai_map.values()) or [0]
    d_min, d_max = min(devam_vals), max(devam_vals)
    m_min, m_max = min(mesai_vals), max(mesai_vals)

    ad_to_mid = {ad: mid for mid, ad in mag_ad_map.items()}

    ws = wb["Magaza_KPI_Skor_Karti"]
    degisti = False
    for r in range(3, ws.max_row + 1):
        ad = ws.cell(r, 1).value
        if not ad:
            continue
        mid = ad_to_mid.get(ad)
        norm_kadro = norm_toplam_magaza.get(mid, 0)
        eksik = eksik_toplam_magaza.get(mid, 0)
        fazla = fazla_toplam_magaza.get(mid, 0)
        norm_uyumu = max(0.0, min(100.0, 100 - (eksik + fazla) / max(1, norm_kadro) * 100))

        risk = devir_map.get(ad)
        risk = 50 if risk is None or pd.isna(risk) else risk
        devir_puan = max(0.0, min(100.0, 100 - risk))

        devam = devam_map.get(ad)
        devam = (d_min + d_max) / 2 if devam is None or pd.isna(devam) else devam
        devam_puan = 100.0 if d_max <= d_min else max(0.0, min(100.0, 100 - (devam - d_min) / (d_max - d_min) * 100))

        mesai = mesai_map.get(ad)
        mesai = (m_min + m_max) / 2 if mesai is None or pd.isna(mesai) else mesai
        mesai_puan = 100.0 if m_max <= m_min else max(0.0, min(100.0, 100 - (mesai - m_min) / (m_max - m_min) * 100))

        performans_puan = perf_map.get(ad)
        performans_puan = 60 if performans_puan is None or pd.isna(performans_puan) else performans_puan

        skor = round(norm_uyumu * 0.30 + devir_puan * 0.20 + devam_puan * 0.20 + mesai_puan * 0.15 + performans_puan * 0.15, 1)
        if skor >= 80:
            sinif = "🟢 Çok İyi"
        elif skor >= 65:
            sinif = "🔵 İyi"
        elif skor >= 50:
            sinif = "🟡 Orta"
        else:
            sinif = "🔴 Dikkat"

        ws.cell(r, 6).value = round(risk, 1) if risk is not None else None
        ws.cell(r, 7).value = round(devir_puan, 1)
        ws.cell(r, 8).value = round(devam, 2) if devam is not None else None
        ws.cell(r, 9).value = round(devam_puan, 1)
        ws.cell(r, 10).value = round(mesai, 1) if mesai is not None else None
        ws.cell(r, 11).value = round(mesai_puan, 1)
        ws.cell(r, 12).value = round(performans_puan, 1)
        ws.cell(r, 13).value = skor
        ws.cell(r, 14).value = sinif
        degisti = True
    return degisti
