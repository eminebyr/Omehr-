from __future__ import annotations

"""
EV-MAĞAZA YAKINLIK MOTORU
===========================
Personel_Adresleri sayfasındaki (Ev Enlem, Ev Boylam) ile Magaza_Adres
sayfasındaki (Enlem, Boylam) koordinatlarını kullanarak her personel için:

  1) Ev-Mevcut Mağaza Mesafesi (km) + Google Maps rotası
     -> personelin bugün çalıştığı mağazaya evinden uzaklığı ve rotası
  2) Evine En Yakın Mağaza 1/2/3 (ad, mesafe, Google Maps rotası)
     -> koordinatı bilinen TÜM mağazalar arasında evine en yakın 3 tanesi
  3) Mevcut Mağaza Evine En Yakın Mı -> "Evet"/"Hayır" (bugünkü mağazası
     zaten 1 numaralı en yakınsa transfer önerisine gerek yok demektir)

hesaplar ve bu sütunları doğrudan input Excel dosyasındaki Personel_Adresleri
sayfasına YAZAR. Sadece koordinatı GERÇEK/geçerli (Enlem/Boylam dolu)
personel ve mağazalar için hesaplama yapılır; koordinatı olmayanlar için
sütunlar boş bırakılır ve mevcut hesaplama akışını bozmaz.

Bu fonksiyon, motor her çalıştığında (web arayüzü açılışında VEYA
rapor/PDF-Excel üretim akışında) otomatik tetiklenecek şekilde
web/app.py ve src/engine_core.py içine bağlanmıştır.
"""

import math
from pathlib import Path
from urllib.parse import quote_plus

import openpyxl
import pandas as pd

REQUIRED_PERSON_COLS = {"PersonelID", "Ev Enlem", "Ev Boylam"}
TOP_N = 3

NEW_COLUMNS = [
    "Ev-Mevcut Mağaza Mesafesi (km)",
    "Ev-Mevcut Mağaza Google Maps Rota",
    "Mevcut Mağaza Evine En Yakın Mı",
] + [
    col
    for i in range(1, TOP_N + 1)
    for col in (
        f"Evine En Yakın Mağaza {i}",
        f"Evine En Yakın Mağaza {i} Mesafesi (km)",
        f"Evine En Yakın Mağaza {i} Google Maps Rota",
    )
]


def _haversine_km(lat1, lon1, lat2, lon2):
    try:
        a1, b1, a2, b2 = map(float, (lat1, lon1, lat2, lon2))
    except (TypeError, ValueError):
        return None
    r = 6371.0088
    p1, p2 = math.radians(a1), math.radians(a2)
    dp = math.radians(a2 - a1)
    dl = math.radians(b2 - b1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


def maps_route(lat1, lon1, lat2, lon2) -> str | None:
    """Ev koordinatından mağaza koordinatına Google Maps sürüş rotası linki."""
    if any(v is None or pd.isna(v) for v in (lat1, lon1, lat2, lon2)):
        return None
    origin = quote_plus(f"{lat1},{lon1}")
    destination = quote_plus(f"{lat2},{lon2}")
    return (
        "https://www.google.com/maps/dir/?api=1"
        f"&origin={origin}&destination={destination}&travelmode=driving"
    )


# Geriye dönük uyumluluk için eski özel isim de aynı fonksiyona işaret eder.
_maps_route = maps_route


def _store_coordinate_map(input_path: Path) -> dict:
    """MağazaID -> (lat, lon, Mağaza adı); yalnız gerçek/geçerli koordinatlar."""
    try:
        from services.cached_excel_reader import read_sheet_cached
        magaza_adres = read_sheet_cached(input_path, "Magaza_Adres")
    except ValueError:
        return {}
    if magaza_adres.empty or "MağazaID" not in magaza_adres.columns:
        return {}
    out = {}
    for _, row in magaza_adres.iterrows():
        mid = str(row.get("MağazaID", "")).strip()
        lat = pd.to_numeric(row.get("Enlem"), errors="coerce")
        lon = pd.to_numeric(row.get("Boylam"), errors="coerce")
        if mid and pd.notna(lat) and pd.notna(lon):
            out[mid] = (float(lat), float(lon), str(row.get("Mağaza", "")).strip())
    return out


def _current_store_map(input_path: Path) -> dict:
    """PersonelID -> MağazaID (Fact_Mevcut'taki güncel/aktif ataması)."""
    try:
        from services.cached_excel_reader import read_sheet_cached
        fm = read_sheet_cached(input_path, "Fact_Mevcut")
    except ValueError:
        return {}
    if fm.empty or "PersonelID" not in fm.columns:
        return {}
    out = {}
    for _, row in fm.iterrows():
        pid = str(row.get("PersonelID", "")).strip()
        mid = str(row.get("MağazaID", "")).strip()
        if pid and mid:
            out[pid] = mid
    return out


def compute_home_proximity(input_path: Path) -> pd.DataFrame:
    """Hesaplama sonuçlarını (yazmadan) bir DataFrame olarak döndürür."""
    input_path = Path(input_path)
    try:
        from services.cached_excel_reader import read_sheet_cached
        personel = read_sheet_cached(input_path, "Personel_Adresleri")
    except ValueError:
        return pd.DataFrame()
    if personel.empty or not REQUIRED_PERSON_COLS.issubset(personel.columns):
        return pd.DataFrame()

    store_coords = _store_coordinate_map(input_path)
    current_store = _current_store_map(input_path)
    if not store_coords:
        return pd.DataFrame()

    results = []
    for _, row in personel.iterrows():
        pid = str(row.get("PersonelID", "")).strip()
        lat = pd.to_numeric(row.get("Ev Enlem"), errors="coerce")
        lon = pd.to_numeric(row.get("Ev Boylam"), errors="coerce")
        rec = {"PersonelID": pid}
        if not pid or pd.isna(lat) or pd.isna(lon):
            rec.update({c: None for c in NEW_COLUMNS})
            results.append(rec)
            continue
        lat, lon = float(lat), float(lon)

        # Koordinatı bilinen tüm mağazalara mesafeyi hesapla, yakınlığa göre sırala
        distances = []
        for mid, (slat, slon, sname) in store_coords.items():
            km = _haversine_km(lat, lon, slat, slon)
            if km is not None:
                distances.append((km, mid, sname, slat, slon))
        distances.sort(key=lambda x: x[0])
        top = distances[:TOP_N]

        for i in range(TOP_N):
            idx = i + 1
            if i < len(top):
                km, mid, sname, slat, slon = top[i]
                rec[f"Evine En Yakın Mağaza {idx}"] = sname
                rec[f"Evine En Yakın Mağaza {idx} Mesafesi (km)"] = round(km, 1)
                rec[f"Evine En Yakın Mağaza {idx} Google Maps Rota"] = _maps_route(lat, lon, slat, slon)
            else:
                rec[f"Evine En Yakın Mağaza {idx}"] = None
                rec[f"Evine En Yakın Mağaza {idx} Mesafesi (km)"] = None
                rec[f"Evine En Yakın Mağaza {idx} Google Maps Rota"] = None

        # Mevcut çalıştığı mağazaya mesafe + rota
        cur_mid = current_store.get(pid)
        cur_km = None
        cur_route = None
        if cur_mid and cur_mid in store_coords:
            slat, slon, _ = store_coords[cur_mid]
            cur_km = _haversine_km(lat, lon, slat, slon)
            cur_route = _maps_route(lat, lon, slat, slon)

        rec["Ev-Mevcut Mağaza Mesafesi (km)"] = round(cur_km, 1) if cur_km is not None else None
        rec["Ev-Mevcut Mağaza Google Maps Rota"] = cur_route
        if cur_mid and top:
            rec["Mevcut Mağaza Evine En Yakın Mı"] = "Evet" if cur_mid == top[0][1] else "Hayır"
        else:
            rec["Mevcut Mağaza Evine En Yakın Mı"] = None
        results.append(rec)

    return pd.DataFrame(results)


def refresh_home_proximity(input_path: Path) -> int:
    """Hesaplar VE doğrudan input dosyasındaki Personel_Adresleri sayfasına
    yazar. Motor her çalıştığında (web açılışı / rapor üretimi) çağrılması
    beklenir. Döndürdüğü değer güncellenen satır sayısıdır."""
    input_path = Path(input_path)
    calc = compute_home_proximity(input_path)
    if calc.empty:
        return 0

    wb = openpyxl.load_workbook(input_path)
    if "Personel_Adresleri" not in wb.sheetnames:
        return 0
    ws = wb["Personel_Adresleri"]
    headers = [c.value for c in ws[1]]

    # Eksik yeni sütunları başlık satırına ekle
    col_idx = {}
    for h in headers:
        if h:
            col_idx[h] = headers.index(h) + 1
    next_col = len(headers) + 1
    for col_name in NEW_COLUMNS:
        if col_name not in col_idx:
            ws.cell(1, next_col).value = col_name
            col_idx[col_name] = next_col
            next_col += 1

    pid_col = col_idx.get("PersonelID")
    if not pid_col:
        return 0

    calc_by_pid = {r["PersonelID"]: r for _, r in calc.iterrows()}
    guncellenen = 0
    for r in range(2, ws.max_row + 1):
        pid = str(ws.cell(r, pid_col).value or "").strip()
        if pid not in calc_by_pid:
            continue
        rec = calc_by_pid[pid]
        for col_name in NEW_COLUMNS:
            ws.cell(r, col_idx[col_name]).value = rec.get(col_name)
        guncellenen += 1

    wb.save(input_path)
    return guncellenen


if __name__ == "__main__":
    import sys
    p = Path(sys.argv[1]) if len(sys.argv) > 1 else None
    if not p:
        print("Kullanım: python3 services/home_proximity.py <input.xlsx>")
        sys.exit(1)
    n = refresh_home_proximity(p)
    print(f"Güncellenen personel satırı: {n}")
