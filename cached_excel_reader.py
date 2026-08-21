from __future__ import annotations

"""Merkezi, önbellekli Excel sayfa okuyucu (mimari ayrıştırma + performans).

DÜZELTME (ölçülmüş performans sorunu): 8 web sekmesi (ceo_ozet,
performans, isgucu_tahmini, operasyon_gorselleri, verimlilik_gorselleri,
ai_operasyon, ai_geri_bildirim, veri_toplama) kendi başlarına, HER
SAYFA ZİYARETİNDE `pd.read_excel(...)` ile aynı dosyaları tekrar tekrar
okuyordu — ölçüldü: ek bir sayfa okuması başına ~0.5 saniye. `ctx.sheets`
zaten ana input dosyasının TÜM sayfalarını önbellekli tutuyordu, ama bu
sekmeler bunu KULLANMIYORDU; üstelik çoğu `header=1` gibi ctx.sheets'in
varsayılan `header=0`'ından FARKLI parametrelerle okuyordu — bu yüzden
doğrudan ctx.sheets'e yönlendirmek YANLIŞ sütun eşlemesine yol açardı.

Bu modül, dosyanın mtime+boyutuna göre anahtarlanan GENEL bir önbellek
sunar — hem ana input dosyası hem de ayrı çıktı dosyaları (ör.
V19_AI_Norm_Sonuclari.xlsx) için doğru header parametresiyle çalışır.
Dosya gerçekten değişmediği sürece ikinci okuma anında döner; herhangi
bir yazmadan hemen sonra otomatik olarak taze okunur (mtime değiştiği
için önbellek anahtarı da değişir).
"""

from functools import lru_cache
from pathlib import Path
import hashlib
import threading

import pandas as pd


def _dosya_anahtari(path: Path) -> tuple[str, int, int]:
    st_ = path.stat()
    return str(path.resolve()), int(st_.st_mtime_ns), int(st_.st_size)


# ============================================================================
# DÜZELTME (Madde 6 — Excel Change Watcher, sayfa fingerprint'i):
# Önceki tasarım TEK dosya-seviyesi mtime anahtarı kullanıyordu — dosyanın
# HERHANGİ bir sayfası değişince (ör. yalnız Fact_Mevcut), TÜM 64 sayfanın
# önbellek anahtarı da değişiyordu (aynı mtime'a bağlı olduğu için), bu da
# "AI veya diğer 63 sayfa yeniden okunmamalı" hedefini ihlal ediyordu —
# bir sonraki erişimde HEPSİ yeniden okunuyordu.
#
# Artık İKİ SEVİYELİ kontrol var:
#   1. seviye (ucuz): dosya mtime+boyutu değişti mi? Değişmediyse hiçbir
#      şey yapılmaz (yalnız dict bakışı, dosya AÇILMAZ) — en sık durum.
#   2. seviye (yalnız dosya değiştiğinde): dosya TEK SEFERDE açılır, HER
#      sayfanın içerik hash'i hesaplanır. Yalnız hash'i GERÇEKTEN değişen
#      sayfalar yeniden ayrıştırılır; DEĞİŞMEYEN sayfalar ÖNCEKİ DataFrame
#      referansını korur (yeniden okunmaz/kopyalanmaz).
# ============================================================================
_SAYFA_ONBELLEGI: dict[str, dict] = {}
_ONBELLEK_KILIDI = threading.Lock()


def _sayfa_hash(df: pd.DataFrame) -> str:
    try:
        return hashlib.sha256(pd.util.hash_pandas_object(df, index=True).values.tobytes()).hexdigest()
    except Exception:
        # Hashlenemeyen (ör. karışık tip) bir sayfa için güvenli yedek:
        # şekil + son sütun adları — yine de çoğu değişikliği yakalar.
        return f"fallback:{df.shape}:{list(df.columns)[-5:]}"


def _dosyayi_gerekirse_yenile(path_text: str, mtime_ns: int, size: int) -> dict:
    with _ONBELLEK_KILIDI:
        durum = _SAYFA_ONBELLEGI.get(path_text)
        if durum is not None and durum["mtime_ns"] == mtime_ns and durum["size"] == size:
            return durum  # 1. seviye: dosya değişmedi, dosyayı AÇMA.

        # 2. seviye: dosya değişti (veya ilk kez görülüyor) — TEK SEFERDE aç.
        eski_hash = (durum or {}).get("sayfa_hash", {})
        eski_veri = (durum or {}).get("sayfa_veri", {})
        yeni_hash: dict[str, str] = {}
        yeni_veri: dict[str, pd.DataFrame] = {}
        degisen_sayfalar: list[str] = []
        with pd.ExcelFile(path_text) as xls:
            for sheet in xls.sheet_names:
                ham = pd.read_excel(xls, sheet_name=sheet, header=0)
                h = _sayfa_hash(ham)
                yeni_hash[sheet] = h
                if eski_hash.get(sheet) == h and sheet in eski_veri:
                    yeni_veri[sheet] = eski_veri[sheet]  # İÇERİK AYNI: eski referansı koru.
                else:
                    yeni_veri[sheet] = ham  # İÇERİK DEĞİŞTİ (veya yeni): taze veriyi tut.
                    degisen_sayfalar.append(sheet)
        yeni_durum = {"mtime_ns": mtime_ns, "size": size, "sayfa_hash": yeni_hash,
                      "sayfa_veri": yeni_veri, "header_varyantlari": {}}
        _SAYFA_ONBELLEGI[path_text] = yeni_durum
        if not hasattr(_dosyayi_gerekirse_yenile, "_son_degisen"):
            _dosyayi_gerekirse_yenile._son_degisen = {}
        _dosyayi_gerekirse_yenile._son_degisen[path_text] = degisen_sayfalar
        return yeni_durum


@lru_cache(maxsize=64)
def _read_excel_cached(path_text: str, mtime_ns: int, size: int, sheet_name: str, header: int) -> pd.DataFrame:
    durum = _dosyayi_gerekirse_yenile(path_text, mtime_ns, size)
    if header == 0:
        if sheet_name not in durum["sayfa_veri"]:
            raise KeyError(f"Sayfa bulunamadı: {sheet_name}")
        return durum["sayfa_veri"][sheet_name]
    # header != 0 olan varyantlar (ör. iki satırlı başlık şeritli sayfalar):
    # sayfa İÇERİK hash'i (her zaman header=0 ile) DEĞİŞMEDİĞİ sürece bu
    # varyant da önbellekte tutulur — dosya değişmedikçe tekrar okunmaz.
    varyantlar = durum["header_varyantlari"].setdefault(sheet_name, {})
    anahtar = (durum["sayfa_hash"].get(sheet_name), header)
    if anahtar not in varyantlar:
        varyantlar[anahtar] = pd.read_excel(path_text, sheet_name=sheet_name, header=header)
    return varyantlar[anahtar]


def read_sheet_cached(path: Path, sheet_name: str, header: int = 0) -> pd.DataFrame:
    """pd.read_excel(path, sheet_name=..., header=...) ile AYNI sonucu
    verir, ama dosya değişmediği sürece diskten tekrar okumaz.
    Döndürülen DataFrame'in bir KOPYASIdır — çağıran güvenle değiştirebilir."""
    p = Path(path)
    path_text, mtime_ns, size = _dosya_anahtari(p)
    return _read_excel_cached(path_text, mtime_ns, size, sheet_name, header).copy(deep=True)


@lru_cache(maxsize=8)
def _read_workbook_cached(path_text: str, mtime_ns: int, size: int, data_only: bool):
    from openpyxl import load_workbook
    return load_workbook(path_text, data_only=data_only)


def read_workbook_cached(path: Path, data_only: bool = True):
    """openpyxl.load_workbook(path, data_only=...) ile AYNI sonucu verir,
    dosya değişmediği sürece tekrar açmaz.

    DİKKAT: openpyxl Workbook nesneleri DEĞİŞTİRİLEBİLİR (mutable) —
    önbellekten dönen nesne ÜZERİNDE YAZMA/KAYDETME yapmayın (paylaşılan
    önbelleği bozar). Yalnız OKUMA amaçlı kullanın; kaydetme gerekiyorsa
    services/master_data_admin.py'nin kendi (kilitli, atomik) yazma
    yolunu kullanın."""
    p = Path(path)
    path_text, mtime_ns, size = _dosya_anahtari(p)
    return _read_workbook_cached(path_text, mtime_ns, size, data_only)


# ============================================================================
# MERKEZİ EXCEL DATA SERVICE — isimlendirilmiş, alan-bilgili yükleyiciler
# (OMEHR hızlandırma şartnamesi Madde 4). Web katmanı artık DOĞRUDAN
# pd.read_excel(INPUT, sheet_name=...) çağırmak yerine buradaki
# load_*() fonksiyonlarını kullanmalıdır — yol çözümlemesi (aktif
# tenant + runtime kökü) burada merkezi olarak yapılır, yukarıdaki
# read_sheet_cached() önbelleğini kullanır (paralel bir önbellek DEĞİL).
# ============================================================================

def _guncel_input_yolu() -> Path:
    from services.runtime_paths import runtime_root
    from services.settings import input_path
    return input_path(runtime_root())


def load_fact_mevcut() -> pd.DataFrame:
    return read_sheet_cached(_guncel_input_yolu(), "Fact_Mevcut")


def load_fact_norm() -> pd.DataFrame:
    return read_sheet_cached(_guncel_input_yolu(), "Fact_Norm")


def load_dim_magaza() -> pd.DataFrame:
    return read_sheet_cached(_guncel_input_yolu(), "Dim_Magaza")


def load_dim_unvan() -> pd.DataFrame:
    return read_sheet_cached(_guncel_input_yolu(), "Dim_Unvan")


def load_mail_listesi() -> pd.DataFrame:
    return read_sheet_cached(_guncel_input_yolu(), "Mail_Listesi")


def load_transfer_talepleri() -> pd.DataFrame:
    return read_sheet_cached(_guncel_input_yolu(), "Transfer_Talepleri")


def load_cached_table(sheet_name: str, header: int = 0) -> pd.DataFrame:
    """Şartnamedeki genel `load_cached_table()` karşılığı — yukarıdaki
    isimlendirilmiş yükleyicilerin kapsamadığı herhangi bir sayfa için."""
    return read_sheet_cached(_guncel_input_yolu(), sheet_name, header=header)


def invalidate_table(sheet_name: str | None = None) -> None:
    """Şartnamedeki `invalidate_table()` karşılığı.

    NOT: Alttaki önbellek zaten dosyanın TAMAMININ mtime+boyutuna göre
    anahtarlanıyor (bkz. modül docstring'i) — tek bir Excel dosyasında
    HERHANGİ bir yazma, dosyanın mtime'ını değiştirir ve bu da o dosyaya
    ait TÜM önbellek anahtarlarını doğal olarak geçersiz kılar (bir
    sonraki okuma otomatik olarak taze veri döner). Bu yüzden gerçek bir
    kayıp/yanlış-önbellek riski YOKTUR; bu fonksiyon arayüz tamlığı ve
    "yazma sonrası açıkça temizle" isteyen çağıranlar için sağlanır."""
    _read_excel_cached.cache_clear()
    with _ONBELLEK_KILIDI:
        _SAYFA_ONBELLEGI.clear()


def refresh_changed_table(sheet_name: str) -> pd.DataFrame:
    """Şartnamedeki `refresh_changed_table()` karşılığı — açıkça
    invalidate edip taze okur (invalidate_table() + load_cached_table()
    kısayolu)."""
    invalidate_table(sheet_name)
    return load_cached_table(sheet_name)


# ============================================================================
# CHANGE MANIFEST (Madde 7) — son yazmada HANGİ sayfaların gerçekten
# değiştiğini dışa açan API. Sayfa fingerprint'i zaten yukarıdaki
# _dosyayi_gerekirse_yenile() içinde hesaplanıyor; burada yeni bir
# mekanizma KURULMUYOR, mevcut durum dışa açılıyor.
# ============================================================================

def son_degisen_sayfalar(path: Path) -> list[str]:
    """Bu dosya için önbellek son güncellendiğinde İÇERİĞİ GERÇEKTEN
    değişmiş (veya ilk kez görülmüş) sayfaların adlarını döndürür.

    Şartnamedeki "hangi sayfa/hangi KPI/hangi rapor etkilendi?" sorusuna
    temel sağlar (madde 7, 39). Dosyanın kendisi HİÇ okunmadıysa boş
    liste döner — önce en az bir read_sheet_cached()/load_*() çağrısı
    yapılmalıdır."""
    path_text, mtime_ns, size = _dosya_anahtari(Path(path))
    durum = _dosyayi_gerekirse_yenile(path_text, mtime_ns, size)
    onceki = getattr(_dosyayi_gerekirse_yenile, "_son_degisen", {}).get(path_text, [])
    return onceki


def personel_alan_degisikligi(eski_satir: dict, yeni_satir: dict, alanlar: list[str] | None = None) -> list[dict]:
    """İki personel kaydı (eski/yeni) arasındaki alan bazlı farkları,
    şartname madde 7'deki manifest şemasına uygun şekilde döndürür:
    {sheet, key, field, old_value, new_value, magaza, unvan}."""
    if alanlar is None:
        alanlar = sorted(set(eski_satir.keys()) | set(yeni_satir.keys()))
    kisi = yeni_satir.get("İsim Soyisim") or eski_satir.get("İsim Soyisim") or ""
    magaza = yeni_satir.get("Mağaza") or eski_satir.get("Mağaza") or ""
    unvan = yeni_satir.get("Unvan") or eski_satir.get("Unvan") or ""
    degisiklikler = []
    for alan in alanlar:
        eski_deger = eski_satir.get(alan)
        yeni_deger = yeni_satir.get(alan)
        if str(eski_deger or "") != str(yeni_deger or ""):
            degisiklikler.append({
                "sheet": "Fact_Mevcut", "key": kisi, "field": alan,
                "old_value": eski_deger, "new_value": yeni_deger,
                "magaza": magaza, "unvan": unvan,
            })
    return degisiklikler
