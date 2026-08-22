"""Veri Toplama sekmesi.

Bu modül, web/app.py içindeki eski "with tabs[N]:" bloğundan otomatik
olarak çıkarılmıştır. Kod davranışı değiştirilmeden taşınmıştır; tüm
paylaşılan durum (sheets, kullanıcı/rol bilgisi, fm/detail/stores/kpis,
servis fonksiyonları) web.context.PageContext üzerinden gelir.
"""
from __future__ import annotations

import streamlit as st
import pandas as pd
import plotly.express as px

import openpyxl
from web.context import PageContext


def render(ctx: PageContext) -> None:
    """Veri Toplama sekmesinin içeriğini çizer."""
    sheets, acc = ctx.sheets, ctx.acc
    fm, detail, stores, kpis = ctx.fm, ctx.detail, ctx.stores, ctx.kpis
    user, username, role, scope, email = ctx.user, ctx.username, ctx.role, ctx.scope, ctx.email
    is_global = ctx.is_global
    can_view_personal_address = ctx.can_view_personal_address
    approval_level, can_approve = ctx.approval_level, ctx.can_approve
    ROOT, INPUT, OUTPUT, DB = ctx.root, ctx.input_path, ctx.output_path, ctx.db_path
    APPROVERS, BD_RENK = ctx.approvers, ctx.bd_renk
    db, log = ctx.db, ctx.log
    enqueue, job_status, tenant_code = ctx.enqueue, ctx.job_status, ctx.tenant_code
    norm_text, tr_number, tr_money_compact = ctx.norm_text, ctx.tr_number, ctx.tr_money_compact
    set_password, password_error = ctx.set_password, ctx.password_error
    refresh_home_proximity, maps_route = ctx.refresh_home_proximity, ctx.maps_route
    verify_password, transfer_recipients = ctx.verify_password, ctx.transfer_recipients
    cancel_transfer_request, redirect_transfer_request = ctx.cancel_transfer_request, ctx.redirect_transfer_request
    bulk_branch_mail_panel = ctx.bulk_branch_mail_panel
    _enqueue_and_process = ctx.enqueue_and_process
    read_input = ctx.read_input

    st.subheader("Veri Toplama — Saha Ölçümü / İK-Finans Verisi")
    st.caption(
        "Sistemde 'Saha Etüdü Bekleniyor' işaretli operasyonel veriler var (Standart Süre, "
        "Kapasite Politikası, Maliyet). Bu sekme, GERÇEK ölçüm/kayıtları toplayıp sisteme "
        "aktarmanızı sağlar — hiçbir sayı otomatik uydurulmaz, sadece siz ne girerseniz o aktarılır."
    )

    _sablon_dizin = ROOT / "reference" / "veri_toplama_sablonlari"

    vt_c1, vt_c2 = st.columns(2)
    with vt_c1:
        st.markdown("#### 1️⃣ Saha Zaman Etüdü (Standart Süre)")
        st.caption("Fiş işlemi, Kasa açılışı gibi 21 aktivitenin kronometre ile ölçülmesi için.")
        _saha_sablon = _sablon_dizin / "SAHA_OLCUM_FORMU.xlsx"
        if _saha_sablon.is_file():
            st.download_button(
                "📥 Boş formu indir (SAHA_OLCUM_FORMU.xlsx)",
                _saha_sablon.read_bytes(), file_name="SAHA_OLCUM_FORMU.xlsx",
                use_container_width=True, key="vt_saha_indir",
            )
        _saha_yuklenen = st.file_uploader("Doldurulmuş formu yükle", type=["xlsx"], key="vt_saha_yukle")
        if _saha_yuklenen is not None and st.button("Saha ölçümünü sisteme uygula", key="vt_saha_uygula"):
            try:
                import openpyxl as _oxl
                from services.veri_toplama import saha_olcumu_uygula_ve_serialize
                _form_wb = _oxl.load_workbook(_saha_yuklenen, data_only=True)
                _ana_wb = _oxl.load_workbook(INPUT)
                _n, _olcumler, _veri = saha_olcumu_uygula_ve_serialize(_form_wb, _ana_wb)
                if _n == 0:
                    st.warning("Formda henüz gerçekten doldurulmuş (en az 1 ölçümü olan) hiçbir aktivite bulunamadı.")
                else:
                    st.success(f"✅ {_n} aktivite güncellendi: " + ", ".join(o["akt_id"] for o in _olcumler))
                    st.download_button(
                        "📤 Güncellenmiş ana dosyayı indir", _veri,
                        file_name="OMEHR_AI_NORM_TRANSFER_INPUT_SAHA_GUNCEL.xlsx",
                        use_container_width=True, key="vt_saha_indir_sonuc",
                    )
                    st.caption("Bu dosyayı kontrol edip beğenirseniz, input/ klasöründeki ana dosyanın üzerine kopyalayın.")
            except Exception as _exc:
                st.error(f"İşlenemedi: {_exc}")

    with vt_c2:
        st.markdown("#### 2️⃣ İK/Finans Verisi (Kapasite + Maliyet)")
        st.caption("Vardiya/mola politikası ve bordro maliyet rakamları için — saha ölçümü gerekmez.")
        _ik_sablon = _sablon_dizin / "IK_FINANS_VERI_FORMU.xlsx"
        if _ik_sablon.is_file():
            st.download_button(
                "📥 Boş formu indir (IK_FINANS_VERI_FORMU.xlsx)",
                _ik_sablon.read_bytes(), file_name="IK_FINANS_VERI_FORMU.xlsx",
                use_container_width=True, key="vt_ik_indir",
            )
        _ik_yuklenen = st.file_uploader("Doldurulmuş formu yükle", type=["xlsx"], key="vt_ik_yukle")
        if _ik_yuklenen is not None and st.button("İK/Finans verisini sisteme uygula", key="vt_ik_uygula"):
            try:
                import openpyxl as _oxl
                from services.veri_toplama import ik_finans_uygula_ve_serialize
                _form_wb = _oxl.load_workbook(_ik_yuklenen, data_only=True)
                _ana_wb = _oxl.load_workbook(INPUT)
                _sonuc, _veri = ik_finans_uygula_ve_serialize(_form_wb, _ana_wb)
                if _sonuc["kapasite"] == 0 and _sonuc["maliyet"] == 0:
                    st.warning("Formda henüz gerçekten doldurulmuş hiçbir satır bulunamadı.")
                else:
                    st.success(f"✅ Kapasite: {_sonuc['kapasite']} unvan, Maliyet: {_sonuc['maliyet']} unvan güncellendi.")
                    st.download_button(
                        "📤 Güncellenmiş ana dosyayı indir", _veri,
                        file_name="OMEHR_AI_NORM_TRANSFER_INPUT_IK_FINANS_GUNCEL.xlsx",
                        use_container_width=True, key="vt_ik_indir_sonuc",
                    )
            except Exception as _exc:
                st.error(f"İşlenemedi: {_exc}")

    st.markdown("---")
    st.markdown("#### 3️⃣ Vardiya Pik Saat (otomatik — form gerekmiyor)")
    st.caption(
        "Sistemde zaten GERÇEK saatlik yoğunluk verisi var — form doldurmaya gerek yok, "
        "tek tıkla mağaza bazında pik saat/katsayı gerçek veriden hesaplanır."
    )
    if st.button("Pik saatleri gerçek veriden hesapla", key="vt_pik_hesapla"):
        try:
            import openpyxl as _oxl
            from services.veri_toplama import vardiya_pik_turet_ve_serialize
            from services.cached_excel_reader import read_sheet_cached
            _saatlik = read_sheet_cached(INPUT, "Saatlik Yoğunluk", header=1)
            _ana_wb = _oxl.load_workbook(INPUT)
            _n, _veri = vardiya_pik_turet_ve_serialize(_ana_wb, _saatlik)
            st.success(f"✅ {_n} mağaza için pik saat verisi gerçek veriden türetildi.")
            st.download_button(
                "📤 Güncellenmiş ana dosyayı indir", _veri,
                file_name="OMEHR_AI_NORM_TRANSFER_INPUT_PIK_GUNCEL.xlsx",
                use_container_width=True, key="vt_pik_indir_sonuc",
            )
        except Exception as _exc:
            st.error(f"İşlenemedi: {_exc}")
