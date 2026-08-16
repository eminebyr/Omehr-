"""Verimlilik Görselleri sekmesi.

Bu modül, web/app.py içindeki eski "with tabs[N]:" bloğundan otomatik
olarak çıkarılmıştır. Kod davranışı değiştirilmeden taşınmıştır; tüm
paylaşılan durum (sheets, kullanıcı/rol bilgisi, fm/detail/stores/kpis,
servis fonksiyonları) web.context.PageContext üzerinden gelir.
"""
from __future__ import annotations

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go

import openpyxl
from web.context import PageContext
from services.safe_exec import log_swallowed


def render(ctx: PageContext) -> None:
    """Verimlilik Görselleri sekmesinin içeriğini çizer."""
    sheets, acc = ctx.sheets, ctx.acc
    fm, detail, stores, kpis = ctx.fm, ctx.detail, ctx.stores, ctx.kpis
    user, username, role, scope, email = ctx.user, ctx.username, ctx.role, ctx.scope, ctx.email
    is_global = ctx.is_global
    can_view_personal_address = ctx.can_view_personal_address
    approval_level, can_approve = ctx.approval_level, ctx.can_approve
    ROOT, INPUT, OUTPUT, DB = ctx.root, ctx.input_path, ctx.output_path, ctx.db_path
    APPROVERS, BD_RENK = ctx.approvers, ctx.bd_renk
    db, log = ctx.db, ctx.log
    enqueue, job_status, tenant_code = ctx.enqueue, ctx.job_status, ctx.tenant_code
    norm_text, tr_number, tr_money_compact = ctx.norm_text, ctx.tr_number, ctx.tr_money_compact
    set_password, password_error = ctx.set_password, ctx.password_error
    refresh_home_proximity, maps_route = ctx.refresh_home_proximity, ctx.maps_route
    verify_password, transfer_recipients = ctx.verify_password, ctx.transfer_recipients
    cancel_transfer_request, redirect_transfer_request = ctx.cancel_transfer_request, ctx.redirect_transfer_request
    bulk_branch_mail_panel = ctx.bulk_branch_mail_panel
    _enqueue_and_process = ctx.enqueue_and_process
    read_input = ctx.read_input

    st.subheader("Verimlilik Görselleri")
    st.caption("İş yükü endeksi, personel maliyeti, devir riski, fazla mesai ve devamsızlık — interaktif grafikler.")

    try:
        from services.cached_excel_reader import read_sheet_cached
        isyuku = read_sheet_cached(INPUT, "İş Yükü Endeksi", header=1)
    except Exception as _exc:
        log_swallowed("web.tab_modules.verimlilik_gorselleri.render: beklenmeyen hata", _exc)
        isyuku = pd.DataFrame()
    if not isyuku.empty:
        if not is_global:
            isyuku = isyuku[isyuku["Bölge"].astype(str).map(norm_text).eq(norm_text(scope))]
        try:
            from services.cached_excel_reader import read_sheet_cached
            _maliyet_metrik = read_sheet_cached(INPUT, "Personel Maliyeti", header=1).dropna(how="all")
            _mesai_metrik = read_sheet_cached(INPUT, "Fazla Mesai", header=1).dropna(how="all")
            _devam_metrik = read_sheet_cached(INPUT, "Devamsızlık", header=1).dropna(how="all")
        except Exception as _exc:
            log_swallowed("web.tab_modules.verimlilik_gorselleri.render: beklenmeyen hata", _exc)
            _maliyet_metrik = _mesai_metrik = _devam_metrik = pd.DataFrame()
        vc1, vc2, vc3, vc4 = st.columns(4)
        vc1.metric("Ortalama İş Yükü Endeksi", tr_number(isyuku["İş Yükü Endeksi"].mean(),1))
        if not _maliyet_metrik.empty and "Personel Maliyeti" in _maliyet_metrik.columns:
            son_ay_m = _maliyet_metrik["Ay"].max()
            vc2.metric("Son Ay Personel Maliyeti", tr_money_compact(_maliyet_metrik[_maliyet_metrik["Ay"]==son_ay_m]["Personel Maliyeti"].sum()))
        if not _mesai_metrik.empty and "Fazla Mesai Saat" in _mesai_metrik.columns:
            son_ay_fm = _mesai_metrik["Ay"].max()
            vc3.metric("Son Ay Toplam Fazla Mesai", tr_number(_mesai_metrik[_mesai_metrik["Ay"]==son_ay_fm]["Fazla Mesai Saat"].sum(),0," sa"))
        if not _devam_metrik.empty and "Fiili Kayıp FTE" in _devam_metrik.columns:
            son_ay_d = _devam_metrik["Ay"].max()
            vc4.metric("Son Ay Devamsızlık Kaybı", tr_number(_devam_metrik[_devam_metrik["Ay"]==son_ay_d]["Fiili Kayıp FTE"].sum(),1," FTE"))
        isyuku_sirali = isyuku.sort_values("İş Yükü Endeksi", ascending=False).head(20)
        fig_isyuku = px.bar(
            isyuku_sirali, x="İş Yükü Endeksi", y="Mağaza", orientation="h", color="Seviye",
            text="İş Yükü Endeksi", title="Mağaza Bazında İş Yükü Endeksi (En Yüksek 20)",
            color_discrete_map={"Kritik": "#C00000", "Yüksek": "#118B94", "Normal": "#70AD47", "Düşük": "#70AD47"},
        )
        fig_isyuku.update_traces(textposition="outside", cliponaxis=False)
        st.plotly_chart(fig_isyuku, use_container_width=True)
    else:
        st.info("İş Yükü Endeksi verisi bulunamadı.")

    c3, c4 = st.columns(2)
    try:
        from services.cached_excel_reader import read_sheet_cached
        maliyet = read_sheet_cached(INPUT, "Personel Maliyeti", header=1).dropna(how="all")
    except Exception as _exc:
        log_swallowed("web.tab_modules.verimlilik_gorselleri.render: beklenmeyen hata", _exc)
        maliyet = pd.DataFrame()
    if not maliyet.empty:
        maliyet_trend = maliyet.groupby("Ay", as_index=False)["Personel Maliyeti"].sum()
        fig_maliyet = px.bar(maliyet_trend, x="Ay", y="Personel Maliyeti", title="Aylık Toplam Personel Maliyeti", color_discrete_sequence=["#4472C4"])
        c3.plotly_chart(fig_maliyet, use_container_width=True)
    else:
        c3.info("Personel Maliyeti verisi bulunamadı.")

    try:
        from services.cached_excel_reader import read_sheet_cached
        mesai = read_sheet_cached(INPUT, "Fazla Mesai", header=1).dropna(how="all")
    except Exception as _exc:
        log_swallowed("web.tab_modules.verimlilik_gorselleri.render: beklenmeyen hata", _exc)
        mesai = pd.DataFrame()
    if not mesai.empty:
        mesai_top = mesai.groupby("Mağaza", as_index=False)["Fazla Mesai Saat"].sum().sort_values("Fazla Mesai Saat", ascending=False).head(15)
        fig_mesai = px.bar(mesai_top, x="Fazla Mesai Saat", y="Mağaza", orientation="h", text="Fazla Mesai Saat", title="En Yüksek Fazla Mesai (Dönem Toplamı, saat)", color_discrete_sequence=["#118B94"])
        fig_mesai.update_traces(textposition="outside", cliponaxis=False)
        c4.plotly_chart(fig_mesai, use_container_width=True)
    else:
        c4.info("Fazla Mesai verisi bulunamadı.")

    c5, c6 = st.columns(2)
    try:
        from services.cached_excel_reader import read_sheet_cached
        devir = read_sheet_cached(INPUT, "Devir Riski", header=1).dropna(how="all")
    except Exception as _exc:
        log_swallowed("web.tab_modules.verimlilik_gorselleri.render: beklenmeyen hata", _exc)
        devir = pd.DataFrame()
    if not devir.empty:
        devir_top = devir.groupby("Mağaza", as_index=False)["Risk Skoru"].mean().sort_values("Risk Skoru", ascending=False).head(15)
        fig_devir = px.bar(devir_top, x="Risk Skoru", y="Mağaza", orientation="h", text="Risk Skoru", title="En Yüksek Personel Devir Riski (15 Mağaza)", color="Risk Skoru", color_continuous_scale="Reds")
        fig_devir.update_traces(textposition="outside", cliponaxis=False)
        c5.plotly_chart(fig_devir, use_container_width=True)
    else:
        c5.info("Devir Riski verisi bulunamadı.")

    try:
        from services.cached_excel_reader import read_sheet_cached
        devamsizlik = read_sheet_cached(INPUT, "Devamsızlık", header=1).dropna(how="all")
    except Exception as _exc:
        log_swallowed("web.tab_modules.verimlilik_gorselleri.render: beklenmeyen hata", _exc)
        devamsizlik = pd.DataFrame()
    if not devamsizlik.empty:
        devam_trend = devamsizlik.groupby("Ay", as_index=False)["Fiili Kayıp FTE"].sum()
        fig_devam = px.line(devam_trend, x="Ay", y="Fiili Kayıp FTE", markers=True, title="Aylık Devamsızlıktan Kaynaklanan Fiili Kayıp (FTE)", color_discrete_sequence=["#C00000"])
        c6.plotly_chart(fig_devam, use_container_width=True)
    else:
        c6.info("Devamsızlık verisi bulunamadı.")

    st.markdown("---")
    st.markdown("#### En İyi ML Modeline Göre İş Yükü Tahmini (Dinamik)")
    st.caption(
        "Ridge/ElasticNet/Random Forest/Extra Trees yarıştırılıp en düşük hatalı model seçilir "
        "(ai_operations_engine.py). Bu grafik, formül tabanlı 'İş Yükü FTE' ile "
        "seçilen en iyi ML modelinin ürettiği tahmini karşılaştırır — main.py her "
        "çalıştığında otomatik güncellenir."
    )
    try:
        from services.cached_excel_reader import read_sheet_cached
        ai_norm_dinamik = read_sheet_cached(OUTPUT / "V19_AI_Norm_Sonuclari.xlsx", "AI_Norm_Sonuclari")
    except Exception as _exc:
        log_swallowed("web.tab_modules.verimlilik_gorselleri.render: beklenmeyen hata", _exc)
        ai_norm_dinamik = pd.DataFrame()
    if not ai_norm_dinamik.empty and "ML Tahmini İş Yükü FTE" in ai_norm_dinamik.columns:
        ai_norm_dinamik = ai_norm_dinamik.dropna(subset=["Mağaza", "Unvan"])
        if not is_global:
            ai_norm_dinamik = ai_norm_dinamik[ai_norm_dinamik["Bölge"].astype(str).map(norm_text).eq(norm_text(scope))]
        magaza_ozet = ai_norm_dinamik.groupby("Mağaza", as_index=False).agg(
            **{"İş Yükü FTE (Formül)": ("İş Yükü FTE", "sum"), "ML Tahmini İş Yükü FTE": ("ML Tahmini İş Yükü FTE", "sum")}
        )
        magaza_ozet["Fark (ML - Formül)"] = magaza_ozet["ML Tahmini İş Yükü FTE"] - magaza_ozet["İş Yükü FTE (Formül)"]
        en_farkli = magaza_ozet.reindex(magaza_ozet["Fark (ML - Formül)"].abs().sort_values(ascending=False).index).head(15)
        fig_ml = go.Figure()
        fig_ml.add_trace(go.Bar(x=en_farkli["Mağaza"], y=en_farkli["İş Yükü FTE (Formül)"], name="Formül Tabanlı", marker_color="#4472C4"))
        fig_ml.add_trace(go.Bar(x=en_farkli["Mağaza"], y=en_farkli["ML Tahmini İş Yükü FTE"], name="En İyi ML Modeli", marker_color="#118B94"))
        fig_ml.update_layout(barmode="group", title="Formül vs En İyi ML Modeli — En Çok Farklılaşan 15 Mağaza (Toplam İş Yükü FTE)", height=450)
        st.plotly_chart(fig_ml, use_container_width=True)
        kullanilan_model = ai_norm_dinamik["AI Modeli"].dropna().unique()
        if len(kullanilan_model):
            st.caption(f"Kullanılan model(ler): {', '.join(str(m) for m in kullanilan_model if m and str(m)!='nan')}")
    else:
        st.info("ML tahmin verisi henüz oluşmadı — main.py'yi çalıştırın (V19_AI_Norm_Sonuclari.xlsx üretir).")

    try:
        from services.cached_excel_reader import read_workbook_cached
        wb_isyuku_tahmin = read_workbook_cached(INPUT, data_only=True)
        ws_ciro_isyuku = wb_isyuku_tahmin["Verimlilik_Operasyon_Tahmini"]
        _isyuku_tahmin_satiri = None
        for r in range(1, ws_ciro_isyuku.max_row + 1):
            if ws_ciro_isyuku.cell(r, 1).value == "İş Yükü Endeksi (Ciro Tahmininden Türetilmiş)":
                _isyuku_tahmin_satiri = r
                break
        if _isyuku_tahmin_satiri:
            _degerler = [ws_ciro_isyuku.cell(_isyuku_tahmin_satiri, c).value for c in range(2, 6)]
            if all(v is not None for v in _degerler):
                fig_isyuku_gelecek = px.line(
                    x=["+30 Gün", "+60 Gün", "+90 Gün", "+120 Gün"], y=_degerler, markers=True,
                    labels={"x": "Ufuk", "y": "İş Yükü Endeksi"},
                    title="İş Yükü Endeksi — Ciro Tahminine Dayalı İleriye Dönük Projeksiyon (r=0,91)",
                )
                st.plotly_chart(fig_isyuku_gelecek, use_container_width=True)
    except Exception as _exc:
        log_swallowed("web.tab_modules.verimlilik_gorselleri.render: beklenmeyen hata", _exc)
        pass

    st.markdown("---")
    st.markdown("#### Verimlilik ve Operasyon Tahmini — 30/60/90/120 Gün")
    st.caption(
        "12 aylık geçmiş veri üzerinden Excel FORECAST fonksiyonu ile hesaplanır. "
        "Doğruluk, ilk 9 ayla son 3 ay geriye dönük tahmin edilip gerçekle karşılaştırılarak (backtesting) ölçülmüştür."
    )
    try:
        from services.cached_excel_reader import read_workbook_cached
        wb_tahmin = read_workbook_cached(INPUT, data_only=True)
        ws_t = wb_tahmin["Verimlilik_Operasyon_Tahmini"]
        tahmin_satirlari = []
        for r in range(1, ws_t.max_row + 1):
            etiket = ws_t.cell(r, 1).value
            if isinstance(etiket, str) and etiket.endswith("(Tahmin)"):
                degerler = [ws_t.cell(r, c).value for c in range(2, 6)]
                if all(v is not None for v in degerler):
                    tahmin_satirlari.append((etiket.replace(" (Tahmin)", ""), degerler))
        ws_d = wb_tahmin["Tahmin_Dogruluk_Testi"]
        dogruluk_satirlari = []
        metrik_baslik_satiri = None
        for r in range(1, ws_d.max_row + 1):
            if ws_d.cell(r, 1).value == "Metrik":
                metrik_baslik_satiri = r
                break
        if metrik_baslik_satiri:
            for r in range(metrik_baslik_satiri + 1, ws_d.max_row + 1):
                etiket = ws_d.cell(r, 1).value
                mape = ws_d.cell(r, 8).value
                if not etiket:
                    break
                if isinstance(mape, (int, float)):
                    dogruluk_satirlari.append((etiket, round(mape, 1)))
    except Exception as _exc:
        log_swallowed("web.tab_modules.verimlilik_gorselleri.render: beklenmeyen hata", _exc)
        tahmin_satirlari = []
        dogruluk_satirlari = []

    if tahmin_satirlari:
        metrik_adlari = [t[0] for t in tahmin_satirlari]
        metrik_secim = st.selectbox("Tahmin edilecek metrik", metrik_adlari, key="tahmin_metrik_secim")
        secili = next(t for t in tahmin_satirlari if t[0] == metrik_secim)
        fig_tahmin = px.bar(
            x=["+30 Gün", "+60 Gün", "+90 Gün", "+120 Gün"], y=secili[1],
            labels={"x": "Ufuk", "y": metrik_secim}, title=f"{metrik_secim} — İleriye Dönük Tahmin",
            color_discrete_sequence=["#4472C4"], text_auto=".2s",
        )
        st.plotly_chart(fig_tahmin, use_container_width=True)
    else:
        st.info("Tahmin verisi bulunamadı.")

    if dogruluk_satirlari:
        st.markdown("**Model Doğruluğu (MAPE — geriye dönük test, düşük olması iyidir):**")
        for etiket, mape in dogruluk_satirlari:
            st.write(f"- {etiket}: %{mape}")

