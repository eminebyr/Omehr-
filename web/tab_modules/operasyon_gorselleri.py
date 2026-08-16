"""Operasyon Görselleri sekmesi.

Bu modül, web/app.py içindeki eski "with tabs[N]:" bloğundan otomatik
olarak çıkarılmıştır. Kod davranışı değiştirilmeden taşınmıştır; tüm
paylaşılan durum (sheets, kullanıcı/rol bilgisi, fm/detail/stores/kpis,
servis fonksiyonları) web.context.PageContext üzerinden gelir.
"""
from __future__ import annotations

import streamlit as st
import pandas as pd
import plotly.express as px

import openpyxl
from web.context import PageContext
from services.safe_exec import log_swallowed


def render(ctx: PageContext) -> None:
    """Operasyon Görselleri sekmesinin içeriğini çizer."""
    sheets, acc = ctx.sheets, ctx.acc
    fm, detail, stores, kpis = ctx.fm, ctx.detail, ctx.stores, ctx.kpis
    user, username, role, scope, email = ctx.user, ctx.username, ctx.role, ctx.scope, ctx.email
    is_global = ctx.is_global
    can_view_personal_address = ctx.can_view_personal_address
    approval_level, can_approve = ctx.approval_level, ctx.can_approve
    ROOT, INPUT, OUTPUT, DB = ctx.root, ctx.input_path, ctx.output_path, ctx.db_path
    APPROVERS, BD_RENK = ctx.approvers, ctx.bd_renk
    db, log = ctx.db, ctx.log
    enqueue, job_status, tenant_code = ctx.enqueue, ctx.job_status, ctx.tenant_code
    norm_text, tr_number, tr_money_compact = ctx.norm_text, ctx.tr_number, ctx.tr_money_compact
    set_password, password_error = ctx.set_password, ctx.password_error
    refresh_home_proximity, maps_route = ctx.refresh_home_proximity, ctx.maps_route
    verify_password, transfer_recipients = ctx.verify_password, ctx.transfer_recipients
    cancel_transfer_request, redirect_transfer_request = ctx.cancel_transfer_request, ctx.redirect_transfer_request
    bulk_branch_mail_panel = ctx.bulk_branch_mail_panel
    _enqueue_and_process = ctx.enqueue_and_process
    read_input = ctx.read_input

    st.subheader("Operasyon Görselleri")
    st.caption("Günlük/aylık ciro, fiş, saatlik yoğunluk ve kasa kullanımı — interaktif grafikler.")

    def _read_two_row(sheet_name):
        try:
            from services.cached_excel_reader import read_sheet_cached
            df = read_sheet_cached(INPUT, sheet_name, header=1)
            df = df.dropna(how="all")
            return df
        except Exception as _exc:
            log_swallowed("web.tab_modules.operasyon_gorselleri._read_two_row: beklenmeyen hata", _exc)
            return pd.DataFrame()

    gunluk = _read_two_row("Günlük Operasyon")
    if not gunluk.empty:
        if not is_global:
            gunluk = gunluk[gunluk["Bölge"].astype(str).map(norm_text).eq(norm_text(scope))]
        try:
            from services.cached_excel_reader import read_sheet_cached
            kasa_ozet_metrik = read_sheet_cached(INPUT, "Kasa Kullanımı", header=1)
        except Exception as _exc:
            log_swallowed("web.tab_modules.operasyon_gorselleri._read_two_row: beklenmeyen hata", _exc)
            kasa_ozet_metrik = pd.DataFrame()
        oc1, oc2, oc3 = st.columns(3)
        oc1.metric("Dönem Toplam Ciro", tr_money_compact(gunluk["Ciro"].sum()))
        oc2.metric("Dönem Toplam Fiş Adedi", tr_number(gunluk["Fiş Adedi"].sum()))
        if not kasa_ozet_metrik.empty and "Kullanım Oranı %" in kasa_ozet_metrik.columns:
            oc3.metric("Ortalama Kasa Kullanım Oranı", tr_number(kasa_ozet_metrik["Kullanım Oranı %"].mean(),1,"%"))
        gunluk["Tarih"] = pd.to_datetime(gunluk["Tarih"], errors="coerce")
        gunluk_trend = gunluk.groupby("Tarih", as_index=False).agg(Ciro=("Ciro", "sum"), **{"Fiş Adedi": ("Fiş Adedi", "sum")})
        c1, c2 = st.columns(2)
        fig_ciro = px.line(gunluk_trend, x="Tarih", y="Ciro", markers=True, title="Günlük Toplam Ciro Trendi")
        c1.plotly_chart(fig_ciro, use_container_width=True)
        fig_fis = px.line(gunluk_trend, x="Tarih", y="Fiş Adedi", markers=True, title="Günlük Toplam Fiş Adedi Trendi", color_discrete_sequence=["#118B94"])
        c2.plotly_chart(fig_fis, use_container_width=True)
        magaza_ciro = gunluk.groupby("Mağaza", as_index=False)["Ciro"].sum().sort_values("Ciro", ascending=False).head(15)
        fig_top = px.bar(magaza_ciro, x="Ciro", y="Mağaza", orientation="h", text="Ciro", title="En Yüksek Ciro Yapan 15 Mağaza (Dönem Toplamı)", color_discrete_sequence=["#4472C4"])
        fig_top.update_traces(textposition="outside", cliponaxis=False)
        st.plotly_chart(fig_top, use_container_width=True)
    else:
        st.info("Günlük Operasyon verisi bulunamadı.")

    try:
        from services.cached_excel_reader import read_sheet_cached
        saatlik = read_sheet_cached(INPUT, "Saatlik Yoğunluk", header=1)
    except Exception as _exc:
        log_swallowed("web.tab_modules.operasyon_gorselleri.render: beklenmeyen hata", _exc)
        saatlik = pd.DataFrame()
    if not saatlik.empty:
        magazalar_opt = sorted(saatlik["Mağaza"].dropna().unique().tolist())
        secili_magazalar = st.multiselect("Saatlik yoğunluk için mağaza seç (boş bırakılırsa ilk 10)", magazalar_opt, key="saatlik_magaza_secim")
        gosterilecek = secili_magazalar if secili_magazalar else magazalar_opt[:10]
        heat_df = saatlik[saatlik["Mağaza"].isin(gosterilecek)]
        pivot = heat_df.pivot_table(index="Mağaza", columns="Saat", values="Yoğunluk Skoru", aggfunc="mean")
        fig_heat = px.imshow(pivot, aspect="auto", color_continuous_scale="OrRd", title="Saatlik Yoğunluk Skoru (Mağaza x Saat)", labels=dict(color="Yoğunluk"))
        # Az mağaza + çok saat sütunu durumunda hücrelerin ince/sütun gibi görünmesini
        # önlemek için yükseklik satır sayısına göre ayarlanır.
        fig_heat.update_layout(height=max(420,55*len(pivot.index)+120))
        st.plotly_chart(fig_heat, use_container_width=True)
    else:
        st.info("Saatlik Yoğunluk verisi bulunamadı.")

    try:
        from services.cached_excel_reader import read_sheet_cached
        kasa = read_sheet_cached(INPUT, "Kasa Kullanımı", header=1)
    except Exception as _exc:
        log_swallowed("web.tab_modules.operasyon_gorselleri.render: beklenmeyen hata", _exc)
        kasa = pd.DataFrame()
    if not kasa.empty:
        kasa_ozet = kasa.groupby("Mağaza", as_index=False)["Kullanım Oranı %"].mean().sort_values("Kullanım Oranı %", ascending=False).head(20)
        fig_kasa = px.bar(kasa_ozet, x="Kullanım Oranı %", y="Mağaza", orientation="h", text="Kullanım Oranı %", title="Ortalama Kasa Kullanım Oranı (%) — Mağaza Bazında", color="Kullanım Oranı %", color_continuous_scale="Blues")
        fig_kasa.update_traces(textposition="outside", cliponaxis=False)
        st.plotly_chart(fig_kasa, use_container_width=True)
    else:
        st.info("Kasa Kullanımı verisi bulunamadı.")

    st.markdown("---")
    st.markdown("#### Enflasyona Göre Büyüme — Mağaza Mağaza")
    DEPO_HARIC = {"Z1  ET DEPO", "Z1 ET DEPO", "MERKEZ", "ÇİFTLİK", "BUCA2", "HAL DEPO", "MENDERES DEPO"}
    try:
        from services.cached_excel_reader import read_sheet_cached
        aylik = read_sheet_cached(INPUT, "Aylık Operasyon KPI", header=1).dropna(how="all")
    except Exception as _exc:
        log_swallowed("web.tab_modules.operasyon_gorselleri.render: beklenmeyen hata", _exc)
        aylik = pd.DataFrame()
    if not aylik.empty:
        aylik = aylik[~aylik["Mağaza"].astype(str).str.strip().isin(DEPO_HARIC)]
        if not is_global:
            aylik = aylik.merge(sheets.get("Dim_Magaza", pd.DataFrame())[["MağazaID", "Bölge Sorumlusu"]], left_on="MagazaID", right_on="MağazaID", how="left")
            aylik = aylik[aylik["Bölge Sorumlusu"].astype(str).map(norm_text).eq(norm_text(scope))]
        aylik = aylik.sort_values("Ay")
        ilk_ay, son_ay = aylik["Ay"].min(), aylik["Ay"].max()
        c1, c2 = st.columns([1, 3])
        enflasyon_orani = c1.number_input(
            "Yıllık enflasyon oranı (%) — TÜİK, Haziran 2026: %32,11",
            min_value=0.0, max_value=200.0, value=32.11, step=0.1, key="enflasyon_input",
        )
        ilk = aylik[aylik["Ay"] == ilk_ay].set_index("Mağaza")["Aylık Ciro"]
        son = aylik[aylik["Ay"] == son_ay].set_index("Mağaza")["Aylık Ciro"]
        nominal_oran = (son / ilk) - 1
        enflasyon_carpani = 1 + (enflasyon_orani / 100.0)
        reel_oran = ((1 + nominal_oran) / enflasyon_carpani) - 1
        buyume_df = pd.DataFrame({
            "Nominal Büyüme %": (nominal_oran * 100).round(1),
            "Reel Büyüme %": (reel_oran * 100).round(1),
        }).replace([float("inf"), float("-inf")], pd.NA).dropna().reset_index().sort_values("Reel Büyüme %", ascending=False)
        pozitif_sayisi = int((buyume_df["Reel Büyüme %"] > 0).sum())
        c2.info(
            f"{ilk_ay} → {son_ay} dönemi | Enflasyon %{tr_number(enflasyon_orani,1)} | "
            f"{tr_number(pozitif_sayisi)} / {tr_number(len(buyume_df))} mağaza enflasyonun üzerinde (reel) büyüdü. "
            "Reel büyüme bileşik formülle hesaplanır: (1+nominal)/(1+enflasyon)-1."
        )
        fig_reel = px.bar(
            buyume_df, x="Reel Büyüme %", y="Mağaza", orientation="h", text="Reel Büyüme %",
            title=f"Mağaza Bazında Reel Büyüme ({ilk_ay} → {son_ay}, Enflasyon %{tr_number(enflasyon_orani,1)} Baz Alınarak)",
            color="Reel Büyüme %", color_continuous_scale=["#C00000", "#F2F2F2", "#70AD47"], color_continuous_midpoint=0,
            height=max(500, 22 * len(buyume_df)),
        )
        fig_reel.update_traces(textposition="outside", cliponaxis=False)
        fig_reel.add_vline(x=0, line_dash="dash", line_color="grey")
        st.plotly_chart(fig_reel, use_container_width=True)
        with st.expander("Nominal büyüme tablosu (enflasyondan arındırılmamış)"):
            st.dataframe(buyume_df[["Mağaza", "Nominal Büyüme %", "Reel Büyüme %"]], use_container_width=True, hide_index=True)
    else:
        st.info("Aylık Operasyon KPI verisi bulunamadı.")

    st.markdown("---")
    st.markdown("#### Ciro — 30/60/90/120 Gün İleriye Dönük Tahmin (Dinamik)")
    st.caption(
        "Excel FORECAST fonksiyonuyla, 12 aylık geçmiş veriden hesaplanır (Verimlilik_Operasyon_Tahmini sayfası). "
        "Dosya her güncellendiğinde bu grafik de otomatik yenilenir."
    )
    try:
        from services.cached_excel_reader import read_workbook_cached
        wb_ciro_tahmin = read_workbook_cached(INPUT, data_only=True)
        ws_ct = wb_ciro_tahmin["Verimlilik_Operasyon_Tahmini"]
        _ciro_satiri = None
        for r in range(1, ws_ct.max_row + 1):
            if ws_ct.cell(r, 1).value == "Toplam Ciro (TL) (Tahmin)":
                _ciro_satiri = r
                break
        if _ciro_satiri:
            _ciro_degerler = [ws_ct.cell(_ciro_satiri, c).value for c in range(2, 6)]
            if all(v is not None for v in _ciro_degerler):
                fig_ciro_gelecek = px.bar(
                    x=["+30 Gün", "+60 Gün", "+90 Gün", "+120 Gün"], y=_ciro_degerler,
                    labels={"x": "Ufuk", "y": "Toplam Ciro (TL)"}, title="Toplam Ciro — İleriye Dönük Tahmin",
                    color_discrete_sequence=["#4472C4"], text_auto=".2s",
                )
                st.plotly_chart(fig_ciro_gelecek, use_container_width=True)
                st.caption(
                    "⚠️ Bu zaman-trend tahmininin istatistiksel anlamlılığı düşüktür (R²≈0, bkz. "
                    "Istatistiksel_Model_Testi). Verimlilik Görselleri sekmesindeki Ciro-tabanlı İş Yükü "
                    "Endeksi projeksiyonu (r=0,91) daha güvenilirdir."
                )
    except Exception as _exc:
        log_swallowed("web.tab_modules.operasyon_gorselleri.render: beklenmeyen hata", _exc)
        pass

