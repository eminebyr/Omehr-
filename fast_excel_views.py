from __future__ import annotations

"""Web panelindeki yalnız-okuma Excel görselleri için hafif önbellek.

Ana veri kaynağı Excel olarak kalır. Bu modül veri yazmaz; yalnızca pahalı
``openpyxl.load_workbook`` çağrılarını dosya değişmediği sürece tekrar etmez.
"""

from functools import lru_cache
from pathlib import Path

import openpyxl


def _sig(path: str | Path) -> tuple[str, int, int]:
    p = Path(path).resolve()
    st = p.stat()
    return str(p), int(st.st_mtime_ns), int(st.st_size)


def forecast_payload(path: str | Path) -> dict:
    p, mtime_ns, size = _sig(path)
    return _forecast_payload_cached(p, mtime_ns, size)


@lru_cache(maxsize=4)
def _forecast_payload_cached(path_text: str, mtime_ns: int, size: int) -> dict:
    """Tahmin sayfalarından webde kullanılan küçük veri özetini tek açılışta al."""
    wb = openpyxl.load_workbook(path_text, data_only=True, read_only=True)
    try:
        result = {
            "ciro": None,
            "isyuku": None,
            "tahmin_satirlari": [],
            "dogruluk_satirlari": [],
        }

        if "Verimlilik_Operasyon_Tahmini" in wb.sheetnames:
            ws = wb["Verimlilik_Operasyon_Tahmini"]
            for row in ws.iter_rows(values_only=True):
                if not row:
                    continue
                etiket = row[0]
                if etiket == "Toplam Ciro (TL) (Tahmin)":
                    vals = list(row[1:5])
                    if len(vals) == 4 and all(v is not None for v in vals):
                        result["ciro"] = vals
                elif etiket == "İş Yükü Endeksi (Ciro Tahmininden Türetilmiş)":
                    vals = list(row[1:5])
                    if len(vals) == 4 and all(v is not None for v in vals):
                        result["isyuku"] = vals
                elif isinstance(etiket, str) and etiket.endswith("(Tahmin)"):
                    vals = list(row[1:5])
                    if len(vals) == 4 and all(v is not None for v in vals):
                        result["tahmin_satirlari"].append((etiket.replace(" (Tahmin)", ""), vals))

        if "Tahmin_Dogruluk_Testi" in wb.sheetnames:
            ws = wb["Tahmin_Dogruluk_Testi"]
            header_seen = False
            for row in ws.iter_rows(values_only=True):
                if not row:
                    continue
                if not header_seen:
                    if row[0] == "Metrik":
                        header_seen = True
                    continue
                etiket = row[0]
                if not etiket:
                    break
                mape = row[7] if len(row) > 7 else None
                if isinstance(mape, (int, float)):
                    result["dogruluk_satirlari"].append((etiket, round(float(mape), 1)))

        return result
    finally:
        wb.close()


def clear() -> None:
    _forecast_payload_cached.cache_clear()
