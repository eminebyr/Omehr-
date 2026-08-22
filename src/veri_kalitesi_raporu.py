from __future__ import annotations

"""
VERİ KALİTESİ RAPORU (V19.9 — dış inceleme sonrası eklendi)
=====================================================================
SORUN: Sistem, dummy adresler/saha etüdü bekleyen süreler/dummy mail
adresleri/tanımsız unvanlar gibi veri kalitesi sorunlarını KOD İÇİNDE
biliyordu (loglara "sessiz uyarı" olarak yazıyordu — bkz. services/
safe_exec.py::log_swallowed çağrıları) ama bir yöneticinin kolayca
açıp okuyabileceği AYRI, GÖRÜNÜR bir özet dosyası YOKTU. Bu modül o
boşluğu kapatır: input dosyasındaki bilinen 4 veri kalitesi işaretini
tarar ve output/OMEHR_Veri_Kalitesi_Raporu.xlsx üretir.

Bu modül YENİ veri kalitesi sorunu ÜRETMEZ/TAHMİN ETMEZ — sadece
zaten var olan işaretleri (ör. "Veri Durumu" sütunundaki "Dummy" metni)
sayar ve okunabilir bir tabloya döker. Yani bu bir RAPORLAMA katmanıdır,
veri düzeltme/tahmin katmanı değildir.
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from src.excel_report import write_df
from src.text_utils import product_name

REPORT_FILENAME = "OMEHR_Veri_Kalitesi_Raporu.xlsx"


def _oran(sayi: int, toplam: int) -> str:
    if not toplam:
        return "%0.0"
    return f"%{(sayi / toplam * 100):.1f}"


def data_quality_findings(sheets: dict) -> dict:
    """Bilinen 4 veri kalitesi işaretini tarar, hem özet hem detay
    DataFrame'lerini döndürür. Excel üretmeden de (ör. testte) kullanılabilir."""
    bulgular = {}

    adres = sheets.get("Personel_Adresleri", pd.DataFrame())
    if not adres.empty and "Veri Durumu" in adres.columns:
        durum = adres["Veri Durumu"].astype(str)
        dummy_maske = durum.str.contains("Dummy", case=False, na=False)
        bulgular["adres"] = {
            "sayi": int(dummy_maske.sum()),
            "toplam": int(len(adres)),
            "detay": adres.loc[dummy_maske, [c for c in ["PersonelID", "İsim Soyisim", "Mevcut Mağaza", "Veri Durumu"] if c in adres.columns]],
        }
    else:
        bulgular["adres"] = {"sayi": 0, "toplam": 0, "detay": pd.DataFrame()}

    sure = sheets.get("Standart_Sure_Kutuphanesi", pd.DataFrame())
    if not sure.empty and "Kaynak" in sure.columns:
        saha_maske = sure["Kaynak"].astype(str).str.contains("Saha Etüdü Bekleniyor", case=False, na=False)
        bulgular["sure"] = {
            "sayi": int(saha_maske.sum()),
            "toplam": int(len(sure)),
            "detay": sure.loc[saha_maske, [c for c in ["AktiviteID", "Unvan", "Aktivite", "Standart Süre (Dk)", "Kaynak"] if c in sure.columns]],
        }
    else:
        bulgular["sure"] = {"sayi": 0, "toplam": 0, "detay": pd.DataFrame()}

    mail = sheets.get("Mail_Listesi", pd.DataFrame())
    if not mail.empty and "E-posta" in mail.columns:
        dummy_mail_maske = mail["E-posta"].astype(str).str.contains("dummy", case=False, na=False)
        bulgular["mail"] = {
            "sayi": int(dummy_mail_maske.sum()),
            "toplam": int(len(mail)),
            "detay": mail.loc[dummy_mail_maske, [c for c in ["Bölge", "Sorumlu", "E-posta", "Aktif"] if c in mail.columns]],
        }
    else:
        bulgular["mail"] = {"sayi": 0, "toplam": 0, "detay": pd.DataFrame()}

    norm = sheets.get("Fact_Norm", pd.DataFrame())
    if not norm.empty and "Unvan" in norm.columns:
        tanimsiz_maske = norm["Unvan"].astype(str).str.contains("TANIMSIZ", case=False, na=False)
        bulgular["unvan"] = {
            "sayi": int(tanimsiz_maske.sum()),
            "toplam": int(len(norm)),
            "detay": norm.loc[tanimsiz_maske, [c for c in ["MağazaID", "Mağaza", "UnvanID", "Unvan", "Norm Kadro"] if c in norm.columns]],
        }
    else:
        bulgular["unvan"] = {"sayi": 0, "toplam": 0, "detay": pd.DataFrame()}

    return bulgular


def _ozet_tablosu(bulgular: dict) -> pd.DataFrame:
    satirlar = [
        {
            "Kategori": "Dummy Personel Adresi",
            "Sayfa": "Personel_Adresleri",
            "Sayı": bulgular["adres"]["sayi"],
            "Toplam Satır": bulgular["adres"]["toplam"],
            "Oran": _oran(bulgular["adres"]["sayi"], bulgular["adres"]["toplam"]),
            "Etki": "Transfer/yakınlık puanları gerçek ev konumunu yansıtmıyor",
            "Önerilen Aksiyon": "İK'dan gerçek adres/koordinat toplanmalı",
        },
        {
            "Kategori": "Saha Etüdü Bekleyen Standart Süre",
            "Sayfa": "Standart_Sure_Kutuphanesi",
            "Sayı": bulgular["sure"]["sayi"],
            "Toplam Satır": bulgular["sure"]["toplam"],
            "Oran": _oran(bulgular["sure"]["sayi"], bulgular["sure"]["toplam"]),
            "Etki": "AI Önerilen Norm, doğrulanmamış süre varsayımına dayanıyor",
            "Önerilen Aksiyon": "Operasyon ekibi öncelikli aktiviteler için gerçek zaman etüdü yapmalı",
        },
        {
            "Kategori": "Dummy E-posta Adresi",
            "Sayfa": "Mail_Listesi",
            "Sayı": bulgular["mail"]["sayi"],
            "Toplam Satır": bulgular["mail"]["toplam"],
            "Oran": _oran(bulgular["mail"]["sayi"], bulgular["mail"]["toplam"]),
            "Etki": "Bu adresler otomatik dışlanır (mail gitmez) — kapsam eksik kalabilir",
            "Önerilen Aksiyon": "Gerçek Outlook adresleriyle değiştirilmeli",
        },
        {
            "Kategori": "Tanımsız Unvan (Fact_Norm)",
            "Sayfa": "Fact_Norm",
            "Sayı": bulgular["unvan"]["sayi"],
            "Toplam Satır": bulgular["unvan"]["toplam"],
            "Oran": _oran(bulgular["unvan"]["sayi"], bulgular["unvan"]["toplam"]),
            "Etki": "Bu satırlar için norm/mevcut karşılaştırması güvenilmez",
            "Önerilen Aksiyon": "İK ilgili mağaza için gerçek unvanı teyit etmeli",
        },
    ]
    return pd.DataFrame(satirlar)


def generate_data_quality_report(sheets: dict, outdir) -> Path:
    bulgular = data_quality_findings(sheets)
    ozet = _ozet_tablosu(bulgular)

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Özet", 0)
    ws.merge_cells("A1:G1")
    ws["A1"] = product_name().upper() + " - VERİ KALİTESİ ÖZETİ"
    ws["A1"].fill = PatternFill("solid", fgColor="B00000")
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"] = ("Bu rapor, sistemin ZATEN bildiği ama daha önce ayrı/görünür bir dosyada "
                "sunulmayan bilinen veri kalitesi işaretlerini özetler. Bunlar KOD HATASI "
                "DEĞİLDİR — gerçek saha/İK veri toplama işleridir.")
    ws.merge_cells("A2:G2")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 30

    headers = list(ozet.columns)
    header_row = 4
    for col_index, header in enumerate(headers, 1):
        ws.cell(header_row, col_index, header)
    for cell in ws[header_row]:
        cell.fill = PatternFill("solid", fgColor="102F64")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for row_offset, (_, row) in enumerate(ozet.iterrows(), header_row + 1):
        for col_index, value in enumerate(row, 1):
            ws.cell(row_offset, col_index, value)
    widths = {"A": 34, "B": 26, "C": 8, "D": 12, "E": 8, "F": 48, "G": 48}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    for sheet_name, key in [
        ("Detay - Dummy Adresler", "adres"),
        ("Detay - Saha Etüdü Bekleyen Süreler", "sure"),
        ("Detay - Dummy Mail Adresleri", "mail"),
        ("Detay - Tanımsız Unvanlar", "unvan"),
    ]:
        detay = bulgular[key]["detay"]
        write_df(wb, sheet_name, detay if not detay.empty else pd.DataFrame([{"Durum": "Bulunamadı — bu kategori temiz"}]))

    outdir = Path(outdir)
    outdir.mkdir(parents=True, exist_ok=True)
    path = outdir / REPORT_FILENAME
    wb.save(path)
    return path
