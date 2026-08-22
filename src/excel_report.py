from __future__ import annotations

"""
EXCEL RAPOR ÜRETİM KATMANI (P2 — engine_core.py modülerleştirme, dördüncü adım)
=====================================================================
Yönetici/bölge Excel çıktısını (openpyxl) üreten fonksiyonlar. engine_core.py
içindeki state()/kpis() hesaplama çekirdeğine bağımlı DEĞİLDİR — girdi olarak
zaten hesaplanmış kpi/st/tt/ai/staff DataFrame'lerini alır. _title_key artık
src/text_utils.py'de tanımlıdır (saf fonksiyon), bu yüzden engine_core.py'ye
dönüp dolanan bir bağımlılık YOKTUR.
"""

import math
import shutil
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from services.runtime_paths import runtime_root
from services.personnel_notes import format_person_note, note_kind
from services.family_balance import balance_detail_table, balance_store_title_rows
from src.text_utils import canon, col, numeric, product_name, req, txt, unvan_sirali, _store_key, _title_key



def _store_columns_adjacent(df):
    """Raporlarda MağazaID görüldüğünde Mağaza adını hemen yanına taşır."""
    x=df.copy()
    id_col=next((c for c in ('MağazaID','MagazaID','Mağaza ID','Magaza ID') if c in x.columns),None)
    name_col=next((c for c in ('Mağaza','Magaza','Mağaza Adı','Magaza Adi') if c in x.columns),None)
    if id_col and name_col:
        cols=list(x.columns)
        cols.remove(name_col)
        cols.insert(cols.index(id_col)+1,name_col)
        x=x[cols]
    return x


def write_df(wb,name,df):
    df=_store_columns_adjacent(df)
    ws=wb.create_sheet(name[:31]); ws.append(list(df.columns))
    for row in df.itertuples(index=False,name=None):ws.append(list(row))
    navy='102F64'; thin=Side(style='thin',color='D9E1F2'); note_yellow='FFF2CC'
    for c in ws[1]:c.fill=PatternFill('solid',fgColor=navy);c.font=Font(color='FFFFFF',bold=True);c.alignment=Alignment(horizontal='center',wrap_text=True)
    for row in ws.iter_rows():
        for c in row:c.border=Border(left=thin,right=thin,top=thin,bottom=thin);c.alignment=Alignment(vertical='top',wrap_text=True)
    # Açıklama bulunan satırları sarı vurgula ve ilgili kişi/unvan hücresine Excel notu ekle.
    headers={txt(ws.cell(1,i).value):i for i in range(1,ws.max_column+1)}
    note_idx=next((headers[h] for h in ('Açıklama','Aciklama','AÇIKLAMA','ACIKLAMA') if h in headers),None)
    target_idx=next((headers[h] for h in ('Personel Adı Soyadı','İsim Soyisim','Ad Soyad','Unvan','Departman') if h in headers),1)
    if note_idx:
        for r in range(2,ws.max_row+1):
            note=txt(ws.cell(r,note_idx).value).strip()
            if not note:
                continue
            person_name=txt(ws.cell(r,target_idx).value).strip()
            sentence=format_person_note(person_name,note)
            fill_color='F4CCCC' if note_kind(note)=='departure' else note_yellow
            for c in range(1,ws.max_column+1):
                ws.cell(r,c).fill=PatternFill('solid',fgColor=fill_color)
            ws.cell(r,target_idx).comment=Comment(sentence or note,'OMEHR Personel Açıklaması')
    ws.freeze_panes='A2';ws.auto_filter.ref=ws.dimensions
    for i in range(1,ws.max_column+1):ws.column_dimensions[get_column_letter(i)].width=min(38,max(11,max(len(txt(ws.cell(r,i).value)) for r in range(1,min(ws.max_row,150)+1))+2))
    return ws


def _personnel_names_by_store_title(staff):
    """Fact_Mevcut içinden mağaza+departman bazında ad ve gerçek unvanları birleştirir."""
    sm=req(staff,'Mağaza','Magaza')
    dep=req(staff,'Departman')
    actual=req(staff,'Unvan')
    pname=req(staff,'İsim Soyisim','Isim Soyisim','Ad Soyad')
    cols=[sm,dep,pname,actual]
    x=staff[cols].copy()
    x.columns=['Mağaza','Departman','Personel Adı Soyadı','Gerçek Unvan']
    x['_Mağaza']=x['Mağaza'].map(_store_key)
    x['_Unvan']=x['Departman'].map(_title_key)
    x['Personel Adı Soyadı']=x['Personel Adı Soyadı'].map(txt)
    x['Gerçek Unvan']=x['Gerçek Unvan'].map(txt)
    x=x[x['Personel Adı Soyadı']!='']
    if x.empty:
        return pd.DataFrame(columns=['_Mağaza','_Unvan','Personel Adı Soyadı','Gerçek Unvanlar','Gerçek Unvan / Personel'])
    x['Gerçek Unvan / Personel']=x.apply(lambda r:f"{r['Personel Adı Soyadı']} ({r['Gerçek Unvan']})",axis=1)
    return (x.groupby(['_Mağaza','_Unvan'],dropna=False)
             .agg(**{
                 'Personel Adı Soyadı':('Personel Adı Soyadı',lambda values:', '.join(dict.fromkeys(v for v in values if v))),
                 'Gerçek Unvanlar':('Gerçek Unvan',lambda values:', '.join(dict.fromkeys(v for v in values if v))),
                 'Gerçek Unvan / Personel':('Gerçek Unvan / Personel',lambda values:', '.join(dict.fromkeys(v for v in values if v))),
             }).reset_index())



def _title_report_with_names(tt,staff):
    """Teknik UnvanID yerine personel ad-soyadını kullanıcı raporuna taşır."""
    x=tt.copy()
    x['_Mağaza']=x['Mağaza'].map(_store_key)
    x['_Unvan']=x['Unvan'].map(_title_key)
    names=_personnel_names_by_store_title(staff)
    x=x.merge(names,on=['_Mağaza','_Unvan'],how='left')
    for c in ['Personel Adı Soyadı','Gerçek Unvanlar','Gerçek Unvan / Personel']:
        x[c]=x[c].fillna('')
    x=x.drop(columns=['_Mağaza','_Unvan'],errors='ignore')
    return x



def _personnel_detail_report(tt,staff):
    """Mağaza-departman satırlarını gerçek unvan ve ad-soyadla kişi bazına açar."""
    sm=req(staff,'Mağaza','Magaza'); dep=req(staff,'Departman')
    actual=req(staff,'Unvan'); pname=req(staff,'İsim Soyisim','Isim Soyisim','Ad Soyad')
    staff_note=col(staff,'Açıklama','Aciklama','AÇIKLAMA','ACIKLAMA','Personel Açıklaması','Personel Aciklamasi')
    people_cols=[sm,dep,actual,pname]+([staff_note] if staff_note else [])
    people=staff[people_cols].copy()
    people.columns=['_Personel Mağaza','Departman','Unvan','Personel Adı Soyadı']+(['Açıklama'] if staff_note else [])
    if 'Açıklama' not in people.columns:
        people['Açıklama']=''
    people['_Mağaza']=people['_Personel Mağaza'].map(_store_key)
    people['_Unvan']=people['Departman'].map(_title_key)
    people['Personel Adı Soyadı']=people['Personel Adı Soyadı'].map(txt)
    people['Unvan']=people['Unvan'].map(txt)
    base=tt.copy()
    base['_Mağaza']=base['Mağaza'].map(_store_key)
    base['_Unvan']=base['Unvan'].map(_title_key)
    base=base.rename(columns={'Unvan':'Departman'})
    out=base.merge(
        people[['_Mağaza','_Unvan','Unvan','Personel Adı Soyadı','Açıklama']],
        on=['_Mağaza','_Unvan'],how='left'
    )
    out['Personel Adı Soyadı']=out['Personel Adı Soyadı'].fillna('')
    out['Unvan']=out['Unvan'].fillna('')
    out['Açıklama']=out['Açıklama'].fillna('')
    # Her personel satırı bir mevcut kişiyi temsil eder. Departman norm/eksik/fazla
    # değeri yalnızca ilk satırda gösterilir; böylece Excel toplamları şişmez.
    out['_Sıra']=out.groupby(['_Mağaza','_Unvan'],dropna=False).cumcount()
    has_person=out['Personel Adı Soyadı'].map(txt).ne('')
    out['Aktif Mevcut']=has_person.astype(int)
    for c in ['Norm Kadro','Norm Eksiği','Norm Fazlası','Net Fark']:
        out[c]=out[c].astype(object)
        out.loc[out['_Sıra']>0,c]=''
    return out.drop(columns=['_Mağaza','_Unvan','_Sıra'],errors='ignore')



def _surplus_people_report(tt,staff):
    """Norm fazlası kadar personeli, en yeni işe girişten başlayarak listeler."""
    # SÜTUN GARANTİSİ: aşağıdaki döngü hiç eşleşme bulamazsa (ör. bir bölgede
    # hiç "Norm Fazlası>0" satırı yoksa) `rows` boş kalır ve pd.DataFrame([])
    # HİÇBİR SÜTUNU OLMAYAN bir DataFrame üretir. Bu durumda çağıran taraf
    # surplus_source[surplus_cols] ile sütun seçmeye çalışınca "None of [...]
    # are in the columns" hatasıyla çöker. Bunu önlemek için sabit sütun
    # listesiyle BOŞ (ama doğru şemalı) bir DataFrame her zaman garanti edilir.
    surplus_cols=['Bölge Sorumlusu','MağazaID','Mağaza','Departman','Unvan','Personel Adı Soyadı','Açıklama','Aktif Mevcut','Norm Kadro','Norm Eksiği','Norm Fazlası','Net Fark']
    sm=req(staff,'Mağaza','Magaza'); dep=req(staff,'Departman')
    actual=req(staff,'Unvan'); pname=req(staff,'İsim Soyisim','Isim Soyisim','Ad Soyad')
    entry=col(staff,'İşe Giriş','Ise Giris')
    rows=[]
    for _,r in tt[numeric(tt['Norm Fazlası'])>0].iterrows():
        candidates=staff[
            (staff[sm].map(_store_key)==_store_key(r['Mağaza'])) &
            (staff[dep].map(_title_key)==_title_key(r['Unvan']))
        ].copy()
        if entry:
            candidates['_İşe Giriş']=pd.to_datetime(candidates[entry],errors='coerce')
            candidates=candidates.sort_values('_İşe Giriş',ascending=False,na_position='last')
        candidates=candidates.head(int(r['Norm Fazlası']))
        if candidates.empty:
            item=r.to_dict()
            item.update({'Departman':r['Unvan'],'Unvan':'','Personel Adı Soyadı':''})
            rows.append(item)
            continue
        for index,(_,person) in enumerate(candidates.iterrows()):
            item=r.to_dict()
            item.update({
                'Departman':r['Unvan'],
                'Unvan':txt(person[actual]),
                'Personel Adı Soyadı':txt(person[pname]),
            })
            if index>0:
                for c in ['Aktif Mevcut','Norm Kadro','Norm Eksiği','Norm Fazlası','Net Fark']:
                    item[c]=''
            rows.append(item)
    if not rows:
        return pd.DataFrame(columns=surplus_cols)
    out=pd.DataFrame(rows)
    for c in surplus_cols:
        if c not in out.columns:
            out[c]=''
    return out



def executive_excel(kpi,st,tt,scens,risk,hash_,ai=None,validation=None,validation_summary=None,input_sheets=None,staff=None):
    tt=balance_detail_table(tt)
    """Sadece Fact_Mevcut ve Fact_Norm karşılaştırmasını içeren sade yönetici Excel'i."""
    out=runtime_root()/'output'/'OMEHR_Executive_Data.xlsx'
    wb=Workbook(); wb.remove(wb.active)
    ws=wb.create_sheet('Yönetici Özeti')
    ws.append([product_name().upper(),'']); ws.merge_cells('A1:B1')
    ws.append(['KPI','Değer'])
    labels=[('Aktif Mevcut','Aktif Mevcut'),('Yönetim Normu','Toplam Norm'),('Norm Eksiği','Norm Eksiği'),('Norm Fazlası','Norm Fazlası'),('Net İhtiyaç','Net İhtiyaç')]
    for label,key in labels: ws.append([label,int(kpi[key])])
    ws.append([]); ws.append(['Kaynak SHA256',hash_]); ws.append(['Üretim Zamanı',datetime.now().strftime('%d.%m.%Y %H:%M:%S')])
    ws['A1'].fill=PatternFill('solid',fgColor='102F64'); ws['A1'].font=Font(color='FFFFFF',bold=True,size=14)
    for c in ws[2]: c.fill=PatternFill('solid',fgColor='4472C4'); c.font=Font(color='FFFFFF',bold=True)
    ws.column_dimensions['A'].width=30; ws.column_dimensions['B'].width=70

    store_cols=['Bölge Sorumlusu','MağazaID','Mağaza','Aktif Mevcut','Norm Kadro','Norm Eksiği','Norm Fazlası','Net Fark']
    title_source=_personnel_detail_report(tt,staff) if staff is not None else tt.rename(columns={'Unvan':'Departman'}).assign(
        **{'Unvan':'','Personel Adı Soyadı':''}
    )
    deficit_source=tt.rename(columns={'Unvan':'Departman'})
    surplus_source=_surplus_people_report(tt,staff) if staff is not None else pd.DataFrame()
    title_cols=['Bölge Sorumlusu','MağazaID','Mağaza','Departman','Unvan','Personel Adı Soyadı','Açıklama','Aktif Mevcut','Norm Kadro','Norm Eksiği','Norm Fazlası','Net Fark']
    deficit_cols=['Bölge Sorumlusu','MağazaID','Mağaza','Departman','Aktif Mevcut','Norm Kadro','Norm Eksiği','Norm Fazlası','Net Fark']
    surplus_cols=['Bölge Sorumlusu','MağazaID','Mağaza','Departman','Unvan','Personel Adı Soyadı','Açıklama','Aktif Mevcut','Norm Kadro','Norm Eksiği','Norm Fazlası','Net Fark']
    write_df(wb,'Mağaza Bazlı',st[store_cols].sort_values(['Bölge Sorumlusu','Mağaza']))
    write_df(wb,'Mağaza-Unvan Bazlı',unvan_sirali(title_source[title_cols],['Bölge Sorumlusu','Mağaza'],unvan_kolonu='Departman'))
    region=st.groupby('Bölge Sorumlusu',dropna=False)[['Aktif Mevcut','Norm Kadro','Norm Eksiği','Norm Fazlası']].sum().reset_index()
    write_df(wb,'Bölge Müdürü Bazlı',region.sort_values('Bölge Sorumlusu'))
    write_df(wb,'Norm Eksikleri',deficit_source.loc[deficit_source['Norm Eksiği']>0,deficit_cols].sort_values(['Bölge Sorumlusu','Mağaza','Norm Eksiği'],ascending=[True,True,False]))
    write_df(wb,'Norm Fazlaları',surplus_source[surplus_cols].sort_values(['Bölge Sorumlusu','Mağaza'],ascending=[True,True]))
    if input_sheets is not None:
        inventory=[]
        for sheet_name,frame in input_sheets.items():
            inventory.append({'Sayfa':str(sheet_name),'Satır Sayısı':int(len(frame)),'Sütun Sayısı':int(len(frame.columns)),'Kaynak SHA256':hash_})
        write_df(wb,'Input Envanteri',pd.DataFrame(inventory))
    for scenario_name,frame in (scens or {}).items():
        if frame is not None and not frame.empty:
            write_df(wb,('Transfer - '+scenario_name)[:31],frame)
    if risk is not None and not risk.empty:
        write_df(wb,'Transfer Riskleri',risk)
    wb.save(out)
    return out


# ============================================================================
# FINAL OUTPUT LAYER
# Dynamic KPI + management norm / AI norm separation + professional reports.
# ============================================================================



def _store_roster_rows(store_id, store_name, norm, staff, ai):
    nmid=req(norm,'MağazaID','MagazaID'); nuid=req(norm,'UnvanID'); nu=req(norm,'Unvan'); nn=req(norm,'Norm Kadro','Norm')
    smid=req(staff,'MağazaID','MagazaID'); suid=req(staff,'UnvanID'); su=req(staff,'Unvan'); pname=req(staff,'İsim Soyisim','Isim Soyisim','Ad Soyad')
    entry=col(staff,'İşe Giriş','Ise Giris')
    ns=norm[norm[nmid].map(txt)==txt(store_id)].copy(); ps=staff[staff[smid].map(txt)==txt(store_id)].copy()
    all_titles=[]
    for _,r in ns.iterrows():
        key=txt(r[nuid]); all_titles.append((key,txt(r[nu]),int(float(r[nn] or 0))))
    known={x[0] for x in all_titles}
    for _,r in ps.iterrows():
        key=txt(r[suid])
        if key not in known: all_titles.append((key,txt(r[su]),0)); known.add(key)
    rows=[]
    for uid,title,norm_count in all_titles:
        people=ps[ps[suid].map(txt)==uid].copy()
        if entry:
            people['_entry']=pd.to_datetime(people[entry],errors='coerce')
            people=people.sort_values(['_entry',pname],ascending=[True,True])
        else:
            people=people.sort_values(pname)
        excess=max(len(people)-norm_count,0)
        statuses=['MEVCUT']*(len(people)-excess)+['FAZLA']*excess
        first=True
        for (_,person),status in zip(people.iterrows(),statuses):
            rows.append([txt(person[pname]),title,norm_count if first else '',status])
            first=False
        for i in range(max(norm_count-len(people),0)):
            rows.append([f'BOŞ POZİSYON {i+1}',title,norm_count if first else '','EKSİK'])
            first=False
    return rows



def _gap_text(mevcut, ai_norm):
    gap=int(ai_norm)-int(mevcut)
    if gap>0:return f'AI\'ya göre {gap} kişi eksik'
    if gap<0:return f'AI\'ya göre {abs(gap)} kişi fazla'
    return 'AI normu ile mevcut eşit'




def _region_excel(path, region, kpi, st, tt, ai, staff):
    tt=balance_detail_table(tt)
    """Bölge müdürü için yalnızca mevcut, norm, eksik ve fazla raporu."""
    wb=Workbook(); wb.remove(wb.active)
    rs=st[st['Bölge Sorumlusu'].map(canon)==canon(region)].copy()
    rt=tt[tt['Bölge Sorumlusu'].map(canon)==canon(region)].copy()
    ws=wb.create_sheet('Bölge Özeti')
    ws.append([product_name().upper(),'']); ws.merge_cells('A1:B1')
    ws.append(['Bölge Sorumlusu',region]); ws.append([]); ws.append(['KPI','Değer'])
    local={'Aktif Mevcut':int(rs['Aktif Mevcut'].sum()),'Yönetim Normu':int(rs['Norm Kadro'].sum()),'Norm Eksiği':int(rs['Norm Eksiği'].sum()),'Norm Fazlası':int(rs['Norm Fazlası'].sum())}
    for k,v in local.items(): ws.append([k,v])
    ws['A1'].fill=PatternFill('solid',fgColor='102F64'); ws['A1'].font=Font(color='FFFFFF',bold=True,size=14)
    ws.column_dimensions['A'].width=30; ws.column_dimensions['B'].width=45
    store_cols=['Bölge Sorumlusu','MağazaID','Mağaza','Aktif Mevcut','Norm Kadro','Norm Eksiği','Norm Fazlası','Net Fark']
    title_source=_personnel_detail_report(rt,staff)
    deficit_source=rt.rename(columns={'Unvan':'Departman'})
    surplus_source=_surplus_people_report(rt,staff)
    title_cols=['Bölge Sorumlusu','MağazaID','Mağaza','Departman','Unvan','Personel Adı Soyadı','Açıklama','Aktif Mevcut','Norm Kadro','Norm Eksiği','Norm Fazlası','Net Fark']
    deficit_cols=['Bölge Sorumlusu','MağazaID','Mağaza','Departman','Aktif Mevcut','Norm Kadro','Norm Eksiği','Norm Fazlası','Net Fark']
    surplus_cols=['Bölge Sorumlusu','MağazaID','Mağaza','Departman','Unvan','Personel Adı Soyadı','Açıklama','Aktif Mevcut','Norm Kadro','Norm Eksiği','Norm Fazlası','Net Fark']
    write_df(wb,'Mağaza Bazlı',rs[store_cols].sort_values('Mağaza'))
    write_df(wb,'Mağaza-Unvan Bazlı',unvan_sirali(title_source[title_cols],['Mağaza'],unvan_kolonu='Departman'))
    write_df(wb,'Norm Eksikleri',deficit_source.loc[deficit_source['Norm Eksiği']>0,deficit_cols].sort_values(['Mağaza','Norm Eksiği'],ascending=[True,False]))
    write_df(wb,'Norm Fazlaları',surplus_source[surplus_cols].sort_values('Mağaza'))
    wb.save(path)



def enhanced_excel_reports(kpi,st,tt,ai,staff):
    tt=balance_detail_table(tt)
    outdir=runtime_root()/'output'/'Bolge_Raporlari'; outdir.mkdir(exist_ok=True)
    temp_dir=runtime_root()/'output'/'Bolge_Excel_Yeni'; shutil.rmtree(temp_dir,ignore_errors=True); temp_dir.mkdir(exist_ok=True)
    for region in sorted(st['Bölge Sorumlusu'].dropna().map(txt).unique()):
        safe=''.join(c if c.isalnum() else '_' for c in region).strip('_')
        _region_excel(temp_dir/f'OMEHR_Bolge_{safe}.xlsx',region,kpi,st,tt,ai,staff)
    for old in outdir.glob('*.xlsx'): old.unlink()
    for fresh in temp_dir.glob('*.xlsx'): fresh.replace(outdir/fresh.name)
    shutil.rmtree(temp_dir,ignore_errors=True)


# ============================================================================
# OUTPUT ORCHESTRATION
# Adds executive dashboard, scenario impact, cost-ready optimization,
# stronger explainability and version/audit outputs while preserving the current BASDAS outputs.
# ============================================================================



def _v16_enrich_explanations(ai):
    if ai is None or ai.empty:
        return ai
    x=ai.copy()
    def explain(r):
        def sn(v):
            try:
                z=float(v)
                return 0.0 if not math.isfinite(z) else z
            except Exception:
                # NOT: Bilerek loglanmıyor — bkz. src/ai_norm.py safe_num
                # (aynı gerekçe: boş/metin hücrelerde sık ve normal başarısızlık).
                return 0.0
        mevcut=int(round(sn(r.get('Aktif Mevcut',0)))); yon=int(round(sn(r.get('Norm Kadro',0)))); ain=int(round(sn(r.get('AI Önerilen Norm',0))))
        transfer=int(round(sn(r.get('Doğrulanmış Transfer',0)))); hire=int(round(sn(r.get('Doğrulanmış İşe Alım',0))))
        confidence=sn(r.get('Güven Skoru',0))
        direction='eksik' if ain>mevcut else ('fazla' if ain<mevcut else 'dengede')
        delta=abs(ain-mevcut)
        official='Yönetim normuna göre '+(('eksik' if yon>mevcut else 'fazla') if yon!=mevcut else 'dengede')
        action=[]
        if transfer: action.append(f'{transfer} kişi transfer')
        if hire: action.append(f'{hire} kişi işe alım')
        if ain<mevcut: action.append(f'{delta} kişi transfer havuzunda değerlendirme')
        if not action: action.append('mevcut yapının korunması')
        operational=[]
        fte=sn(r.get('İş Yükü FTE',0)); peak=sn(r.get('Pik Katsayısı',0)); minimum=int(round(sn(r.get('Minimum Kadro',0))))
        if fte>0: operational.append(f'hesaplanan iş yükü {fte:.2f} FTE')
        if peak>1.001: operational.append(f'pik dönem katsayısı {peak:.2f}')
        if minimum>0: operational.append(f'operasyonel minimum {minimum}')
        if not operational: operational.append('mevcut veri setindeki norm ve operasyon göstergeleri')
        return (f"{txt(r.get('Unvan'))} için mevcut {mevcut}, yönetim normu {yon}, AI normu {ain}. "
                f"AI değerlendirmesinde {delta} kişi {direction}; {official}. "
                f"Karar dayanakları: {', '.join(operational)}. Önerilen uygulama: {', '.join(action)}. "
                f"Güven skoru %{confidence:.0f}. Yönetim normu resmi kadro planıdır; AI sonucu bağımsız karar desteğidir.")
    x['OMEHR Yönetici Açıklaması']=x.apply(explain,axis=1)
    x['AI Durum Metni']=x.apply(lambda r:_gap_text(r.get('Aktif Mevcut',0),r.get('AI Önerilen Norm',0)),axis=1)
    return x



def _v16_scenario_impact(kpi,scens):
    rows=[]
    deficit=int(kpi.get('Norm Eksiği',0)); excess=int(kpi.get('Norm Fazlası',0)); net=int(kpi.get('Net İhtiyaç',0))
    for name,df in scens.items():
        transfers=0 if df is None or df.empty else int(len(df))
        remaining=max(deficit-transfers,0)
        rows.append({'Senaryo':name,'Başlangıç Norm Eksiği':deficit,'Transfer Önerisi':transfers,'Transfer Sonrası Kalan Açık':remaining,'Norm Fazlası Havuzu':excess,'Net İşe Alım İhtiyacı':max(net-transfers,0),'Karar Notu':('En çok açığı kapatan senaryo' if transfers==max([0]+[len(d) for d in scens.values() if d is not None]) else 'Alternatif optimizasyon senaryosu')})
    return pd.DataFrame(rows)



def _v16_add_workbook_layers(wb,kpi,st,tt,ai,scens,risk):
    from openpyxl.chart import BarChart,Reference
    for s in ['Yönetim Paneli','Senaryo Etkisi','Maliyet Hazırlığı','Kritik Aksiyonlar','Kritik Norm Durumları']:
        if s in wb.sheetnames: del wb[s]
    ws=wb.create_sheet('Yönetim Paneli',0)
    ws.sheet_view.showGridLines=False
    ws.merge_cells('A1:H2');ws['A1']=product_name().upper()+' - YÖNETİM PANELİ';ws['A1'].fill=PatternFill('solid',fgColor='102F64');ws['A1'].font=Font(color='FFFFFF',bold=True,size=18);ws['A1'].alignment=Alignment(horizontal='center',vertical='center')
    labels=['Aktif Mevcut','Yönetim Normu','Norm Eksiği','Norm Fazlası','Net İhtiyaç']
    vals=[kpi.get('Aktif Mevcut',0),kpi.get('Toplam Norm',0),kpi.get('Norm Eksiği',0),kpi.get('Norm Fazlası',0),kpi.get('Net İhtiyaç',0)]
    for i,(lab,val) in enumerate(zip(labels,vals),start=1):
        c=1+(i-1)*2;ws.merge_cells(start_row=4,start_column=c,end_row=4,end_column=c+1);ws.merge_cells(start_row=5,start_column=c,end_row=6,end_column=c+1)
        ws.cell(4,c,lab);ws.cell(5,c,val)
        ws.cell(4,c).fill=PatternFill('solid',fgColor='4472C4');ws.cell(4,c).font=Font(color='FFFFFF',bold=True);ws.cell(4,c).alignment=Alignment(horizontal='center')
        ws.cell(5,c).font=Font(bold=True,size=20);ws.cell(5,c).alignment=Alignment(horizontal='center',vertical='center')
    ws['A8']='En Kritik 10 Mağaza';ws['A8'].font=Font(bold=True,size=13,color='102F64')
    crit=st.sort_values(['Norm Eksiği','Risk Puanı'] if 'Risk Puanı' in st.columns else ['Norm Eksiği'],ascending=False).head(10)
    headers=['Mağaza','Bölge Sorumlusu','Mevcut','Norm','Eksik','Fazla']
    for j,h in enumerate(headers,1):ws.cell(9,j,h)
    for ri,(_,r) in enumerate(crit.iterrows(),10):
        vals=[r.get('Mağaza',''),r.get('Bölge Sorumlusu',''),r.get('Aktif Mevcut',0),r.get('Norm Kadro',0),r.get('Norm Eksiği',0),r.get('Norm Fazlası',0)]
        for j,v in enumerate(vals,1):ws.cell(ri,j,v)
    for cell in ws[9]:cell.fill=PatternFill('solid',fgColor='102F64');cell.font=Font(color='FFFFFF',bold=True)
    for column,width in {'A':28,'B':28,'C':12,'D':12,'E':12,'F':12,'G':16,'H':16,'I':16,'J':16}.items():ws.column_dimensions[column].width=width
    chart=BarChart();chart.title='En Yüksek Norm Açıkları';chart.y_axis.title='Kişi';chart.x_axis.title='Mağaza'
    chart.add_data(Reference(ws,min_col=5,min_row=9,max_row=9+len(crit)),titles_from_data=True);chart.set_categories(Reference(ws,min_col=1,min_row=10,max_row=9+len(crit)));chart.height=7;chart.width=14;ws.add_chart(chart,'H9')
    impact=_v16_scenario_impact(kpi,scens);write_df(wb,'Senaryo Etkisi',impact)
    # Cost-ready layer: monetary figures remain zero unless the input provides cost data.
    cost=pd.DataFrame([{'Kalem':'Net işe alım ihtiyacı','Kişi':int(kpi.get('Net İhtiyaç',0)),'Birim Aylık Maliyet':0,'Tahmini Aylık Maliyet':0,'Not':'Birim maliyet inputa eklendiğinde otomatik hesaplanmaya hazır.'},{'Kalem':'Transferle kapatılabilecek açık','Kişi':max([0]+[len(d) for d in scens.values() if d is not None]),'Birim Aylık Maliyet':0,'Tahmini Aylık Maliyet':0,'Not':'Transfer maliyeti ve tasarruf parametreleri girilebilir.'}])
    write_df(wb,'Maliyet Hazırlığı',cost)
    # Kritik liste yalnızca resmi yönetim normu ile mevcut arasındaki farkı gösterir.
    critical=tt[(pd.to_numeric(tt.get('Norm Eksiği',0),errors='coerce').fillna(0)>0) | (pd.to_numeric(tt.get('Norm Fazlası',0),errors='coerce').fillna(0)>0)].copy()
    if not critical.empty:
        cols=[c for c in ['Bölge Sorumlusu','Mağaza','Unvan','Aktif Mevcut','Norm Kadro','Norm Eksiği','Norm Fazlası'] if c in critical.columns]
        critical=critical.sort_values(['Norm Eksiği','Norm Fazlası'],ascending=False)[cols]
    write_df(wb,'Kritik Norm Durumları',critical if not critical.empty else pd.DataFrame([{'Durum':'Kritik norm farkı bulunmadı'}]))
    ai_critical=ai.copy() if ai is not None else pd.DataFrame()
    if not ai_critical.empty:
        ai_gap=pd.to_numeric(ai_critical.get('AI-Mevcut Fark',0),errors='coerce').fillna(0)
        priority=ai_critical.get('Öncelik Seviyesi',pd.Series('',index=ai_critical.index)).astype(str).str.casefold()
        ai_critical=ai_critical[(ai_gap.ne(0)) | priority.isin({'kritik','yüksek','yuksek'})].copy()
        critical_columns=[c for c in [
            'Bölge Sorumlusu','Mağaza','Unvan','Aktif Mevcut','Norm Kadro',
            'AI Önerilen Norm','AI-Mevcut Fark','Öncelik Seviyesi',
            'Önerilen Aksiyon','Aksiyon Gerekçesi','OMEHR Yönetici Açıklaması'
        ] if c in ai_critical.columns]
        ai_critical=ai_critical[critical_columns]
    if ai is not None and not ai.empty:
        write_df(wb,'Kritik Aksiyonlar',ai_critical if not ai_critical.empty else pd.DataFrame([{'Durum':'Kritik AI aksiyonu bulunmadı'}]))


import functools


@functools.lru_cache(maxsize=4)


def _executive_analysis_frames(input_path):
    """Son mevcut dönemi kullanarak mükerrer ay toplamı oluşturmadan finansal/operasyonel analiz üretir.

    PERFORMANS DÜZELTMESİ (P2 — engine_core.py incelemesi): Bu fonksiyon TEK
    bir run_all() çalıştırmasında 5 FARKLI yerden çağrılıyordu ve her
    seferinde AYNI Excel dosyasını (tüm sayfaları) baştan okuyordu —
    profilleme ile ölçülen gerçek maliyet: ~40 saniye (toplam ~184 saniyelik
    çalıştırmanın %22'si). @lru_cache ile process-içi önbelleğe alınarak
    bu maliyet TEK okumaya indirildi (aynı input_path için sonraki çağrılar
    anında döner). NOT: dönüş değeri (DataFrame'ler) çağıranlar tarafından
    DEĞİŞTİRİLMEMELİDİR — önbellek paylaşılan referans döndürür.
    """
    def _read(path,sheet):
        try:
            raw=pd.read_excel(path,sheet_name=sheet,header=None)
        except ValueError:
            # DÜZELTME: bu sayfa (ör. 'Aylık Operasyon KPI', 'Fazla Mesai')
            # İSTEĞE BAĞLIDIR — henüz doldurulmamış bir kurulumda hiç
            # bulunmayabilir. Önceden bu durumda ValueError yükselip TÜM
            # yönetici analizi (ve onu çağıran run_all() adımı) çöküyordu;
            # artık aşağıdaki latest()/canonical_locations() zaten boş bir
            # DataFrame'i güvenle işleyebildiği için buradan boş döndürmek
            # yeterli ve doğru.
            return pd.DataFrame()
        best=0;best_score=-1
        terms={'MAGAZAID','MAGAZA','AY','DONEM','TARIH','AYLIKCIRO','AYLIKFIS',
               'PERSONELMALIYETI','MESAIMALIYETI','TOPLAMMALIYET','ISYUKUENDEKSI'}
        for row_index in range(min(14,len(raw))):
            score=sum(1 for value in raw.iloc[row_index].dropna() if canon(value).replace(' ','').upper() in terms)
            if score>best_score:best,best_score=row_index,score
        return pd.read_excel(path,sheet_name=sheet,header=best)

    def latest(frame):
        if frame.empty:return frame
        period=next((c for c in frame.columns if canon(c) in {'ay','donem','tarih'}),None)
        if not period:return frame
        values=frame[period].dropna().astype(str)
        return frame[frame[period].astype(str).eq(values.max())].copy() if not values.empty else frame

    def number(frame,name):
        c=col(frame,name)
        return numeric(frame[c]) if c else pd.Series(0,index=frame.index,dtype=float)

    master=pd.read_excel(input_path,sheet_name='Dim_Magaza')
    master_id=col(master,'MağazaID','MagazaID')
    master_store=col(master,'Mağaza','Magaza')
    store_by_id={}
    primary_id_by_store={}
    if master_id and master_store:
        for _,row in master[[master_id,master_store]].dropna().drop_duplicates().iterrows():
            store_id=txt(row[master_id])
            store_name=txt(row[master_store])
            store_by_id[store_id]=store_name
            key=canon(store_name)
            if key not in primary_id_by_store or store_id<primary_id_by_store[key]:
                primary_id_by_store[key]=store_id

    def canonical_locations(frame):
        """MağazaID ana verisini esas alır; aynı fiziksel birimin ikinci kodunu dışlar."""
        if frame.empty:return frame
        mid=col(frame,'MagazaID','MağazaID')
        store=col(frame,'Mağaza','Magaza')
        if not mid or not store:return frame
        result=frame.copy()
        result[store]=result[mid].astype(str).map(store_by_id).fillna(result[store])
        keep=result.apply(
            lambda row: primary_id_by_store.get(canon(row[store]),txt(row[mid]))==txt(row[mid]),
            axis=1,
        )
        return result[keep].copy()

    def unit_type(value):
        name=canon(value).upper()
        if 'MERKEZ' in name:return 'Merkez'
        if 'DEPO' in name or 'CIFTLIK' in name:return 'Depo'
        return 'Mağaza'

    ops=canonical_locations(latest(_read(input_path,'Aylık Operasyon KPI')))
    costs=canonical_locations(latest(_read(input_path,'Personel Maliyeti')))
    overtime=latest(_read(input_path,'Fazla Mesai'))
    fire=_read(input_path,'Fire ve İade')
    workload=_read(input_path,'İş Yükü Endeksi')

    def keys(frame):
        mid=col(frame,'MagazaID','MağazaID'); store=col(frame,'Mağaza','Magaza')
        return mid,store

    financial=pd.DataFrame()
    cmid,cstore=keys(costs)
    if cmid and cstore:
        financial=costs[[cmid,cstore]].copy()
        financial.columns=['MağazaID','Mağaza']
        financial['Personel Maliyeti']=number(costs,'Personel Maliyeti')
        financial['Mesai Maliyeti']=number(costs,'Mesai Maliyeti')
        financial['Prim/Kesinti']=number(costs,'Prim/Kesinti')
        financial['Toplam İş Gücü Maliyeti']=financial[['Personel Maliyeti','Mesai Maliyeti','Prim/Kesinti']].sum(axis=1)
    omid,ostore=keys(ops)
    if omid and ostore:
        revenue=ops[[omid,ostore]].copy();revenue.columns=['MağazaID','Mağaza']
        revenue['Aylık Ciro']=number(ops,'Aylık Ciro')
        financial=revenue.merge(financial,on=['MağazaID','Mağaza'],how='outer') if not financial.empty else revenue
    if not financial.empty:
        financial.insert(2,'Birim Tipi',financial['Mağaza'].map(unit_type))
        for c in ['Aylık Ciro','Personel Maliyeti','Mesai Maliyeti','Prim/Kesinti','Toplam İş Gücü Maliyeti']:
            if c not in financial:financial[c]=0
            financial[c]=pd.to_numeric(financial[c],errors='coerce').fillna(0)
        financial['İş Gücü Maliyeti / Ciro %']=np.where(
            financial['Aylık Ciro']>0,
            financial['Toplam İş Gücü Maliyeti']/financial['Aylık Ciro']*100,
            np.nan,
        ).round(2)

    operational=pd.DataFrame()
    if omid and ostore:
        operational=ops[[omid,ostore]].copy();operational.columns=['MağazaID','Mağaza']
        for source,target in [
            ('Aylık Fiş','Aylık Fiş'),('Aylık Ciro','Aylık Ciro'),
            ('Ort. Sepet','Ortalama Sepet'),('Online Sipariş','Online Sipariş'),
            ('Mal Kabul','Mal Kabul')
        ]:operational[target]=number(ops,source)
    fmid,fstore=keys(fire)
    if fmid and fstore:
        f=fire[[fmid,fstore]].copy();f.columns=['MağazaID','Mağaza'];f['Fire Oranı %']=number(fire,'Fire Oranı %')
        operational=f if operational.empty else operational.merge(f,on=['MağazaID','Mağaza'],how='outer')
    wmid,wstore=keys(workload)
    if wmid and wstore:
        w=workload[[wmid,wstore]].copy();w.columns=['MağazaID','Mağaza'];w['İş Yükü Endeksi']=number(workload,'İş Yükü Endeksi')
        operational=w if operational.empty else operational.merge(w,on=['MağazaID','Mağaza'],how='outer')
    if not operational.empty:
        operational.insert(2,'Birim Tipi',operational['Mağaza'].map(unit_type))
        numeric_cols=operational.select_dtypes(include='number').columns
        operational[numeric_cols]=operational[numeric_cols].fillna(0)

    total_revenue=float(financial.get('Aylık Ciro',pd.Series(dtype=float)).sum())
    total_cost=float(financial.get('Toplam İş Gücü Maliyeti',pd.Series(dtype=float)).sum())
    total_receipts=float(operational.get('Aylık Fiş',pd.Series(dtype=float)).sum())
    overtime_hours=float(number(overtime,'Fazla Mesai Saat').sum()) if not overtime.empty else 0
    online_orders=float(operational.get('Online Sipariş',pd.Series(dtype=float)).sum())
    summary=pd.DataFrame([
        {'Gösterge':'Son Dönem Toplam Ciro','Değer':round(total_revenue),'Birim':'TL','Açıklama':'Aylık Operasyon KPI son mevcut dönem toplamı'},
        {'Gösterge':'Toplam İş Gücü Maliyeti','Değer':round(total_cost),'Birim':'TL','Açıklama':'Personel, mesai ve prim/kesinti toplamı'},
        {'Gösterge':'İş Gücü Maliyeti / Ciro','Değer':round(total_cost/total_revenue*100,2) if total_revenue else 0,'Birim':'%','Açıklama':'Toplam iş gücü maliyetinin ciroya oranı'},
        {'Gösterge':'Aylık Toplam Fiş','Değer':round(total_receipts),'Birim':'Adet','Açıklama':'Son mevcut dönem toplam fiş'},
        {'Gösterge':'Fazla Mesai','Değer':round(overtime_hours),'Birim':'Saat','Açıklama':'Son mevcut dönem fazla mesai saati'},
        {'Gösterge':'Online Sipariş','Değer':round(online_orders),'Birim':'Adet','Açıklama':'Son mevcut dönem online sipariş'},
    ])
    return summary,financial,operational




def _add_executive_analysis_sheets(wb,input_path):
    grouped_sheets=['Maliyet - Mağaza','Maliyet - Depo','Maliyet - Merkez']
    for name in ['Yönetici Analiz Özeti','Yönetici Finansal Analiz','Yönetici Operasyonel Analiz',*grouped_sheets]:
        if name in wb.sheetnames:del wb[name]
    summary,financial,operational=_executive_analysis_frames(input_path)
    write_df(wb,'Yönetici Analiz Özeti',summary)
    write_df(wb,'Yönetici Finansal Analiz',financial)
    write_df(wb,'Yönetici Operasyonel Analiz',operational)
    for unit,sheet in [('Mağaza','Maliyet - Mağaza'),('Depo','Maliyet - Depo'),('Merkez','Maliyet - Merkez')]:
        subset=financial[financial.get('Birim Tipi',pd.Series('',index=financial.index)).eq(unit)].copy()
        write_df(wb,sheet,subset if not subset.empty else pd.DataFrame([{'Durum':f'{unit} verisi bulunamadı'}]))




def _add_visible_ai_dashboard(wb,kpi,ai,input_path):
    from openpyxl.chart import BarChart,Reference
    if '00 AI Analiz Paneli' in wb.sheetnames:del wb['00 AI Analiz Paneli']
    ws=wb.create_sheet('00 AI Analiz Paneli',0)
    ws.merge_cells('A1:J1');ws['A1']=product_name()+' - AI, OPERASYON VE MALİYET YÖNETİCİ PANELİ'
    ws['A1'].fill=PatternFill('solid',fgColor='102F64');ws['A1'].font=Font(color='FFFFFF',bold=True,size=15);ws['A1'].alignment=Alignment(horizontal='center')
    summary,financial,operational=_executive_analysis_frames(input_path)
    ai_gap=pd.to_numeric(ai.get('AI-Mevcut Fark',0),errors='coerce').fillna(0)
    ai_metrics=[
        ('Aktif Mevcut',int(kpi['Aktif Mevcut'])),('Yönetim Normu',int(kpi['Toplam Norm'])),
        ('AI Önerilen Norm',int(pd.to_numeric(ai.get('AI Önerilen Norm',0),errors='coerce').fillna(0).sum())),
        ('AI Kapasite Açığı',int(ai_gap.clip(lower=0).sum())),('AI Transfer Adayı',int((-ai_gap.clip(upper=0)).sum())),
    ]
    for index,(label,value) in enumerate(ai_metrics,1):
        column=(index-1)*2+1
        ws.cell(3,column,label);ws.cell(4,column,value)
        ws.merge_cells(start_row=3,start_column=column,end_row=3,end_column=column+1)
        ws.merge_cells(start_row=4,start_column=column,end_row=4,end_column=column+1)
        ws.cell(3,column).fill=PatternFill('solid',fgColor='4472C4');ws.cell(3,column).font=Font(color='FFFFFF',bold=True)
        ws.cell(4,column).font=Font(bold=True,size=18);ws.cell(3,column).alignment=ws.cell(4,column).alignment=Alignment(horizontal='center')
    ws['A6']='OPERASYON VE MALİYET ÖZETİ';ws['A6'].font=Font(bold=True,size=12,color='102F64')
    headers=['Gösterge','Değer','Birim','Yönetici Açıklaması']
    for col_index,header in enumerate(headers,1):ws.cell(7,col_index,header)
    for row_index,(_,row) in enumerate(summary.iterrows(),8):
        values=[row['Gösterge'],row['Değer'],row['Birim'],row['Açıklama']]
        for col_index,value in enumerate(values,1):ws.cell(row_index,col_index,value)
    for cell in ws[7]:cell.fill=PatternFill('solid',fgColor='102F64');cell.font=Font(color='FFFFFF',bold=True)
    ws['A16']='ÖNCELİKLİ AI NORM AKSİYONLARI';ws['A16'].font=Font(bold=True,size=12,color='102F64')
    action_headers=['Mağaza','Unvan','Mevcut','Yönetim Normu','AI Önerilen Norm','AI Fark','Güven','Önerilen Aksiyon','Akıllı Açıklama']
    for col_index,header in enumerate(action_headers,1):ws.cell(17,col_index,header)
    top=ai.assign(_Gap=ai_gap).sort_values(['_Gap','Güven Skoru'],ascending=False)
    top=top[top['_Gap']>0].head(20)
    for row_index,(_,row) in enumerate(top.iterrows(),18):
        values=[row.get('Mağaza'),row.get('Unvan'),row.get('Aktif Mevcut'),row.get('Norm Kadro'),
                row.get('AI Önerilen Norm'),row.get('_Gap'),row.get('Güven Skoru'),
                row.get('Önerilen Aksiyon'),row.get('Aksiyon Gerekçesi')]
        for col_index,value in enumerate(values,1):ws.cell(row_index,col_index,value)
    for cell in ws[17]:cell.fill=PatternFill('solid',fgColor='4472C4');cell.font=Font(color='FFFFFF',bold=True)
    ws.freeze_panes='A17';ws.auto_filter.ref=f'A17:I{17+len(top)}'
    widths={'A':24,'B':25,'C':11,'D':16,'E':18,'F':10,'G':10,'H':42,'I':85,'J':4}
    for column,width in widths.items():ws.column_dimensions[column].width=width
    for row in ws.iter_rows(min_row=1,max_row=17+len(top),min_col=1,max_col=9):
        for cell in row:cell.alignment=Alignment(vertical='top',wrap_text=True)
    if len(top):
        chart=BarChart();chart.type='bar';chart.style=10;chart.title='En Yüksek AI Kapasite Açıkları';chart.x_axis.title='Kişi';chart.y_axis.title='Mağaza / Unvan'
        chart.add_data(Reference(ws,min_col=6,min_row=17,max_row=17+len(top)),titles_from_data=True)
        chart.set_categories(Reference(ws,min_col=1,min_row=18,max_row=17+len(top)))
        chart.height=8;chart.width=15;ws.add_chart(chart,'K17')




def _build_admin_report_pack(kpi, st, tt, ai, input_path, main_pdf, main_excel):
    """Üst yönetim için ham veri yerine kısa, karar odaklı rapor dosyaları üretir."""
    from openpyxl import Workbook
    outdir=runtime_root()/'output'
    summary,financial,operational=_executive_analysis_frames(input_path)
    critical=tt[(numeric(tt.get('Norm Eksiği',0))>0)|(numeric(tt.get('Norm Fazlası',0))>0)].copy()
    critical_cols=[c for c in ['Bölge Sorumlusu','Mağaza','Unvan','Aktif Mevcut','Norm Kadro','Norm Eksiği','Norm Fazlası'] if c in critical]
    critical=critical[critical_cols].sort_values(['Norm Eksiği','Norm Fazlası'],ascending=False)
    ai_gap=numeric(ai.get('AI-Mevcut Fark',0))
    actions=ai[ai_gap.ne(0)].copy()
    action_cols=[c for c in ['Bölge Sorumlusu','Mağaza','Unvan','Aktif Mevcut','Norm Kadro','AI Önerilen Norm','AI-Mevcut Fark','Güven Skoru','Önerilen Aksiyon','Aksiyon Gerekçesi'] if c in actions]
    actions=actions[action_cols].sort_values('AI-Mevcut Fark',ascending=False)

    def make_book(path, sheets):
        wb=Workbook()
        wb.remove(wb.active)
        for name,frame in sheets:
            write_df(wb,name,frame)
        wb.save(path)
        return path

    norm_path=make_book(outdir/'OMEHR_Admin_Norm_ve_Aksiyonlar.xlsx',[
        ('KPI Özeti',pd.DataFrame([kpi])),
        ('Kritik Norm Durumları',critical),
        ('AI Aksiyonları',actions),
    ])
    finance_path=make_book(outdir/'OMEHR_Admin_Maliyet_ve_Operasyon.xlsx',[
        ('Yönetici Özeti',summary),
        ('Maliyet Analizi',financial),
        ('Operasyon Analizi',operational),
        ('Maliyet - Mağaza',financial[financial.get('Birim Tipi',pd.Series('',index=financial.index)).eq('Mağaza')]),
        ('Maliyet - Depo',financial[financial.get('Birim Tipi',pd.Series('',index=financial.index)).eq('Depo')]),
        ('Maliyet - Merkez',financial[financial.get('Birim Tipi',pd.Series('',index=financial.index)).eq('Merkez')]),
    ])
    from pypdf import PdfReader,PdfWriter
    admin_pdf=outdir/'OMEHR_Admin_Yonetici_Ozeti.pdf'
    reader=PdfReader(str(main_pdf));writer=PdfWriter()
    for page in reader.pages[:min(3,len(reader.pages))]:writer.add_page(page)
    with admin_pdf.open('wb') as stream:writer.write(stream)

    def extract_pdf(filename,page_indexes):
        target=outdir/filename
        subset=PdfWriter()
        for index in page_indexes:
            if index<len(reader.pages):subset.add_page(reader.pages[index])
        with target.open('wb') as stream:subset.write(stream)
        return target

    ai_pdf=extract_pdf('OMEHR_AI_Karar_Analizi.pdf',[0,1])
    operation_pdf=extract_pdf('OMEHR_Operasyon_Verimlilik_Analizi.pdf',[0,1])
    cost_pdf=extract_pdf('OMEHR_Maliyet_Analizi.pdf',[1,2])
    # Tüm yönetici çıktıları doğrudan kurumsal BASDAS adlarıyla üretilir.
    return [admin_pdf,ai_pdf,operation_pdf,cost_pdf,norm_path,finance_path,Path(main_excel)]


def build_boxed_manager_excel(st, norm, staff, kpi=None, output_path=None):
    """Kutucuklu yönetici PDF'sinin personel satırlı Excel karşılığını üretir.

    Her yönetici ayrı sayfadadır. Her mağaza ayrı kutudur ve PDF ile aynı
    ayrıntı mantığını taşır: Gerçek Unvan, Ad Soyad, M, N, E, F.
    Eksik satırlar mavi, norm fazlası personel satırları yeşil gösterilir.
    """
    out = Path(output_path) if output_path else runtime_root() / 'output' / 'OMEHR_Kutucuklu_Yonetici_Raporu.xlsx'
    out.parent.mkdir(parents=True, exist_ok=True)

    def _norm_name(value):
        return ' '.join(canon(txt(value)).split())

    # Gerekli kolonları güvenli biçimde hazırla.
    stx = st.copy()
    for c in ['Bölge Sorumlusu','Mağaza','Aktif Mevcut','Norm Kadro','Norm Eksiği','Norm Fazlası','Net Fark']:
        if c not in stx.columns:
            stx[c] = '' if c in {'Bölge Sorumlusu','Mağaza'} else 0
    stx['_yonetici'] = stx['Bölge Sorumlusu'].map(_norm_name)

    # DÜZELTME (KRİTİK, CANLI ÜRETİM HATASI — çok kiracılı SaaS):
    # önceden burada 4 SABİT, orijinal firmanın gerçek bölge sorumlusu
    # ismi vardı ("Ali Çelik", "Derya Yardımcı", "Cüneyt & Ayşe Avcu",
    # "Ertan Teki") — HER sekme, o mağazaların "_yonetici" alanı bu
    # sabit isimlerden BİRİYLE eşleşirse dolardı. BAŞKA HER kiracı için
    # (kendi bölge sorumluları bambaşka isimler taşıdığı için) bu 4
    # sekmenin TAMAMI SESSİZCE BOŞ kalırdı — rapor "başarıyla" üretilir
    # ama hiçbir mağaza verisi göstermezdi. Artık gruplar, TENANT'IN
    # KENDİ verisinde GERÇEKTEN var olan "Bölge Sorumlusu" değerlerinden
    # dinamik olarak türetilir — canon() ile aynı normalizasyon (Türkçe
    # karakter/boşluk farklarını zaten tekilleştiriyor), sabit bir isim
    # listesine ihtiyaç duymadan.
    groups = []
    for _norm_key in dict.fromkeys(v for v in stx['_yonetici'] if v):
        _orijinal_adlar = stx.loc[stx['_yonetici'].eq(_norm_key), 'Bölge Sorumlusu']
        _goruntu_adi = txt(_orijinal_adlar.iloc[0]).strip() or _norm_key.title()
        groups.append((_goruntu_adi, {_norm_key}))

    nm = req(norm,'Mağaza','Magaza'); nu = req(norm,'Unvan'); nn = req(norm,'Norm Kadro','Norm')
    norm_note_col = col(norm,'Açıklama','Aciklama','AÇIKLAMA','ACIKLAMA')
    sm = req(staff,'Mağaza','Magaza'); su = req(staff,'Unvan'); dep = req(staff,'Departman')
    pname = req(staff,'İsim Soyisim','Isim Soyisim','Ad Soyad')
    staff_note_col = col(staff,'Açıklama','Aciklama','AÇIKLAMA','ACIKLAMA','Personel Açıklaması','Personel Aciklamasi')
    entry_col = col(staff,'İşe Giriş','Ise Giris')

    wb = Workbook(); wb.remove(wb.active)
    navy='102F64'; blue='4472C4'; shortage_blue='D9EAF7'; surplus_green='E2F0D9'
    pale='F3F6FA'; white='FFFFFF'; gold='BF9000'; dark='1F2937'; gray='D9E1F2'
    thin=Side(style='thin',color='7F8C9A'); medium=Side(style='medium',color=navy)

    def _store_payload(store_name, region_name):
        ns = norm[norm[nm].map(canon) == canon(store_name)].copy()
        ps = staff[staff[sm].map(canon) == canon(store_name)].copy()

        msum = (ps.assign(_Key=ps[dep].map(_title_key))
                  .groupby('_Key',dropna=False)[pname].count().reset_index(name='Mevcut'))
        nsum = (ns.assign(_Key=ns[nu].map(_title_key))
                  .groupby('_Key',dropna=False)[nn].sum().reset_index(name='Norm'))
        role_names = (ns.assign(_Key=ns[nu].map(_title_key))
                        .drop_duplicates('_Key').set_index('_Key')[nu].to_dict()) if not ns.empty else {}
        notes_by_key = {}
        if norm_note_col and norm_note_col in ns.columns:
            for _key, _grp in ns.assign(_Key=ns[nu].map(_title_key)).groupby('_Key'):
                _metin = ' '.join(txt(v).strip() for v in _grp[norm_note_col] if txt(v).strip())
                if _metin.strip():
                    notes_by_key[_key] = _metin.strip()
        td = nsum.merge(msum,on='_Key',how='outer').fillna({'Norm':0,'Mevcut':0})
        td['Unvan'] = td['_Key'].map(role_names).fillna(td['_Key'].map(lambda z: txt(z).upper()))
        td['Norm'] = numeric(td['Norm']).astype(int); td['Mevcut'] = numeric(td['Mevcut']).astype(int)
        td['Eksik']=(td['Norm']-td['Mevcut']).clip(lower=0).astype(int)
        td['Fazla']=(td['Mevcut']-td['Norm']).clip(lower=0).astype(int)
        td=balance_store_title_rows(
            td,key_col='_Key',norm_col='Norm',current_col='Mevcut',
            deficit_col='Eksik',surplus_col='Fazla'
        )
        td = unvan_sirali(td,unvan_kolonu='Unvan')

        rows=[]
        for _, r in td.iterrows():
            persons = ps[ps[dep].map(_title_key)==r['_Key']].copy()
            if entry_col and entry_col in persons.columns:
                persons['_Tarih']=pd.to_datetime(persons[entry_col],errors='coerce')
                persons=persons.sort_values('_Tarih',ascending=False,na_position='last')
            excess_count=min(int(r['Fazla']),len(persons))
            excess_idx=set(persons.head(excess_count).index)
            regular=persons.loc[~persons.index.isin(excess_idx)]
            excess=persons.loc[persons.index.isin(excess_idx)]

            if regular.empty and excess.empty:
                rows.append({
                    'Unvan':txt(r['Unvan']), 'Ad Soyad':'BOŞ POZİSYON' if int(r['Eksik']) else '-',
                    'M':0, 'N':int(r['Norm']), 'E':int(r['Eksik']) or '-', 'F':'-',
                    'Durum':'eksik' if int(r['Eksik']) else 'normal', 'Açıklama':'', 'Açıklama Cümlesi':'', 'Açıklama Türü':'',
                    'Norm Açıklaması': notes_by_key.get(r['_Key'], ''),
                })
            else:
                for j,(_,pr) in enumerate(regular.iterrows()):
                    rows.append({
                        'Unvan':txt(pr[su]), 'Ad Soyad':txt(pr[pname]), 'M':1,
                        'N':int(r['Norm']) if j==0 else '',
                        'E':int(r['Eksik']) if j==0 and int(r['Eksik']) else '-', 'F':'-',
                        'Durum':'eksik' if j==0 and int(r['Eksik']) else 'normal', 'Açıklama':txt(pr.get(staff_note_col,'')).strip() if staff_note_col else '', 'Açıklama Cümlesi':format_person_note(pr.get(pname,''),txt(pr.get(staff_note_col,'')).strip() if staff_note_col else ''), 'Açıklama Türü':note_kind(txt(pr.get(staff_note_col,'')).strip() if staff_note_col else ''),
                        'Norm Açıklaması': notes_by_key.get(r['_Key'], ''),
                    })
                for _,pr in excess.iterrows():
                    rows.append({
                        'Unvan':txt(pr[su]), 'Ad Soyad':txt(pr[pname]), 'M':1,
                        'N':'', 'E':'-', 'F':1, 'Durum':'fazla', 'Açıklama':txt(pr.get(staff_note_col,'')).strip() if staff_note_col else '', 'Açıklama Cümlesi':format_person_note(pr.get(pname,''),txt(pr.get(staff_note_col,'')).strip() if staff_note_col else ''), 'Açıklama Türü':note_kind(txt(pr.get(staff_note_col,'')).strip() if staff_note_col else ''),
                        'Norm Açıklaması': notes_by_key.get(r['_Key'], ''),
                    })

        manager_mask = ps[dep].map(canon).isin({'yonetici','magaza yoneticisi','sube muduru'})
        if su in ps.columns:
            rt=ps[su].map(canon)
            manager_mask = manager_mask | (rt.str.contains('yonetici',na=False) & ~rt.str.contains('yardimci',na=False))
        managers=ps.loc[manager_mask,pname].dropna().map(txt).drop_duplicates().tolist()
        manager_text=', '.join(managers[:2]) if managers else 'Yönetici bilgisi girilmemiş'
        return {
            'store':store_name, 'region':region_name, 'manager':manager_text, 'rows':rows,
            'current':int(len(ps)), 'norm':int(td['Norm'].sum()),
            'deficit':int(td['Eksik'].sum()), 'surplus':int(td['Fazla'].sum()),
            'net':int(len(ps)-td['Norm'].sum()), 'notes':{txt(r.get('Ad Soyad')):txt(r.get('Açıklama Cümlesi')) for r in rows if txt(r.get('Açıklama Cümlesi')).strip()}
        }

    def _write_card(ws, top, left, payload):
        # Kart genişliği 6 sütun: Unvan, Ad Soyad, M, N, E, F
        widths=[25,28,6,6,6,6]
        for offset,w in enumerate(widths):
            col_letter=get_column_letter(left+offset)
            ws.column_dimensions[col_letter].width=max(ws.column_dimensions[col_letter].width or 0,w)

        end=left+5
        ws.merge_cells(start_row=top,start_column=left,end_row=top,end_column=end)
        c=ws.cell(top,left); c.value=payload['store'].upper(); c.fill=PatternFill('solid',fgColor=navy)
        c.font=Font(name='Arial',size=11,bold=True,color=white); c.alignment=Alignment(horizontal='center',vertical='center')
        if payload.get('notes'):
            combined = "\n".join(f"{txt(k).upper()}: {txt(v)}" for k,v in payload['notes'].items() if txt(v).strip())
            if combined:
                c.comment=Comment(combined,'OMEHR Personel Açıklamaları')
        ws.row_dimensions[top].height=23

        ws.merge_cells(start_row=top+1,start_column=left,end_row=top+1,end_column=end)
        c=ws.cell(top+1,left); c.value=f"Mağaza Yöneticisi: {payload['manager']}"
        c.fill=PatternFill('solid',fgColor=pale); c.font=Font(name='Arial',size=8,color=dark)
        c.alignment=Alignment(horizontal='left',vertical='center',wrap_text=True)
        ws.merge_cells(start_row=top+2,start_column=left,end_row=top+2,end_column=end)
        c=ws.cell(top+2,left); c.value=f"Bölge Sorumlusu: {payload['region']}"
        c.fill=PatternFill('solid',fgColor=pale); c.font=Font(name='Arial',size=8,color=dark)
        c.alignment=Alignment(horizontal='left',vertical='center')

        headers=['GERÇEK UNVAN','AD SOYAD','M','N','E','F']
        hr=top+3
        for j,h in enumerate(headers):
            cell=ws.cell(hr,left+j,h); cell.fill=PatternFill('solid',fgColor=blue)
            cell.font=Font(name='Arial',size=8,bold=True,color=white)
            cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)

        rr=hr+1
        for item in payload['rows']:
            vals=[item['Unvan'],item['Ad Soyad'],item['M'],item['N'],item['E'],item['F']]
            fill = ('F4CCCC' if item.get('Açıklama Türü')=='departure' else 'FFF2CC') if txt(item.get('Açıklama','')).strip() else shortage_blue if item['Durum']=='eksik' else surplus_green if item['Durum']=='fazla' else white
            for j,val in enumerate(vals):
                cell=ws.cell(rr,left+j,val); cell.fill=PatternFill('solid',fgColor=fill)
                cell.font=Font(name='Arial',size=8,color=dark)
                cell.alignment=Alignment(horizontal='left' if j<2 else 'center',vertical='center',wrap_text=True)
            # Fact_Norm Açıklama sütunundaki unvan-düzeyi (kişiye bağlı
            # olmayan) not, Unvan hücresinde yorum olarak görünür — DÜZELTME:
            # önceden notes_by_key hiç doldurulmuyordu ("Fact_Norm açıklaması
            # kullanılmaz" yorumuyla açıkça atlanıyordu); artık gerçekten
            # okunup buraya yazılıyor.
            norm_note_text=txt(item.get('Norm Açıklaması','')).strip()
            if norm_note_text:
                ws.cell(rr,left).comment=Comment(norm_note_text,'OMEHR Norm Açıklaması')
            # Fact_Mevcut H sütunundaki açıklama, Ad Soyad hücresinde not olarak görünür.
            note_text=txt(item.get('Açıklama Cümlesi') or item.get('Açıklama','')).strip()
            if note_text:
                ws.cell(rr,left+1).comment=Comment(note_text,'OMEHR Personel Açıklaması')
            rr+=1

        # Toplam satırı PDF ile aynı M/N/E/F mantığıyla.
        ws.merge_cells(start_row=rr,start_column=left,end_row=rr,end_column=left+1)
        ws.cell(rr,left,'TOPLAM').fill=PatternFill('solid',fgColor=gold)
        ws.cell(rr,left).font=Font(name='Arial',size=8,bold=True,color=white)
        ws.cell(rr,left).alignment=Alignment(horizontal='center')
        totals=[payload['current'],payload['norm'],payload['deficit'],payload['surplus']]
        for j,val in enumerate(totals, start=2):
            cell=ws.cell(rr,left+j,val); cell.font=Font(name='Arial',size=9,bold=True,color=dark)
            cell.alignment=Alignment(horizontal='center')
            cell.fill=PatternFill('solid',fgColor=shortage_blue if j==4 else surplus_green if j==5 else pale)

        # Kart dışı ve iç çizgiler.
        for r in range(top,rr+1):
            for col_i in range(left,end+1):
                cell=ws.cell(r,col_i)
                cell.border=Border(
                    left=medium if col_i==left else thin,
                    right=medium if col_i==end else thin,
                    top=medium if r==top else thin,
                    bottom=medium if r==rr else thin,
                )
        return rr-top+1

    for sheet_name, aliases in groups:
        ws=wb.create_sheet(sheet_name[:31]); ws.sheet_view.showGridLines=False
        ws.freeze_panes='A9'; ws.page_setup.orientation='landscape'; ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
        ws.page_setup.paperSize=ws.PAPERSIZE_A4
        ws.page_margins.left=0.2; ws.page_margins.right=0.2; ws.page_margins.top=0.3; ws.page_margins.bottom=0.3
        ws.merge_cells('A1:M2'); ws['A1']=f'OMEHR KUTUCUKLU YÖNETİCİ RAPORU — {sheet_name}'
        ws['A1'].font=Font(name='Arial',size=16,bold=True,color=white); ws['A1'].fill=PatternFill('solid',fgColor=navy)
        ws['A1'].alignment=Alignment(horizontal='center',vertical='center'); ws.row_dimensions[1].height=24

        subset=stx[stx['_yonetici'].isin(aliases)].copy().sort_values('Mağaza')
        payloads=[_store_payload(txt(r['Mağaza']),txt(r['Bölge Sorumlusu'])) for _,r in subset.iterrows()]

        # PDF'deki yönetici özet kutularının Excel karşılığı.
        summary = {
            'MAĞAZA': len(payloads),
            'M': sum(x['current'] for x in payloads),
            'N': sum(x['norm'] for x in payloads),
            'E': sum(x['deficit'] for x in payloads),
            'F': sum(x['surplus'] for x in payloads),
        }
        box_specs=[('A4:B6','MAĞAZA',gray),('C4:D6','M',pale),('E4:F6','N',pale),('H4:I6','E',shortage_blue),('J4:K6','F',surplus_green)]
        for rng,label,fill_color in box_specs:
            ws.merge_cells(rng)
            cell=ws[rng.split(':')[0]]
            cell.value=f'{label}\n{summary[label]}'
            cell.fill=PatternFill('solid',fgColor=fill_color)
            cell.font=Font(name='Arial',size=12,bold=True,color=dark)
            cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
            # Özet kutularına belirgin çerçeve.
            for row in ws[rng]:
                for c in row:
                    c.border=Border(left=medium,right=medium,top=medium,bottom=medium)
        ws.merge_cells('L4:M6'); ws['L4']=f'NET\n{summary["M"]-summary["N"]}'
        ws['L4'].fill=PatternFill('solid',fgColor=gold); ws['L4'].font=Font(name='Arial',size=12,bold=True,color=white)
        ws['L4'].alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        for row in ws['L4:M6']:
            for c in row:c.border=Border(left=medium,right=medium,top=medium,bottom=medium)

        if not payloads:
            ws.merge_cells('A9:M12'); ws['A9']='Bu yönetici için eşleşen mağaza kaydı bulunamadı.'
            ws['A9'].alignment=Alignment(horizontal='center',vertical='center')
            last_row=12
        else:
            top=9
            for i in range(0,len(payloads),2):
                left_card=payloads[i]
                right_card=payloads[i+1] if i+1<len(payloads) else None
                h1=_write_card(ws,top,1,left_card)
                h2=_write_card(ws,top,8,right_card) if right_card else 0
                top += max(h1,h2)+2
            last_row=top-1

        ws.sheet_properties.pageSetUpPr.fitToPage=True
        ws.print_title_rows='1:8'
        ws.print_area=f'A1:M{last_row}'
        ws.print_options.horizontalCentered=True
        ws.oddFooter.center.text='OMEHR — İnsan Kaynakları Direktörlüğü'
        ws.oddFooter.right.text='Sayfa &P / &N'

    wb.save(out)
    return out
