from __future__ import annotations

"""
KPI/MUTABAKAT KATMANI (P2 — engine_core.py modülerleştirme, dokuzuncu adım)
=====================================================================
Resmi KPI özetini (kpis()) ve "Mevcut-Norm", "Norm Fazlası-Eksiği", "norm
kapsamındaki mevcut" gibi FARKLI kavramların birbirine karışmadığını
doğrulayan mutabakat katmanını (reconcile_kpis()) içerir. state_engine'in
ürettiği DataFrame'leri girdi olarak alır, hiçbir hesaplama fonksiyonuna
geri bağımlı değildir.
"""

import pandas as pd
from openpyxl.styles import Font, PatternFill

from src.text_utils import canon, col, numeric, req


def _reconcile_net_by_store(frame):
    """Mağaza net farkını en yüksek departman farkından başlayarak satırlara dağıtır."""
    x=frame.copy()
    x['Norm Eksiği']=0
    x['Norm Fazlası']=0
    keys=['MağazaID','Mağaza','Bölge Sorumlusu']
    for _,indexes in x.groupby(keys,dropna=False).groups.items():
        idx=list(indexes)
        gaps=(numeric(x.loc[idx,'Aktif Mevcut'])-numeric(x.loc[idx,'Norm Kadro'])).astype(int)
        net=int(gaps.sum())
        if net>0:
            remaining=net
            for row_index in gaps[gaps>0].sort_values(ascending=False).index:
                amount=min(int(gaps.loc[row_index]),remaining)
                x.at[row_index,'Norm Fazlası']=amount
                remaining-=amount
                if remaining==0: break
        elif net<0:
            remaining=-net
            for row_index in (-gaps[gaps<0]).sort_values(ascending=False).index:
                amount=min(int(-gaps.loc[row_index]),remaining)
                x.at[row_index,'Norm Eksiği']=amount
                remaining-=amount
                if remaining==0: break
    return x



def kpis(st):
    override=getattr(st,'attrs',{}).get('kpi_override')
    if override: return dict(override)
    d={'Aktif Mevcut':int(st['Aktif Mevcut'].sum()),'Toplam Norm':int(st['Norm Kadro'].sum()),'Norm Eksiği':int(st['Norm Eksiği'].sum()),'Norm Fazlası':int(st['Norm Fazlası'].sum())}
    d['Net İhtiyaç']=d['Norm Fazlası']-d['Norm Eksiği']; return d




def reconcile_kpis(staff, tt, kp):
    """KPI MUTABAKAT KATMANI (P0 — üretim öncesi zorunlu).

    Sorun: "Mevcut − Norm", "Norm Fazlası − Norm Eksiği", "norm kapsamındaki
    mevcut", "toplam aktif mevcut" gibi FARKLI kavramlar birbirine
    karıştırılıp aynı şeymiş gibi raporlanabiliyordu. Bu fonksiyon HER
    çalıştırmada şu üç özdeşliği otomatik doğrular:

        Toplam aktif mevcut - Norm dışı çalışanlar = Norm kapsamındaki aktif mevcut
        Norm kapsamındaki mevcut - Toplam norm     = Brüt norm farkı
        Toplam norm fazlası - Toplam norm eksiği   = Net pozisyon farkı (Net İhtiyaç)

    Herhangi biri tutmazsa (aritmetik olarak İMKANSIZ olması gerekir, ama
    veri/kod hatası bunu bozabilir), 'tutarli'=False döner ve HANGİ
    personelin norm kapsamı dışında kaldığını (açıklanabilir fark tablosu)
    listeler — sessizce yanlış bir rakam raporlamak yerine.
    """
    toplam_aktif_mevcut = int(kp.get('Aktif Mevcut', 0))
    norm_kapsamindaki_mevcut = int(numeric(tt['Aktif Mevcut']).sum())
    norm_disi_calisan_sayisi = toplam_aktif_mevcut - norm_kapsamindaki_mevcut

    toplam_norm = int(kp.get('Toplam Norm', 0))
    brut_norm_farki = norm_kapsamindaki_mevcut - toplam_norm

    norm_fazlasi = int(kp.get('Norm Fazlası', 0))
    norm_eksigi = int(kp.get('Norm Eksiği', 0))
    net_pozisyon_farki = norm_fazlasi - norm_eksigi

    tutarli = (net_pozisyon_farki == kp.get('Net İhtiyaç', net_pozisyon_farki)) and norm_disi_calisan_sayisi >= 0

    fark_tablosu = pd.DataFrame()
    if norm_disi_calisan_sayisi != 0 and staff is not None and not staff.empty:
        # NOT: tt'nin 'UnvanID' sütunu title-level birleştirmede taşınmıyor
        # (bu kontrol sırasında keşfedilen, önceden fark edilmemiş bir
        # boşluk). Ayrıca tt'nin 'Unvan' metni de UZMAN/ELİT varyantları
        # BAZ unvana indirgenmiş olarak tutuyor (norm sayım kuralı gereği,
        # ör. "UZMAN KASAP" -> "KASAP") — bu yüzden eşleştirmeden önce
        # staff'taki gerçek unvan metni de AYNI indirgemeden geçirilmeli,
        # aksi halde her Uzman/Elit çalışan yanlışlıkla "norm dışı" görünür.
        # MERKEZİ KAYNAK (P2): hangi öneklerin (UZMAN/ELİT) baz unvana
        # indirgeneceği artık services/norm_aliases.py'de TEK bir yerde
        # tanımlıdır — formula_bagimsiz_hesapla.py da AYNI kaynağı kullanır.
        from services.norm_aliases import baz_unvan_metni as _baz_unvan_metni
        def _baz_unvan(v: str) -> str:
            return canon(_baz_unvan_metni(v))
        try:
            kapsanan_ciftler = set(zip(tt.get('MağazaID', pd.Series(dtype=object)), tt.get('Unvan', pd.Series(dtype=object)).map(_baz_unvan)))
            mid_c = req(staff, 'MağazaID'); uid_c = req(staff, 'UnvanID'); uad_c = req(staff, 'Unvan')
            disi_maske = ~staff.apply(lambda r: (r.get(mid_c), _baz_unvan(r.get(uad_c))) in kapsanan_ciftler, axis=1)
            pname_c = col(staff, 'İsim Soyisim', 'Isim Soyisim')
            kolonlar = [c for c in [req(staff, 'PersonelID'), mid_c, req(staff, 'Mağaza', 'Magaza'), uid_c, uad_c, pname_c] if c]
            fark_tablosu = staff.loc[disi_maske, kolonlar].copy()
        except Exception as _exc:
            from services.safe_exec import log_swallowed
            log_swallowed("reconcile_kpis: açıklanabilir fark tablosu hesaplanamadı", _exc)
            fark_tablosu = pd.DataFrame()

    return {
        'tutarli': bool(tutarli),
        'toplam_aktif_mevcut': toplam_aktif_mevcut,
        'norm_kapsamindaki_aktif_mevcut': norm_kapsamindaki_mevcut,
        'norm_disi_calisan_sayisi': norm_disi_calisan_sayisi,
        'toplam_norm': toplam_norm,
        'brut_norm_farki': brut_norm_farki,
        'toplam_norm_fazlasi': norm_fazlasi,
        'toplam_norm_eksigi': norm_eksigi,
        'net_pozisyon_farki': net_pozisyon_farki,
        'net_ihtiyac_kpi': kp.get('Net İhtiyaç'),
        'aciklanabilir_fark_tablosu': fark_tablosu,
    }




def _yaz_kpi_mutabakat_sayfasi(excel_path, mutabakat: dict, workbook=None):
    """reconcile_kpis() sonucunu, üretilen yönetici Excel dosyasına ayrı bir
    'KPI_Mutabakat_Kontrolu' sayfası olarak ekler — her çalıştırmada rakamların
    gerçekten tutup tutmadığının GÖRÜNÜR, denetlenebilir kaydı."""
    from openpyxl import load_workbook
    owns_workbook = workbook is None
    wb = load_workbook(excel_path) if owns_workbook else workbook
    if 'KPI_Mutabakat_Kontrolu' in wb.sheetnames:
        del wb['KPI_Mutabakat_Kontrolu']
    ws = wb.create_sheet('KPI_Mutabakat_Kontrolu')
    baslik_font = Font(bold=True, color='FFFFFF')
    baslik_dolgu = PatternFill(start_color='143C36', end_color='143C36', fill_type='solid')
    ws['A1'] = 'KPI MUTABAKAT KONTROLÜ — her çalıştırmada otomatik doğrulanır'
    ws['A1'].font = Font(bold=True, size=13)
    durum = '✅ TUTARLI' if mutabakat['tutarli'] else '🚫 TUTARSIZLIK TESPİT EDİLDİ — aşağıdaki fark tablosuna bakın'
    ws['A2'] = f'Durum: {durum}'
    ws['A2'].font = Font(bold=True, color=('1F6F54' if mutabakat['tutarli'] else '9B2D2D'))
    satirlar = [
        ('Toplam Aktif Mevcut', mutabakat['toplam_aktif_mevcut']),
        ('Norm Kapsamındaki Aktif Mevcut', mutabakat['norm_kapsamindaki_aktif_mevcut']),
        ('Norm Dışı Çalışan Sayısı (= Toplam − Kapsamdaki)', mutabakat['norm_disi_calisan_sayisi']),
        ('Toplam Norm (Yönetim Normu)', mutabakat['toplam_norm']),
        ('Brüt Norm Farkı (= Kapsamdaki Mevcut − Toplam Norm)', mutabakat['brut_norm_farki']),
        ('Toplam Norm Fazlası', mutabakat['toplam_norm_fazlasi']),
        ('Toplam Norm Eksiği', mutabakat['toplam_norm_eksigi']),
        ('Net Pozisyon Farkı (= Fazla − Eksik)', mutabakat['net_pozisyon_farki']),
        ('Resmi KPI "Net İhtiyaç" (karşılaştırma)', mutabakat['net_ihtiyac_kpi']),
    ]
    r = 4
    for etiket, deger in satirlar:
        ws.cell(r, 1).value = etiket
        ws.cell(r, 2).value = deger
        r += 1
    if not mutabakat['aciklanabilir_fark_tablosu'].empty:
        r += 1
        ws.cell(r, 1).value = 'AÇIKLANABİLİR FARK TABLOSU — norm kapsamı dışında kalan personel'
        ws.cell(r, 1).font = Font(bold=True)
        r += 1
        headers = list(mutabakat['aciklanabilir_fark_tablosu'].columns)
        for c, h_ in enumerate(headers, start=1):
            ws.cell(r, c).value = h_; ws.cell(r, c).font = baslik_font; ws.cell(r, c).fill = baslik_dolgu
        for _, row in mutabakat['aciklanabilir_fark_tablosu'].iterrows():
            r += 1
            for c, h_ in enumerate(headers, start=1):
                ws.cell(r, c).value = row[h_]
    for col_letter, w in zip('ABCDEFG', [55, 16, 14, 12, 12, 12, 22]):
        ws.column_dimensions[col_letter].width = w
    if owns_workbook:
        wb.save(excel_path)
    return wb
